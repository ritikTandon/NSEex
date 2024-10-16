import os
import shutil
import pandas as pd
from zipfile import ZipFile
import calendar

import openpyxl as xl
from openpyxl.styles import PatternFill
import datetime
from xls2xlsx import XLS2XLSX
from date_variables import date, mnth, yr

fo_share_list = ["ADANI", "APORT", "APOLLO", "AURO", "AXIS", "BAJAJ", "BARODA", "AIRTEL", "BHEL", "BN", "CANBK", "COALIND",
                 "DLF", "DRREDDY", "EICHER", "HCL", "HDFC", "HIND", "HINDUNLVR", "ICICI", "INDUSIND", "JIND", "NIFTY", "REL",
                 "SBIN", "TCHEM", "TCON", "TM", "TS", "TCS", "TITAN", "ULTRA", "VEDL"]
# fo_share_list = ["ADANI"]

fo_daily_aggregate_list = ["APORT", "AURO", "BN", "CANBK", "DLF", "HIND", "ICICI", "JIND", "NIFTY", "REL", "SBIN",
                           "TCON", "TM", "TS", "TCS", "TITAN"]


# potential fix to check if we are or last thu of month, we start looking for next month's expiry
def checkExpiry(day, month_abbr, year):
    # Mapping of month abbreviations to month numbers
    month_map = {
        "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
        "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12
    }

    # Convert the month abbreviation to a month number
    month = month_map.get(month_abbr.upper())

    if month is None:
        raise ValueError("Invalid month abbreviation")

    # Get the number of days in the month
    _, last_day_of_month = calendar.monthrange(year, month)

    # Find the last Thursday of the month
    last_date_of_month = datetime.date(year, month, last_day_of_month)
    while last_date_of_month.weekday() != 3:  # 3 corresponds to Thursday
        last_date_of_month -= datetime.timedelta(days=1)

    # Check if the given day is greater than or equal to the last Thursday
    given_date = datetime.date(year, month, day)
    if given_date >= last_date_of_month:
        # Move to the next month
        if month == 12:
            next_month = "JAN"
        else:
            next_month = list(month_map.keys())[list(month_map.values()).index(month) + 1]
        return next_month
    else:
        return month_abbr


# date variables converted to int to work with checkExpiry function
d = int(date[:2])
m = mnth
y = int(yr)


foHL_wb = xl.load_workbook(r'C:\Users\admin\PycharmProjects\daily data\fo high low.xlsx')
foHL_sheet = foHL_wb['Sheet1']
foHL_row = 2


# converting xls to xlsx for fo1 sheet
x2x = XLS2XLSX(r'E:\Daily Data work\fo1.xls')

wb = x2x.to_xlsx()
wb.save(r'E:\Daily Data work\fo1.xlsx')

fo1_wb = xl.load_workbook(r'E:\Daily Data work\fo1.xlsx')
fo1_sheet = fo1_wb['fo1-Sheet1']
fo1_row = 2

fl_9_25 = 0.392361111       # 9:25 time value in general format

# copying hourlys (.xls) as backup
# path to source directory
src_dir = rf"E:\Daily Data work\hourlys 1 minute FO\{yr}\{mnth}\{date}"

# path to destination directory
dest_dir = rf"C:\Users\admin\PycharmProjects\daily data\Daily Backup hourlys\1 min fo"

# getting all the files in the source directory
src_files = os.listdir(src_dir)
for file_name in src_files:
    full_file_name = os.path.join(src_dir, file_name)
    if os.path.isfile(full_file_name):
        shutil.copy(full_file_name, dest_dir)

print("Files copied as backup!")

for share in fo_share_list:
    path = rf"E:\Daily Data work\hourlys 1 minute FO\{yr}\{mnth}\{date}\{share}.xlsx"
    xls_path = rf"E:\Daily Data work\hourlys 1 minute FO\{yr}\{mnth}\{date}\{share}.xls"

    x2x = XLS2XLSX(xls_path)

    wb = x2x.to_xlsx()
    sheet = wb[f"{share}-Sheet1"]

    start_row = 2
    time_cell = sheet.cell(start_row, 7)

    while time_cell.value < fl_9_25:
        start_row += 1
        time_cell = sheet.cell(start_row, 7)

    print(f"starting row is {start_row}")

    start_row_2 = start_row

    # loop for changing the time cells format (have to close and reopen otherwise it doesn't change format)
    while time_cell.value is not None:
        time_cell = sheet.cell(start_row, 7)
        time_cell.number_format = 'hh:mm AM/PM'

        start_row += 1

    wb.save(path)

    # reloading the worksheet
    wb = xl.load_workbook(path)
    sheet = wb[f"{share}-Sheet1"]

    start_row = start_row_2

    time_cell = sheet.cell(start_row, 7)
    cur_time = time_cell.value
    end_time = datetime.time(15, 30, 0)

    # 9:25 close value
    cl_9_25 = sheet.cell(start_row, 3).value
    sheet.cell(start_row, 3).fill = PatternFill("solid", 'FFFF00')

    # reloading wb otherwise pattern fill doesn't work
    wb.save(path)
    wb = xl.load_workbook(path)
    sheet = wb[f"{share}-Sheet1"]

    HIGH = 0
    LOW = 9999999

    # HIGH and LOW value finding loop
    while cur_time is not None and cur_time <= end_time:
        time_cell = sheet.cell(start_row, 7)
        high_cell = sheet.cell(start_row, 4)
        low_cell = sheet.cell(start_row, 5)

        cur_time = time_cell.value

        if high_cell.value is not None and high_cell.value > HIGH:
            HIGH = high_cell.value

        if low_cell.value is not None and low_cell.value < LOW and low_cell.value != 0:
            LOW = low_cell.value

        start_row += 1

    if share in fo_daily_aggregate_list:
        # high
        foHL_sheet.cell(foHL_row, 2).value = HIGH

        # low
        foHL_sheet.cell(foHL_row, 3).value = LOW

        # close
        foHL_sheet.cell(foHL_row, 4).value = fo1_sheet.cell(fo1_row, 2).value

        # vol
        volume = fo1_sheet.cell(fo1_row, 5).value
        volume //= 100000   # truncating volume to display in lakhs

        foHL_sheet.cell(foHL_row, 6).value = volume

        # 9:25 close
        foHL_sheet.cell(foHL_row, 7).value = cl_9_25

        foHL_row += 1
        fo1_row += 1

    HIGH = 0
    LOW = 9999999

    # 30 MIN FORMATTING IN 1 MIN SHEETS
    sheet.cell(1, 14).value = "HIGH"
    sheet.cell(1, 15).value = "LOW"
    sheet.cell(1, 16).value = "CLOSE"

    start_row = start_row_2     # actual start row

    time_cell = sheet.cell(start_row, 7)
    cur_time = time_cell.value

    count = 0

    while cur_time is not None and cur_time <= end_time:
        high_cell = sheet.cell(start_row, 4)
        low_cell = sheet.cell(start_row, 5)

        # print(cur_time)

        if high_cell.value is not None and high_cell.value > HIGH:
            HIGH = high_cell.value

        if low_cell.value is not None and low_cell.value < LOW and low_cell.value != 0:
            LOW = low_cell.value

        # resetting after 30 mins
        if count == 30:
            sheet.cell(start_row, 14).value = HIGH
            sheet.cell(start_row, 15).value = LOW

            # if 30 min close is empty or 0
            if sheet.cell(start_row, 3).value == 0 or sheet.cell(start_row, 3).value is None:
                temp_row = start_row

                while sheet.cell(temp_row, 3).value == 0 or sheet.cell(temp_row, 3).value is None:
                    temp_row -= 1

                sheet.cell(start_row, 16).value = sheet.cell(temp_row, 3).value  # close

            else:
                sheet.cell(start_row, 16).value = sheet.cell(start_row, 3).value  # close

            count = 1
            HIGH = 0
            LOW = 9999999
            start_row += 1
            continue

        start_row += 1
        count += 1

        time_cell = sheet.cell(start_row, 7)
        cur_time = time_cell.value

    # last any left aggregate (< 30 mins)
    sheet.cell(start_row-1, 14).value = HIGH
    sheet.cell(start_row-1, 15).value = LOW
    sheet.cell(start_row-1, 16).value = sheet.cell(start_row-1, 3).value  # close

    sheet.freeze_panes = sheet["A2"]

    wb.save(path)
    os.remove(xls_path)
    # wb.save(rf'E:\Daily Data work\hourlys 1 minute FO\{yr}\{mnth}\{date}\{share}1.xlsx')
    print(f"{share} done")


# MD file (new (changed on 8 jul 2024)) BhavCopy_NSE_FO_0_0_0_20240708_F_0000.csv
# dict containing share names (NSE) and their respective rows in 'fo high low.xlsx'
share_list = {"BANKNIFTY": 4, "NIFTY": 10, "ADANIPORTS": 2, "AUROPHARMA": 3, "CANBK": 5, "DLF": 6, "HINDALCO": 7,
              "ICICIBANK": 8, "JINDALSTEL": 9, "RELIANCE": 11, "SBIN": 12, "TATACONSUM": 13, "TATAMOTORS": 14,
              "TATASTEEL": 15, "TCS": 16, "TITAN": 17}

md_path_zipped = rf"E:\chrome downloads\BhavCopy_NSE_FO_0_0_0_{yr}{date[3:5]}{date[:2]}_F_0000.csv.zip"     # .zip file path of downloaded cash bhavcopy
md_path = rf"E:\chrome downloads"

# take data from col 22 and col 14 has month as cur month in caps and col 13 is null

# extracting .zip file
with ZipFile(md_path_zipped, 'r') as zObject:
    zObject.extractall(path=md_path)

md_file_path = rf"E:\Daily Data work\MD files\{yr}\{mnth}\fo{date[:2]}{mnth}20{date[6:]}bhav.xlsx"

df = pd.read_csv(md_path_zipped[:-4])       # removing the .zip extension after unzipping
df = df.drop(df.columns[-9:], axis=1)
df = df.drop(columns=['BizDt', 'Src', 'FinInstrmTp', 'FinInstrmId', 'ISIN', 'SctySrs', 'OpnIntrst', 'ChngInOpnIntrst'])
df2 = df.copy()     # this one will actually be saved as md file
df = df[df['OptnTp'].isnull()]

for share in share_list:
    # match name with symbol and check if FinInstrmNm contains month name (it's the NIFTY24AUG24 full share name)
    ltp = df.loc[(df['TckrSymb'] == share) & df['FinInstrmNm'].str.contains(checkExpiry(d, m, y)), 'SttlmPric'].iloc[0]
    foHL_sheet.cell(share_list[share], 5).value = ltp

df2.to_excel(md_file_path, index=False)
foHL_wb.save(r'C:\Users\admin\PycharmProjects\daily data\fo high low.xlsx')
