import shutil
import pandas as pd
import glob
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pyautogui as pg
import pyperclip as pc
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension

from date_variables import date, mnth, yr
from time import sleep
import time
import shutil
import openpyxl as xl
import os
import requests
import datetime

from xls2xlsx import XLS2XLSX

from date_variables import date, mnth, yr
from time import sleep
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException

begin = time.time()

# for close filling
options = Options()
# options.add_argument('--headless=new')
options.add_argument("--disable-blink-features=AutomationControlled")

# Exclude the collection of enable-automation switches
options.add_experimental_option("excludeSwitches", ["enable-automation"])

# maximizing window
options.add_argument("--start-maximized")

# Turn-off userAutomationExtension
options.add_experimental_option("useAutomationExtension", False)

# share_add_list = ['AARTIIND', 'ABB', 'ABCAPITAL', 'ABFRL', 'ADANIENT', 'ADANIPORTS', 'ALKEM', 'AMBUJACEM',
#                      'APOLLOHOSP', 'APOLLOTYRE', 'ASHOKLEY', 'ASTRAL', 'ATUL', 'AUBANK', 'AUROPHARMA', 'BAJAJFINSV',
#                      'BAJFINANCE', 'BALKRISIND', 'BALRAMCHIN', 'BANDHANBNK', 'BANKBARODA', 'BATAINDIA', 'BEL',
#                      'BHARATFORG', 'BIOCON', 'BRITANNIA', 'BSOFT', 'CANBK', 'CANFINHOME', 'CHAMBLFERT', 'CHOLAFIN',
#                      'CIPLA', 'COFORGE', 'CONCOR', 'COROMANDEL', 'CROMPTON', 'CUMMINSIND', 'DABUR', 'DALBHARAT',
#                      'DEEPAKFERT', 'DEEPAKNTR', 'DELTACORP', 'DIVISLAB', 'DIXON', 'DLF', 'DRREDDY', 'ESCORTS',
#                      'EXIDEIND', 'GLENMARK', 'GLS', 'GNFC', 'GODREJCP', 'GODREJPROP', 'GRANULES', 'GRASIM', 'GUJGASLTD',
#                      'HAL', 'HAVELLS', 'HCLTECH', 'HDFCAMC', 'HDFCLIFE', 'HINDALCO', 'HINDCOPPER', 'ICICIGI',
#                      'ICICIPRULI', 'IEX', 'IGL', 'INDHOTEL', 'INDIACEM', 'INDIAMART', 'INDIGO', 'INDUSINDBK',
#                      'INDUSTOWER', 'INTELLECT', 'IPCALAB', 'JINDALSTEL', 'JKCEMENT', 'JSWSTEEL', 'JUBLFOOD',
#                      'KOTAKBANK', 'LALPATHLAB', 'LAURUSLABS', 'LICHSGFIN', 'LTIM', 'LTTS', 'LUPIN', 'M&MFIN',
#                      'MANAPPURAM', 'MARICO', 'MCDOWELL-N', 'MCX', 'METROPOLIS', 'MFSL', 'MGL', 'MPHASIS', 'MUTHOOTFIN',
#                      'NAM-INDIA', 'NAUKRI', 'NAVINFLUOR', 'NMDC', 'NTPC', 'OBEROIRLTY', 'PEL', 'PERSISTENT', 'PETRONET',
#                      'PIDILITIND', 'POLYCAB', 'POWERGRID', 'RAIN', 'RAMCOCEM', 'RBLBANK', 'RECLTD', 'SBICARD',
#                      'SBILIFE', 'SIEMENS', 'SRF', 'STAR', 'SUNPHARMA', 'SYNGENE', 'TATACOMM', 'TATAMOTORS', 'TECHM',
#                      'TORNTPHARM', 'TORNTPOWER', 'TRENT', 'TVSMOTOR', 'UBL', 'ULTRACEMCO', 'UPL', 'VEDL', 'VOLTAS',
#                      'ZEEL', 'ZYDUSLIFE']

share_add_list = ['M%26MFIN',
                     'MANAPPURAM', 'MARICO', 'MCDOWELL-N', 'MCX', 'METROPOLIS', 'MFSL', 'MGL', 'MPHASIS', 'MUTHOOTFIN',
                     'NAM-INDIA', 'NAUKRI', 'NAVINFLUOR', 'NMDC', 'NTPC', 'OBEROIRLTY', 'PEL', 'PERSISTENT', 'PETRONET',
                     'PIDILITIND', 'POLYCAB', 'POWERGRID', 'RAIN', 'RAMCOCEM', 'RBLBANK', 'RECLTD', 'SBICARD',
                     'SBILIFE', 'SIEMENS', 'SRF', 'STAR', 'SUNPHARMA', 'SYNGENE', 'TATACOMM', 'TATAMOTORS', 'TECHM',
                     'TORNTPHARM', 'TORNTPOWER', 'TRENT', 'TVSMOTOR', 'UBL', 'ULTRACEMCO', 'UPL', 'VEDL', 'VOLTAS',
                     'ZEEL', 'ZYDUSLIFE']

double_name = ['CHOLAFIN', 'DEEPAKFERT', 'GNFC']

# share_add_list = ['HINDALCO', 'NTPC']
# share_add_list = ['HINDALCO']


date_list = ['01-Jan-20', '02-Jan-20', '03-Jan-20', '06-Jan-20', '07-Jan-20', '08-Jan-20', '09-Jan-20', '10-Jan-20',
             '13-Jan-20', '14-Jan-20', '15-Jan-20', '16-Jan-20', '17-Jan-20', '20-Jan-20', '21-Jan-20', '22-Jan-20',
             '23-Jan-20', '24-Jan-20', '27-Jan-20', '28-Jan-20', '29-Jan-20', '30-Jan-20', '31-Jan-20', '01-Feb-20',
             '03-Feb-20', '04-Feb-20', '05-Feb-20', '06-Feb-20', '07-Feb-20', '10-Feb-20', '11-Feb-20', '12-Feb-20',
             '13-Feb-20', '14-Feb-20', '17-Feb-20', '18-Feb-20', '19-Feb-20', '20-Feb-20', '21-Feb-20', '24-Feb-20',
             '25-Feb-20', '26-Feb-20', '27-Feb-20', '28-Feb-20', '02-Mar-20', '03-Mar-20', '04-Mar-20', '05-Mar-20',
             '06-Mar-20', '09-Mar-20', '10-Mar-20', '11-Mar-20', '12-Mar-20', '13-Mar-20', '16-Mar-20', '17-Mar-20',
             '18-Mar-20', '19-Mar-20', '20-Mar-20', '23-Mar-20', '24-Mar-20', '25-Mar-20', '26-Mar-20', '27-Mar-20',
             '30-Mar-20', '31-Mar-20', '01-Apr-20', '02-Apr-20', '03-Apr-20', '06-Apr-20', '07-Apr-20', '08-Apr-20',
             '09-Apr-20', '10-Apr-20', '13-Apr-20', '14-Apr-20', '15-Apr-20', '16-Apr-20', '17-Apr-20', '20-Apr-20',
             '21-Apr-20', '22-Apr-20', '23-Apr-20', '24-Apr-20', '27-Apr-20', '28-Apr-20', '29-Apr-20', '30-Apr-20',
             '01-May-20', '04-May-20', '05-May-20', '06-May-20', '07-May-20', '08-May-20', '11-May-20', '12-May-20',
             '13-May-20', '14-May-20', '15-May-20', '18-May-20', '19-May-20', '20-May-20', '21-May-20', '22-May-20',
             '25-May-20', '26-May-20', '27-May-20', '28-May-20', '29-May-20', '01-Jun-20', '02-Jun-20', '03-Jun-20',
             '04-Jun-20', '05-Jun-20', '08-Jun-20', '09-Jun-20', '10-Jun-20', '11-Jun-20', '12-Jun-20', '15-Jun-20',
             '16-Jun-20', '17-Jun-20', '18-Jun-20', '19-Jun-20', '22-Jun-20', '23-Jun-20', '24-Jun-20', '25-Jun-20',
             '26-Jun-20', '29-Jun-20', '30-Jun-20', '01-Jul-20', '02-Jul-20', '03-Jul-20', '06-Jul-20', '07-Jul-20',
             '08-Jul-20', '09-Jul-20', '10-Jul-20', '13-Jul-20', '14-Jul-20', '15-Jul-20', '16-Jul-20', '17-Jul-20',
             '20-Jul-20', '21-Jul-20', '22-Jul-20', '23-Jul-20', '24-Jul-20', '27-Jul-20', '28-Jul-20', '29-Jul-20',
             '30-Jul-20', '31-Jul-20', '03-Aug-20', '04-Aug-20', '05-Aug-20', '06-Aug-20', '07-Aug-20', '10-Aug-20',
             '11-Aug-20', '12-Aug-20', '13-Aug-20', '14-Aug-20', '17-Aug-20', '18-Aug-20', '19-Aug-20', '20-Aug-20',
             '21-Aug-20', '24-Aug-20', '25-Aug-20', '26-Aug-20', '27-Aug-20', '28-Aug-20', '31-Aug-20', '01-Sep-20',
             '02-Sep-20', '03-Sep-20', '04-Sep-20', '07-Sep-20', '08-Sep-20', '09-Sep-20', '10-Sep-20', '11-Sep-20',
             '14-Sep-20', '15-Sep-20', '16-Sep-20', '17-Sep-20', '18-Sep-20', '21-Sep-20', '22-Sep-20', '23-Sep-20',
             '24-Sep-20', '25-Sep-20', '28-Sep-20', '29-Sep-20', '30-Sep-20', '01-Oct-20', '02-Oct-20', '05-Oct-20',
             '06-Oct-20', '07-Oct-20', '08-Oct-20', '09-Oct-20', '12-Oct-20', '13-Oct-20', '14-Oct-20', '15-Oct-20',
             '16-Oct-20', '19-Oct-20', '20-Oct-20', '21-Oct-20', '22-Oct-20', '23-Oct-20', '26-Oct-20', '27-Oct-20',
             '28-Oct-20', '29-Oct-20', '30-Oct-20', '02-Nov-20', '03-Nov-20', '04-Nov-20', '05-Nov-20', '06-Nov-20',
             '09-Nov-20', '10-Nov-20', '11-Nov-20', '12-Nov-20', '13-Nov-20', '16-Nov-20', '17-Nov-20', '18-Nov-20',
             '19-Nov-20', '20-Nov-20', '23-Nov-20', '24-Nov-20', '25-Nov-20', '26-Nov-20', '27-Nov-20', '30-Nov-20',
             '01-Dec-20', '02-Dec-20', '03-Dec-20', '04-Dec-20', '07-Dec-20', '08-Dec-20', '09-Dec-20', '10-Dec-20',
             '11-Dec-20', '14-Dec-20', '15-Dec-20', '16-Dec-20', '17-Dec-20', '18-Dec-20', '21-Dec-20', '22-Dec-20',
             '23-Dec-20', '24-Dec-20', '25-Dec-20', '28-Dec-20', '29-Dec-20', '30-Dec-20', '31-Dec-20', '01-Jan-21',
             '04-Jan-21', '05-Jan-21', '06-Jan-21', '07-Jan-21', '08-Jan-21', '11-Jan-21', '12-Jan-21', '13-Jan-21',
             '14-Jan-21', '15-Jan-21', '18-Jan-21', '19-Jan-21', '20-Jan-21', '21-Jan-21', '22-Jan-21', '25-Jan-21',
             '26-Jan-21', '27-Jan-21', '28-Jan-21', '29-Jan-21', '01-Feb-21', '02-Feb-21', '03-Feb-21', '04-Feb-21',
             '05-Feb-21', '08-Feb-21', '09-Feb-21', '10-Feb-21', '11-Feb-21', '12-Feb-21', '15-Feb-21', '16-Feb-21',
             '17-Feb-21', '18-Feb-21', '19-Feb-21', '22-Feb-21', '23-Feb-21', '24-Feb-21', '25-Feb-21', '26-Feb-21',
             '01-Mar-21', '02-Mar-21', '03-Mar-21', '04-Mar-21', '05-Mar-21', '08-Mar-21', '09-Mar-21', '10-Mar-21',
             '11-Mar-21', '12-Mar-21', '15-Mar-21', '16-Mar-21', '17-Mar-21', '18-Mar-21', '19-Mar-21', '22-Mar-21',
             '23-Mar-21', '24-Mar-21', '25-Mar-21', '26-Mar-21', '29-Mar-21', '30-Mar-21', '31-Mar-21', '01-Apr-21',
             '02-Apr-21', '05-Apr-21', '06-Apr-21', '07-Apr-21', '08-Apr-21', '09-Apr-21', '12-Apr-21', '13-Apr-21',
             '14-Apr-21', '15-Apr-21', '16-Apr-21', '19-Apr-21', '20-Apr-21', '21-Apr-21', '22-Apr-21', '23-Apr-21',
             '26-Apr-21', '27-Apr-21', '28-Apr-21', '29-Apr-21', '30-Apr-21', '03-May-21', '04-May-21', '05-May-21',
             '06-May-21', '07-May-21', '10-May-21', '11-May-21', '12-May-21', '13-May-21', '14-May-21', '17-May-21',
             '18-May-21', '19-May-21', '20-May-21', '21-May-21', '24-May-21', '25-May-21', '26-May-21', '27-May-21',
             '28-May-21', '31-May-21', '01-Jun-21', '02-Jun-21', '03-Jun-21', '04-Jun-21', '07-Jun-21', '08-Jun-21',
             '09-Jun-21', '10-Jun-21', '11-Jun-21', '14-Jun-21', '15-Jun-21', '16-Jun-21', '17-Jun-21', '18-Jun-21',
             '21-Jun-21', '22-Jun-21', '23-Jun-21', '24-Jun-21', '25-Jun-21', '28-Jun-21', '29-Jun-21', '30-Jun-21',
             '01-Jul-21', '02-Jul-21', '05-Jul-21', '06-Jul-21', '07-Jul-21', '08-Jul-21', '09-Jul-21', '12-Jul-21',
             '13-Jul-21', '14-Jul-21', '15-Jul-21', '16-Jul-21', '19-Jul-21', '20-Jul-21', '21-Jul-21', '22-Jul-21',
             '23-Jul-21', '26-Jul-21', '27-Jul-21', '28-Jul-21', '29-Jul-21', '30-Jul-21', '02-Aug-21', '03-Aug-21',
             '04-Aug-21', '05-Aug-21', '06-Aug-21', '09-Aug-21', '10-Aug-21', '11-Aug-21', '12-Aug-21', '13-Aug-21',
             '16-Aug-21', '17-Aug-21', '18-Aug-21', '19-Aug-21', '20-Aug-21', '23-Aug-21', '24-Aug-21', '25-Aug-21',
             '26-Aug-21', '27-Aug-21', '30-Aug-21', '31-Aug-21', '01-Sep-21', '02-Sep-21', '03-Sep-21', '06-Sep-21',
             '07-Sep-21', '08-Sep-21', '09-Sep-21', '10-Sep-21', '13-Sep-21', '14-Sep-21', '15-Sep-21', '16-Sep-21',
             '17-Sep-21', '20-Sep-21', '21-Sep-21', '22-Sep-21', '23-Sep-21', '24-Sep-21', '27-Sep-21', '28-Sep-21',
             '29-Sep-21', '30-Sep-21', '01-Oct-21', '04-Oct-21', '05-Oct-21', '06-Oct-21', '07-Oct-21', '08-Oct-21',
             '11-Oct-21', '12-Oct-21', '13-Oct-21', '14-Oct-21', '15-Oct-21', '18-Oct-21', '19-Oct-21', '20-Oct-21',
             '21-Oct-21', '22-Oct-21', '25-Oct-21', '26-Oct-21', '27-Oct-21', '28-Oct-21', '29-Oct-21', '01-Nov-21',
             '02-Nov-21', '03-Nov-21', '04-Nov-21', '05-Nov-21', '08-Nov-21', '09-Nov-21', '10-Nov-21', '11-Nov-21',
             '12-Nov-21', '15-Nov-21', '16-Nov-21', '17-Nov-21', '18-Nov-21', '19-Nov-21', '22-Nov-21', '23-Nov-21',
             '24-Nov-21', '25-Nov-21', '26-Nov-21', '29-Nov-21', '30-Nov-21', '01-Dec-21', '02-Dec-21', '03-Dec-21',
             '06-Dec-21', '07-Dec-21', '08-Dec-21', '09-Dec-21', '10-Dec-21', '13-Dec-21', '14-Dec-21', '15-Dec-21',
             '16-Dec-21', '17-Dec-21', '20-Dec-21', '21-Dec-21', '22-Dec-21', '23-Dec-21', '24-Dec-21', '27-Dec-21',
             '28-Dec-21', '29-Dec-21', '30-Dec-21', '31-Dec-21', '03-Jan-22', '04-Jan-22', '05-Jan-22', '06-Jan-22',
             '07-Jan-22', '10-Jan-22', '11-Jan-22', '12-Jan-22', '13-Jan-22', '14-Jan-22', '17-Jan-22', '18-Jan-22',
             '19-Jan-22', '20-Jan-22', '21-Jan-22', '24-Jan-22', '25-Jan-22', '26-Jan-22', '27-Jan-22', '28-Jan-22',
             '31-Jan-22', '01-Feb-22', '02-Feb-22', '03-Feb-22', '04-Feb-22', '07-Feb-22', '08-Feb-22', '09-Feb-22',
             '10-Feb-22', '11-Feb-22', '14-Feb-22', '15-Feb-22', '16-Feb-22', '17-Feb-22', '18-Feb-22', '21-Feb-22',
             '22-Feb-22', '23-Feb-22', '24-Feb-22', '25-Feb-22', '28-Feb-22', '01-Mar-22', '02-Mar-22', '03-Mar-22',
             '04-Mar-22', '07-Mar-22', '08-Mar-22', '09-Mar-22', '10-Mar-22', '11-Mar-22', '14-Mar-22', '15-Mar-22',
             '16-Mar-22', '17-Mar-22', '18-Mar-22', '21-Mar-22', '22-Mar-22', '23-Mar-22', '24-Mar-22', '25-Mar-22',
             '28-Mar-22', '29-Mar-22', '30-Mar-22', '31-Mar-22', '01-Apr-22', '04-Apr-22', '05-Apr-22', '06-Apr-22',
             '07-Apr-22', '08-Apr-22', '11-Apr-22', '12-Apr-22', '13-Apr-22', '14-Apr-22', '15-Apr-22', '18-Apr-22',
             '19-Apr-22', '20-Apr-22', '21-Apr-22', '22-Apr-22', '25-Apr-22', '26-Apr-22', '27-Apr-22', '28-Apr-22',
             '29-Apr-22', '02-May-22', '03-May-22', '04-May-22', '05-May-22', '06-May-22', '09-May-22', '10-May-22',
             '11-May-22', '12-May-22', '13-May-22', '16-May-22', '17-May-22', '18-May-22', '19-May-22', '20-May-22',
             '23-May-22', '24-May-22', '25-May-22', '26-May-22', '27-May-22', '30-May-22', '31-May-22', '01-Jun-22',
             '02-Jun-22', '03-Jun-22', '06-Jun-22', '07-Jun-22', '08-Jun-22', '09-Jun-22', '10-Jun-22', '13-Jun-22',
             '14-Jun-22', '15-Jun-22', '16-Jun-22', '17-Jun-22', '20-Jun-22', '21-Jun-22', '22-Jun-22', '23-Jun-22',
             '24-Jun-22', '27-Jun-22', '28-Jun-22', '29-Jun-22', '30-Jun-22', '01-Jul-22', '04-Jul-22', '05-Jul-22',
             '06-Jul-22', '07-Jul-22', '08-Jul-22', '11-Jul-22', '12-Jul-22', '13-Jul-22', '14-Jul-22', '15-Jul-22',
             '18-Jul-22', '19-Jul-22', '20-Jul-22', '21-Jul-22', '22-Jul-22', '25-Jul-22', '26-Jul-22', '27-Jul-22',
             '28-Jul-22', '29-Jul-22', '01-Aug-22', '02-Aug-22', '03-Aug-22', '04-Aug-22', '05-Aug-22', '08-Aug-22',
             '09-Aug-22', '10-Aug-22', '11-Aug-22', '12-Aug-22', '15-Aug-22', '16-Aug-22', '17-Aug-22', '18-Aug-22',
             '19-Aug-22', '22-Aug-22', '23-Aug-22', '24-Aug-22', '25-Aug-22', '26-Aug-22', '29-Aug-22', '30-Aug-22',
             '31-Aug-22', '01-Sep-22', '02-Sep-22', '05-Sep-22', '06-Sep-22', '07-Sep-22', '08-Sep-22', '09-Sep-22',
             '12-Sep-22', '13-Sep-22', '14-Sep-22', '15-Sep-22', '16-Sep-22', '19-Sep-22', '20-Sep-22', '21-Sep-22',
             '22-Sep-22', '23-Sep-22', '26-Sep-22', '27-Sep-22', '28-Sep-22', '29-Sep-22', '30-Sep-22', '03-Oct-22',
             '04-Oct-22', '05-Oct-22', '06-Oct-22', '07-Oct-22', '10-Oct-22', '11-Oct-22', '12-Oct-22', '13-Oct-22',
             '14-Oct-22', '17-Oct-22', '18-Oct-22', '19-Oct-22', '20-Oct-22', '21-Oct-22', '24-Oct-22', '25-Oct-22',
             '26-Oct-22', '27-Oct-22', '28-Oct-22', '31-Oct-22', '01-Nov-22', '02-Nov-22', '03-Nov-22', '04-Nov-22',
             '07-Nov-22', '08-Nov-22', '09-Nov-22', '10-Nov-22', '11-Nov-22', '14-Nov-22', '15-Nov-22', '16-Nov-22',
             '17-Nov-22', '18-Nov-22', '21-Nov-22', '22-Nov-22', '23-Nov-22', '24-Nov-22', '25-Nov-22', '28-Nov-22',
             '29-Nov-22', '30-Nov-22', '01-Dec-22', '02-Dec-22', '05-Dec-22', '06-Dec-22', '07-Dec-22', '08-Dec-22',
             '09-Dec-22', '12-Dec-22', '13-Dec-22', '14-Dec-22', '15-Dec-22', '16-Dec-22', '19-Dec-22', '20-Dec-22',
             '21-Dec-22', '22-Dec-22', '23-Dec-22', '26-Dec-22', '27-Dec-22', '28-Dec-22', '29-Dec-22', '30-Dec-22',
             '02-Jan-23', '03-Jan-23', '04-Jan-23', '05-Jan-23', '06-Jan-23', '09-Jan-23', '10-Jan-23', '11-Jan-23',
             '12-Jan-23', '13-Jan-23', '16-Jan-23', '17-Jan-23', '18-Jan-23', '19-Jan-23', '20-Jan-23', '23-Jan-23',
             '24-Jan-23', '25-Jan-23', '26-Jan-23', '27-Jan-23', '30-Jan-23', '31-Jan-23', '01-Feb-23', '02-Feb-23',
             '03-Feb-23', '06-Feb-23', '07-Feb-23', '08-Feb-23', '09-Feb-23', '10-Feb-23', '13-Feb-23', '14-Feb-23',
             '15-Feb-23', '16-Feb-23', '17-Feb-23', '20-Feb-23', '21-Feb-23', '22-Feb-23', '23-Feb-23', '24-Feb-23',
             '27-Feb-23', '28-Feb-23', '01-Mar-23', '02-Mar-23', '03-Mar-23', '06-Mar-23', '07-Mar-23', '08-Mar-23',
             '09-Mar-23', '10-Mar-23', '13-Mar-23', '14-Mar-23', '15-Mar-23', '16-Mar-23', '17-Mar-23', '20-Mar-23',
             '21-Mar-23', '22-Mar-23', '23-Mar-23', '24-Mar-23', '27-Mar-23', '28-Mar-23', '29-Mar-23', '30-Mar-23',
             '31-Mar-23', '03-Apr-23', '04-Apr-23', '05-Apr-23', '06-Apr-23', '07-Apr-23', '10-Apr-23', '11-Apr-23',
             '12-Apr-23', '13-Apr-23', '14-Apr-23', '17-Apr-23', '18-Apr-23', '19-Apr-23', '20-Apr-23', '21-Apr-23',
             '24-Apr-23', '25-Apr-23', '26-Apr-23', '27-Apr-23', '28-Apr-23', '01-May-23', '02-May-23', '03-May-23',
             '04-May-23', '05-May-23', '08-May-23', '09-May-23', '10-May-23', '11-May-23', '12-May-23', '15-May-23',
             '16-May-23', '17-May-23', '18-May-23', '19-May-23', '22-May-23', '23-May-23', '24-May-23', '25-May-23',
             '26-May-23', '29-May-23', '30-May-23', '31-May-23', '01-Jun-23', '02-Jun-23', '05-Jun-23', '06-Jun-23',
             '07-Jun-23', '08-Jun-23', '09-Jun-23', '12-Jun-23', '13-Jun-23', '14-Jun-23', '15-Jun-23', '16-Jun-23',
             '19-Jun-23', '20-Jun-23', '21-Jun-23', '22-Jun-23', '23-Jun-23', '26-Jun-23', '27-Jun-23', '28-Jun-23',
             '29-Jun-23', '30-Jun-23', '03-Jul-23', '04-Jul-23', '05-Jul-23', '06-Jul-23', '07-Jul-23', '10-Jul-23',
             '11-Jul-23', '12-Jul-23', '13-Jul-23', '14-Jul-23', '17-Jul-23', '18-Jul-23', '19-Jul-23', '20-Jul-23',
             '21-Jul-23', '24-Jul-23', '25-Jul-23', '26-Jul-23', '27-Jul-23', '28-Jul-23', '31-Jul-23', '01-Aug-23',
             '02-Aug-23', '03-Aug-23', '04-Aug-23', '07-Aug-23', '08-Aug-23', '09-Aug-23', '10-Aug-23', '11-Aug-23',
             '14-Aug-23', '15-Aug-23', '16-Aug-23', '17-Aug-23', '18-Aug-23', '21-Aug-23', '22-Aug-23', '23-Aug-23',
             '24-Aug-23', '25-Aug-23', '28-Aug-23', '29-Aug-23', '30-Aug-23', '31-Aug-23', '01-Sep-23', '04-Sep-23',
             '05-Sep-23', '06-Sep-23', '07-Sep-23', '08-Sep-23', '11-Sep-23', '12-Sep-23', '13-Sep-23', '14-Sep-23',
             '15-Sep-23', '18-Sep-23', '19-Sep-23', '20-Sep-23', '21-Sep-23', '22-Sep-23', '25-Sep-23', '26-Sep-23',
             '27-Sep-23', '28-Sep-23', '29-Sep-23', '02-Oct-23', '03-Oct-23', '04-Oct-23', '05-Oct-23', '06-Oct-23',
             '09-Oct-23', '10-Oct-23', '11-Oct-23', '12-Oct-23', '13-Oct-23', '16-Oct-23', '17-Oct-23', '18-Oct-23',
             '19-Oct-23', '20-Oct-23', '23-Oct-23', '24-Oct-23', '25-Oct-23', '26-Oct-23', '27-Oct-23', '30-Oct-23',
             '31-Oct-23', '01-Nov-23', '02-Nov-23', '03-Nov-23', '06-Nov-23', '07-Nov-23', '08-Nov-23', '09-Nov-23',
             '10-Nov-23', '13-Nov-23', '14-Nov-23', '15-Nov-23', '16-Nov-23', '17-Nov-23', '20-Nov-23', '21-Nov-23',
             '22-Nov-23', '23-Nov-23', '24-Nov-23', '27-Nov-23', '28-Nov-23', '29-Nov-23', '30-Nov-23', '01-Dec-23',
             '04-Dec-23', '05-Dec-23', '06-Dec-23', '07-Dec-23', '08-Dec-23', '11-Dec-23', '12-Dec-23', '13-Dec-23',
             '14-Dec-23', '15-Dec-23', '18-Dec-23', '19-Dec-23', '20-Dec-23', '21-Dec-23', '22-Dec-23', '25-Dec-23',
             '26-Dec-23', '27-Dec-23', '28-Dec-23', '29-Dec-23', '01-Jan-24', '02-Jan-24', '03-Jan-24', '04-Jan-24',
             '05-Jan-24', '08-Jan-24', '09-Jan-24', '10-Jan-24', '11-Jan-24', '12-Jan-24', '15-Jan-24', '16-Jan-24',
             '17-Jan-24', '18-Jan-24', '19-Jan-24', '22-Jan-24', '23-Jan-24', '24-Jan-24', '25-Jan-24', '26-Jan-24',
             '29-Jan-24', '30-Jan-24', '31-Jan-24', '01-Feb-24', '02-Feb-24', '05-Feb-24', '06-Feb-24', '07-Feb-24',
             '08-Feb-24', '09-Feb-24', '12-Feb-24', '13-Feb-24', '14-Feb-24', '15-Feb-24', '16-Feb-24', '19-Feb-24',
             '20-Feb-24', '21-Feb-24', '22-Feb-24', '23-Feb-24', '26-Feb-24', '27-Feb-24', '28-Feb-24', '29-Feb-24',
             '01-Mar-24', '04-Mar-24', '05-Mar-24', '06-Mar-24', '07-Mar-24', '08-Mar-24', '11-Mar-24', '12-Mar-24',
             '13-Mar-24', '14-Mar-24', '15-Mar-24', '18-Mar-24', '19-Mar-24', '20-Mar-24', '21-Mar-24', '22-Mar-24',
             '25-Mar-24', '26-Mar-24', '27-Mar-24', '28-Mar-24', '29-Mar-24', '01-Apr-24', '02-Apr-24', '03-Apr-24',
             '04-Apr-24', '05-Apr-24', '08-Apr-24', '09-Apr-24', '10-Apr-24', '11-Apr-24', '12-Apr-24', '15-Apr-24',
             '16-Apr-24', '17-Apr-24', '18-Apr-24', '19-Apr-24', '22-Apr-24', '23-Apr-24', '24-Apr-24', '25-Apr-24',
             '26-Apr-24', '29-Apr-24', '30-Apr-24', '01-May-24', '02-May-24', '03-May-24', '06-May-24', '07-May-24',
             '08-May-24', '09-May-24', '10-May-24', '13-May-24', '14-May-24', '15-May-24', '16-May-24', '17-May-24',
             '20-May-24', '21-May-24', '22-May-24', '23-May-24', '24-May-24', '27-May-24', '28-May-24', '29-May-24',
             '30-May-24', '31-May-24', '03-Jun-24', '04-Jun-24', '05-Jun-24', '06-Jun-24', '07-Jun-24', '10-Jun-24',
             '11-Jun-24', '12-Jun-24', '13-Jun-24', '14-Jun-24', '17-Jun-24', '18-Jun-24', '19-Jun-24', '20-Jun-24',
             '21-Jun-24', '24-Jun-24', '25-Jun-24', '26-Jun-24', '27-Jun-24', '28-Jun-24', '01-Jul-24', '02-Jul-24',
             '03-Jul-24', '04-Jul-24', '05-Jul-24', '08-Jul-24', '09-Jul-24', '10-Jul-24', '11-Jul-24', '12-Jul-24',
             '15-Jul-24', '16-Jul-24', '17-Jul-24', '18-Jul-24', '19-Jul-24', '22-Jul-24', '23-Jul-24', '24-Jul-24',
             '25-Jul-24', '26-Jul-24', '29-Jul-24', '30-Jul-24', '31-Jul-24', '01-Aug-24', '02-Aug-24', '05-Aug-24',
             '06-Aug-24', '07-Aug-24', '08-Aug-24', '09-Aug-24', '12-Aug-24', '13-Aug-24', '14-Aug-24', '15-Aug-24',
             '16-Aug-24', '19-Aug-24', '20-Aug-24', '21-Aug-24', '22-Aug-24', '23-Aug-24', '26-Aug-24', '27-Aug-24',
             '28-Aug-24', '29-Aug-24', '30-Aug-24', '02-Sep-24', '03-Sep-24', '04-Sep-24', '05-Sep-24', '06-Sep-24',
             '09-Sep-24', '10-Sep-24', '11-Sep-24', '12-Sep-24', '13-Sep-24', '16-Sep-24', '17-Sep-24', '18-Sep-24',
             '19-Sep-24', '20-Sep-24', '23-Sep-24', '24-Sep-24', '25-Sep-24', '26-Sep-24', '27-Sep-24', '30-Sep-24',
             '01-Oct-24', '02-Oct-24', '03-Oct-24', '04-Oct-24', '07-Oct-24', '08-Oct-24', '09-Oct-24', '10-Oct-24',
             '11-Oct-24', '14-Oct-24', '15-Oct-24', '16-Oct-24', '17-Oct-24', '18-Oct-24', '21-Oct-24', '22-Oct-24',
             '23-Oct-24', '24-Oct-24', '25-Oct-24', '28-Oct-24', '29-Oct-24', '30-Oct-24', '31-Oct-24', '01-Nov-24',
             '04-Nov-24', '05-Nov-24', '06-Nov-24', '07-Nov-24', '08-Nov-24', '11-Nov-24', '12-Nov-24', '13-Nov-24',
             '14-Nov-24', '15-Nov-24', '18-Nov-24', '19-Nov-24', '20-Nov-24', '21-Nov-24', '22-Nov-24', '25-Nov-24',
             '26-Nov-24', '27-Nov-24', '28-Nov-24', '29-Nov-24', '02-Dec-24', '03-Dec-24', '04-Dec-24', '05-Dec-24',
             '06-Dec-24', '09-Dec-24', '10-Dec-24', '11-Dec-24', '12-Dec-24', '13-Dec-24', '16-Dec-24', '17-Dec-24',
             '18-Dec-24', '19-Dec-24', '20-Dec-24', '23-Dec-24', '24-Dec-24', '25-Dec-24', '26-Dec-24', '27-Dec-24',
             '30-Dec-24', '31-Dec-24', '01-Jan-25', '02-Jan-25', '03-Jan-25', '06-Jan-25', '07-Jan-25', '08-Jan-25',
             '09-Jan-25', '10-Jan-25', '13-Jan-25', '14-Jan-25', '15-Jan-25', '16-Jan-25', '17-Jan-25', '20-Jan-25',
             '21-Jan-25', '22-Jan-25', '23-Jan-25', '24-Jan-25', '27-Jan-25', '28-Jan-25', '29-Jan-25', '30-Jan-25',
             '31-Jan-25', '03-Feb-25', '04-Feb-25', '05-Feb-25', '06-Feb-25', '07-Feb-25', '10-Feb-25', '11-Feb-25',
             '12-Feb-25', '13-Feb-25', '14-Feb-25', '17-Feb-25', '18-Feb-25', '19-Feb-25', '20-Feb-25', '21-Feb-25',
             '24-Feb-25', '25-Feb-25', '26-Feb-25', '27-Feb-25', '28-Feb-25', '03-Mar-25', '04-Mar-25', '05-Mar-25',
             '06-Mar-25', '07-Mar-25', '10-Mar-25', '11-Mar-25', '12-Mar-25', '13-Mar-25', '14-Mar-25', '17-Mar-25',
             '18-Mar-25', '19-Mar-25', '20-Mar-25', '21-Mar-25', '24-Mar-25', '25-Mar-25', '26-Mar-25', '27-Mar-25',
             '28-Mar-25', '31-Mar-25', '01-Apr-25', '02-Apr-25', '03-Apr-25', '04-Apr-25', '07-Apr-25', '08-Apr-25',
             '09-Apr-25', '10-Apr-25', '11-Apr-25', '14-Apr-25', '15-Apr-25', '16-Apr-25', '17-Apr-25', '18-Apr-25',
             '21-Apr-25', '22-Apr-25', '23-Apr-25', '24-Apr-25', '25-Apr-25', '28-Apr-25', '29-Apr-25', '30-Apr-25',
             '01-May-25', '02-May-25', '05-May-25', '06-May-25', '07-May-25', '08-May-25', '09-May-25', '12-May-25',
             '13-May-25', '14-May-25', '15-May-25', '16-May-25', '19-May-25', '20-May-25', '21-May-25', '22-May-25',
             '23-May-25', '26-May-25', '27-May-25', '28-May-25', '29-May-25', '30-May-25', '02-Jun-25', '03-Jun-25',
             '04-Jun-25', '05-Jun-25', '06-Jun-25', '09-Jun-25', '10-Jun-25', '11-Jun-25', '12-Jun-25', '13-Jun-25',
             '16-Jun-25', '17-Jun-25', '18-Jun-25', '19-Jun-25', '20-Jun-25', '23-Jun-25', '24-Jun-25', '25-Jun-25',
             '26-Jun-25', '27-Jun-25', '30-Jun-25', '01-Jul-25', '02-Jul-25', '03-Jul-25', '04-Jul-25', '07-Jul-25',
             '08-Jul-25', '09-Jul-25', '10-Jul-25', '11-Jul-25', '14-Jul-25', '15-Jul-25', '16-Jul-25', '17-Jul-25',
             '18-Jul-25', '21-Jul-25', '22-Jul-25', '23-Jul-25', '24-Jul-25', '25-Jul-25', '28-Jul-25', '29-Jul-25',
             '30-Jul-25', '31-Jul-25', '01-Aug-25', '04-Aug-25', '05-Aug-25', '06-Aug-25', '07-Aug-25', '08-Aug-25',
             '11-Aug-25', '12-Aug-25', '13-Aug-25', '14-Aug-25', '15-Aug-25', '18-Aug-25', '19-Aug-25', '20-Aug-25',
             '21-Aug-25', '22-Aug-25', '25-Aug-25', '26-Aug-25', '27-Aug-25', '28-Aug-25', '29-Aug-25', '01-Sep-25',
             '02-Sep-25', '03-Sep-25', '04-Sep-25', '05-Sep-25', '08-Sep-25', '09-Sep-25', '10-Sep-25', '11-Sep-25',
             '12-Sep-25', '15-Sep-25', '16-Sep-25', '17-Sep-25', '18-Sep-25', '19-Sep-25', '22-Sep-25', '23-Sep-25',
             '24-Sep-25', '25-Sep-25', '26-Sep-25', '29-Sep-25', '30-Sep-25', '01-Oct-25', '02-Oct-25', '03-Oct-25',
             '06-Oct-25', '07-Oct-25', '08-Oct-25', '09-Oct-25', '10-Oct-25', '13-Oct-25', '14-Oct-25', '15-Oct-25',
             '16-Oct-25', '17-Oct-25', '20-Oct-25', '21-Oct-25', '22-Oct-25', '23-Oct-25', '24-Oct-25', '27-Oct-25',
             '28-Oct-25', '29-Oct-25', '30-Oct-25', '31-Oct-25', '03-Nov-25', '04-Nov-25', '05-Nov-25', '06-Nov-25',
             '07-Nov-25', '10-Nov-25', '11-Nov-25', '12-Nov-25', '13-Nov-25', '14-Nov-25', '17-Nov-25', '18-Nov-25',
             '19-Nov-25', '20-Nov-25', '21-Nov-25', '24-Nov-25', '25-Nov-25', '26-Nov-25', '27-Nov-25', '28-Nov-25',
             '01-Dec-25', '02-Dec-25', '03-Dec-25', '04-Dec-25', '05-Dec-25', '08-Dec-25', '09-Dec-25', '10-Dec-25',
             '11-Dec-25', '12-Dec-25', '15-Dec-25', '16-Dec-25', '17-Dec-25', '18-Dec-25', '19-Dec-25', '22-Dec-25',
             '23-Dec-25', '24-Dec-25', '25-Dec-25', '26-Dec-25', '29-Dec-25', '30-Dec-25', '31-Dec-25', '01-Jan-26',
             '02-Jan-26', '05-Jan-26', '06-Jan-26', '07-Jan-26', '08-Jan-26', '09-Jan-26', '12-Jan-26', '13-Jan-26',
             '14-Jan-26', '15-Jan-26', '16-Jan-26', '19-Jan-26', '20-Jan-26', '21-Jan-26', '22-Jan-26', '23-Jan-26',
             '26-Jan-26', '27-Jan-26', '28-Jan-26', '29-Jan-26', '30-Jan-26', '02-Feb-26', '03-Feb-26', '04-Feb-26',
             '05-Feb-26', '06-Feb-26', '09-Feb-26', '10-Feb-26', '11-Feb-26', '12-Feb-26', '13-Feb-26', '16-Feb-26',
             '17-Feb-26', '18-Feb-26', '19-Feb-26', '20-Feb-26', '23-Feb-26', '24-Feb-26', '25-Feb-26', '26-Feb-26',
             '27-Feb-26', '02-Mar-26', '03-Mar-26', '04-Mar-26', '05-Mar-26', '06-Mar-26', '09-Mar-26', '10-Mar-26',
             '11-Mar-26', '12-Mar-26', '13-Mar-26', '16-Mar-26', '17-Mar-26', '18-Mar-26', '19-Mar-26', '20-Mar-26',
             '23-Mar-26', '24-Mar-26', '25-Mar-26', '26-Mar-26', '27-Mar-26', '30-Mar-26', '31-Mar-26', '01-Apr-26',
             '02-Apr-26', '03-Apr-26', '06-Apr-26', '07-Apr-26', '08-Apr-26', '09-Apr-26', '10-Apr-26', '13-Apr-26',
             '14-Apr-26', '15-Apr-26', '16-Apr-26', '17-Apr-26', '20-Apr-26', '21-Apr-26', '22-Apr-26', '23-Apr-26',
             '24-Apr-26', '27-Apr-26', '28-Apr-26', '29-Apr-26', '30-Apr-26', '01-May-26', '04-May-26', '05-May-26',
             '06-May-26', '07-May-26', '08-May-26', '11-May-26', '12-May-26', '13-May-26', '14-May-26', '15-May-26',
             '18-May-26', '19-May-26', '20-May-26', '21-May-26', '22-May-26', '25-May-26', '26-May-26', '27-May-26',
             '28-May-26', '29-May-26', '01-Jun-26', '02-Jun-26', '03-Jun-26', '04-Jun-26', '05-Jun-26', '08-Jun-26',
             '09-Jun-26', '10-Jun-26', '11-Jun-26', '12-Jun-26', '15-Jun-26', '16-Jun-26', '17-Jun-26', '18-Jun-26',
             '19-Jun-26', '22-Jun-26', '23-Jun-26', '24-Jun-26', '25-Jun-26', '26-Jun-26', '29-Jun-26', '30-Jun-26',
             '01-Jul-26', '02-Jul-26', '03-Jul-26', '06-Jul-26', '07-Jul-26', '08-Jul-26', '09-Jul-26', '10-Jul-26',
             '13-Jul-26', '14-Jul-26', '15-Jul-26', '16-Jul-26', '17-Jul-26', '20-Jul-26', '21-Jul-26', '22-Jul-26',
             '23-Jul-26', '24-Jul-26', '27-Jul-26', '28-Jul-26', '29-Jul-26', '30-Jul-26', '31-Jul-26', '03-Aug-26',
             '04-Aug-26', '05-Aug-26', '06-Aug-26', '07-Aug-26', '10-Aug-26', '11-Aug-26', '12-Aug-26', '13-Aug-26',
             '14-Aug-26', '17-Aug-26', '18-Aug-26', '19-Aug-26', '20-Aug-26', '21-Aug-26', '24-Aug-26', '25-Aug-26',
             '26-Aug-26', '27-Aug-26', '28-Aug-26', '31-Aug-26', '01-Sep-26', '02-Sep-26', '03-Sep-26', '04-Sep-26',
             '07-Sep-26', '08-Sep-26', '09-Sep-26', '10-Sep-26', '11-Sep-26', '14-Sep-26', '15-Sep-26', '16-Sep-26',
             '17-Sep-26', '18-Sep-26', '21-Sep-26', '22-Sep-26', '23-Sep-26', '24-Sep-26', '25-Sep-26', '28-Sep-26',
             '29-Sep-26', '30-Sep-26', '01-Oct-26', '02-Oct-26', '05-Oct-26', '06-Oct-26', '07-Oct-26', '08-Oct-26',
             '09-Oct-26', '12-Oct-26', '13-Oct-26', '14-Oct-26', '15-Oct-26', '16-Oct-26', '19-Oct-26', '20-Oct-26',
             '21-Oct-26', '22-Oct-26', '23-Oct-26', '26-Oct-26', '27-Oct-26', '28-Oct-26', '29-Oct-26', '30-Oct-26',
             '02-Nov-26', '03-Nov-26', '04-Nov-26', '05-Nov-26', '06-Nov-26', '09-Nov-26', '10-Nov-26', '11-Nov-26',
             '12-Nov-26', '13-Nov-26', '16-Nov-26', '17-Nov-26', '18-Nov-26', '19-Nov-26', '20-Nov-26', '23-Nov-26',
             '24-Nov-26', '25-Nov-26', '26-Nov-26', '27-Nov-26', '30-Nov-26', '01-Dec-26', '02-Dec-26', '03-Dec-26',
             '04-Dec-26', '07-Dec-26', '08-Dec-26', '09-Dec-26', '10-Dec-26', '11-Dec-26', '14-Dec-26', '15-Dec-26',
             '16-Dec-26', '17-Dec-26', '18-Dec-26', '21-Dec-26', '22-Dec-26', '23-Dec-26', '24-Dec-26', '25-Dec-26',
             '28-Dec-26', '29-Dec-26', '30-Dec-26', '31-Dec-26', '01-Jan-27', '04-Jan-27', '05-Jan-27', '06-Jan-27',
             '07-Jan-27', '08-Jan-27', '11-Jan-27', '12-Jan-27', '13-Jan-27', '14-Jan-27', '15-Jan-27', '18-Jan-27',
             '19-Jan-27', '20-Jan-27', '21-Jan-27', '22-Jan-27', '25-Jan-27', '26-Jan-27', '27-Jan-27', '28-Jan-27',
             '29-Jan-27', '01-Feb-27', '02-Feb-27', '03-Feb-27', '04-Feb-27', '05-Feb-27', '08-Feb-27', '09-Feb-27',
             '10-Feb-27', '11-Feb-27', '12-Feb-27', '15-Feb-27', '16-Feb-27', '17-Feb-27', '18-Feb-27', '19-Feb-27',
             '22-Feb-27', '23-Feb-27', '24-Feb-27', '25-Feb-27', '26-Feb-27', '01-Mar-27', '02-Mar-27', '03-Mar-27',
             '04-Mar-27', '05-Mar-27', '08-Mar-27', '09-Mar-27', '10-Mar-27', '11-Mar-27', '12-Mar-27', '15-Mar-27',
             '16-Mar-27', '17-Mar-27', '18-Mar-27', '19-Mar-27', '22-Mar-27', '23-Mar-27', '24-Mar-27', '25-Mar-27',
             '26-Mar-27', '29-Mar-27', '30-Mar-27', '31-Mar-27', '01-Apr-27', '02-Apr-27', '05-Apr-27', '06-Apr-27',
             '07-Apr-27', '08-Apr-27', '09-Apr-27', '12-Apr-27', '13-Apr-27', '14-Apr-27', '15-Apr-27', '16-Apr-27',
             '19-Apr-27', '20-Apr-27', '21-Apr-27', '22-Apr-27', '23-Apr-27', '26-Apr-27', '27-Apr-27', '28-Apr-27',
             '29-Apr-27', '30-Apr-27', '03-May-27', '04-May-27', '05-May-27', '06-May-27', '07-May-27', '10-May-27',
             '11-May-27', '12-May-27', '13-May-27', '14-May-27', '17-May-27', '18-May-27', '19-May-27', '20-May-27',
             '21-May-27', '24-May-27', '25-May-27', '26-May-27', '27-May-27', '28-May-27', '31-May-27', '01-Jun-27',
             '02-Jun-27', '03-Jun-27', '04-Jun-27', '07-Jun-27', '08-Jun-27', '09-Jun-27', '10-Jun-27', '11-Jun-27',
             '14-Jun-27', '15-Jun-27', '16-Jun-27', '17-Jun-27', '18-Jun-27', '21-Jun-27', '22-Jun-27', '23-Jun-27',
             '24-Jun-27', '25-Jun-27', '28-Jun-27', '29-Jun-27', '30-Jun-27', '01-Jul-27', '02-Jul-27', '05-Jul-27',
             '06-Jul-27', '07-Jul-27', '08-Jul-27', '09-Jul-27', '12-Jul-27', '13-Jul-27', '14-Jul-27', '15-Jul-27',
             '16-Jul-27', '19-Jul-27', '20-Jul-27', '21-Jul-27', '22-Jul-27', '23-Jul-27', '26-Jul-27', '27-Jul-27',
             '28-Jul-27', '29-Jul-27', '30-Jul-27', '02-Aug-27', '03-Aug-27', '04-Aug-27', '05-Aug-27', '06-Aug-27',
             '09-Aug-27', '10-Aug-27', '11-Aug-27', '12-Aug-27', '13-Aug-27', '16-Aug-27', '17-Aug-27', '18-Aug-27',
             '19-Aug-27', '20-Aug-27', '23-Aug-27', '24-Aug-27', '25-Aug-27', '26-Aug-27', '27-Aug-27', '30-Aug-27',
             '31-Aug-27', '01-Sep-27', '02-Sep-27', '03-Sep-27', '06-Sep-27', '07-Sep-27', '08-Sep-27', '09-Sep-27',
             '10-Sep-27', '13-Sep-27', '14-Sep-27', '15-Sep-27', '16-Sep-27', '17-Sep-27', '20-Sep-27', '21-Sep-27',
             '22-Sep-27', '23-Sep-27', '24-Sep-27', '27-Sep-27', '28-Sep-27', '29-Sep-27', '30-Sep-27', '01-Oct-27',
             '04-Oct-27', '05-Oct-27', '06-Oct-27', '07-Oct-27', '08-Oct-27', '11-Oct-27', '12-Oct-27', '13-Oct-27',
             '14-Oct-27', '15-Oct-27', '18-Oct-27', '19-Oct-27', '20-Oct-27', '21-Oct-27', '22-Oct-27', '25-Oct-27',
             '26-Oct-27', '27-Oct-27', '28-Oct-27', '29-Oct-27', '01-Nov-27', '02-Nov-27', '03-Nov-27', '04-Nov-27',
             '05-Nov-27', '08-Nov-27', '09-Nov-27', '10-Nov-27', '11-Nov-27', '12-Nov-27', '15-Nov-27', '16-Nov-27',
             '17-Nov-27', '18-Nov-27', '19-Nov-27', '22-Nov-27', '23-Nov-27', '24-Nov-27', '25-Nov-27', '26-Nov-27',
             '29-Nov-27', '30-Nov-27', '01-Dec-27', '02-Dec-27', '03-Dec-27', '06-Dec-27', '07-Dec-27', '08-Dec-27',
             '09-Dec-27', '10-Dec-27', '13-Dec-27', '14-Dec-27', '15-Dec-27', '16-Dec-27', '17-Dec-27', '20-Dec-27',
             '21-Dec-27', '22-Dec-27', '23-Dec-27', '24-Dec-27', '27-Dec-27', '28-Dec-27', '29-Dec-27', '30-Dec-27',
             '31-Dec-27', '03-Jan-28', '04-Jan-28', '05-Jan-28', '06-Jan-28', '07-Jan-28', '10-Jan-28', '11-Jan-28',
             '12-Jan-28', '13-Jan-28', '14-Jan-28', '17-Jan-28', '18-Jan-28', '19-Jan-28', '20-Jan-28', '21-Jan-28',
             '24-Jan-28', '25-Jan-28', '26-Jan-28', '27-Jan-28', '28-Jan-28', '31-Jan-28', '01-Feb-28', '02-Feb-28',
             '03-Feb-28', '04-Feb-28', '07-Feb-28', '08-Feb-28', '09-Feb-28', '10-Feb-28', '11-Feb-28', '14-Feb-28',
             '15-Feb-28', '16-Feb-28', '17-Feb-28', '18-Feb-28', '21-Feb-28', '22-Feb-28', '23-Feb-28', '24-Feb-28',
             '25-Feb-28', '28-Feb-28', '29-Feb-28', '01-Mar-28', '02-Mar-28', '03-Mar-28', '06-Mar-28', '07-Mar-28',
             '08-Mar-28', '09-Mar-28', '10-Mar-28', '13-Mar-28', '14-Mar-28', '15-Mar-28', '16-Mar-28', '17-Mar-28',
             '20-Mar-28', '21-Mar-28', '22-Mar-28', '23-Mar-28', '24-Mar-28', '27-Mar-28', '28-Mar-28', '29-Mar-28',
             '30-Mar-28', '31-Mar-28', '03-Apr-28', '04-Apr-28', '05-Apr-28', '06-Apr-28', '07-Apr-28', '10-Apr-28',
             '11-Apr-28', '12-Apr-28', '13-Apr-28', '14-Apr-28', '17-Apr-28', '18-Apr-28', '19-Apr-28', '20-Apr-28',
             '21-Apr-28', '24-Apr-28', '25-Apr-28', '26-Apr-28', '27-Apr-28', '28-Apr-28', '01-May-28', '02-May-28',
             '03-May-28', '04-May-28', '05-May-28', '08-May-28', '09-May-28', '10-May-28', '11-May-28', '12-May-28',
             '15-May-28', '16-May-28', '17-May-28', '18-May-28', '19-May-28', '22-May-28', '23-May-28', '24-May-28',
             '25-May-28', '26-May-28', '29-May-28', '30-May-28', '31-May-28', '01-Jun-28', '02-Jun-28', '05-Jun-28',
             '06-Jun-28', '07-Jun-28', '08-Jun-28', '09-Jun-28', '12-Jun-28', '13-Jun-28', '14-Jun-28', '15-Jun-28',
             '16-Jun-28', '19-Jun-28', '20-Jun-28', '21-Jun-28', '22-Jun-28', '23-Jun-28', '26-Jun-28', '27-Jun-28',
             '28-Jun-28', '29-Jun-28', '30-Jun-28', '03-Jul-28', '04-Jul-28', '05-Jul-28', '06-Jul-28', '07-Jul-28',
             '10-Jul-28', '11-Jul-28', '12-Jul-28', '13-Jul-28', '14-Jul-28', '17-Jul-28', '18-Jul-28', '19-Jul-28',
             '20-Jul-28', '21-Jul-28', '24-Jul-28', '25-Jul-28', '26-Jul-28', '27-Jul-28', '28-Jul-28', '31-Jul-28',
             '01-Aug-28', '02-Aug-28', '03-Aug-28', '04-Aug-28', '07-Aug-28', '08-Aug-28', '09-Aug-28', '10-Aug-28',
             '11-Aug-28', '14-Aug-28', '15-Aug-28', '16-Aug-28', '17-Aug-28', '18-Aug-28', '21-Aug-28', '22-Aug-28',
             '23-Aug-28', '24-Aug-28', '25-Aug-28', '28-Aug-28', '29-Aug-28', '30-Aug-28', '31-Aug-28', '01-Sep-28',
             '04-Sep-28', '05-Sep-28', '06-Sep-28', '07-Sep-28', '08-Sep-28', '11-Sep-28', '12-Sep-28', '13-Sep-28',
             '14-Sep-28', '15-Sep-28', '18-Sep-28', '19-Sep-28', '20-Sep-28', '21-Sep-28', '22-Sep-28', '25-Sep-28',
             '26-Sep-28', '27-Sep-28', '28-Sep-28', '29-Sep-28', '02-Oct-28', '03-Oct-28', '04-Oct-28', '05-Oct-28',
             '06-Oct-28', '09-Oct-28', '10-Oct-28', '11-Oct-28', '12-Oct-28', '13-Oct-28', '16-Oct-28', '17-Oct-28',
             '18-Oct-28', '19-Oct-28', '20-Oct-28', '23-Oct-28', '24-Oct-28', '25-Oct-28', '26-Oct-28', '27-Oct-28',
             '30-Oct-28', '31-Oct-28', '01-Nov-28', '02-Nov-28', '03-Nov-28', '06-Nov-28', '07-Nov-28', '08-Nov-28',
             '09-Nov-28', '10-Nov-28', '13-Nov-28', '14-Nov-28', '15-Nov-28', '16-Nov-28', '17-Nov-28', '20-Nov-28',
             '21-Nov-28', '22-Nov-28', '23-Nov-28', '24-Nov-28', '27-Nov-28', '28-Nov-28', '29-Nov-28', '30-Nov-28',
             '01-Dec-28', '04-Dec-28', '05-Dec-28', '06-Dec-28', '07-Dec-28', '08-Dec-28', '11-Dec-28', '12-Dec-28',
             '13-Dec-28', '14-Dec-28', '15-Dec-28', '18-Dec-28', '19-Dec-28', '20-Dec-28', '21-Dec-28', '22-Dec-28',
             '25-Dec-28', '26-Dec-28', '27-Dec-28', '28-Dec-28', '29-Dec-28', '01-Jan-29', '02-Jan-29', '03-Jan-29',
             '04-Jan-29', '05-Jan-29', '08-Jan-29', '09-Jan-29', '10-Jan-29', '11-Jan-29', '12-Jan-29', '15-Jan-29',
             '16-Jan-29', '17-Jan-29', '18-Jan-29', '19-Jan-29', '22-Jan-29', '23-Jan-29', '24-Jan-29', '25-Jan-29',
             '26-Jan-29', '29-Jan-29', '30-Jan-29', '31-Jan-29', '01-Feb-29', '02-Feb-29', '05-Feb-29', '06-Feb-29',
             '07-Feb-29', '08-Feb-29', '09-Feb-29', '12-Feb-29', '13-Feb-29', '14-Feb-29', '15-Feb-29', '16-Feb-29',
             '19-Feb-29', '20-Feb-29', '21-Feb-29', '22-Feb-29', '23-Feb-29', '26-Feb-29', '27-Feb-29', '28-Feb-29',
             '01-Mar-29', '02-Mar-29', '05-Mar-29', '06-Mar-29', '07-Mar-29', '08-Mar-29', '09-Mar-29', '12-Mar-29',
             '13-Mar-29', '14-Mar-29', '15-Mar-29', '16-Mar-29', '19-Mar-29', '20-Mar-29', '21-Mar-29', '22-Mar-29',
             '23-Mar-29', '26-Mar-29', '27-Mar-29', '28-Mar-29', '29-Mar-29', '30-Mar-29', '02-Apr-29', '03-Apr-29',
             '04-Apr-29', '05-Apr-29', '06-Apr-29', '09-Apr-29', '10-Apr-29', '11-Apr-29', '12-Apr-29', '13-Apr-29',
             '16-Apr-29', '17-Apr-29', '18-Apr-29', '19-Apr-29', '20-Apr-29', '23-Apr-29', '24-Apr-29', '25-Apr-29',
             '26-Apr-29', '27-Apr-29', '30-Apr-29', '01-May-29', '02-May-29', '03-May-29', '04-May-29', '07-May-29',
             '08-May-29', '09-May-29', '10-May-29', '11-May-29', '14-May-29', '15-May-29', '16-May-29', '17-May-29',
             '18-May-29', '21-May-29', '22-May-29', '23-May-29', '24-May-29', '25-May-29', '28-May-29', '29-May-29',
             '30-May-29', '31-May-29', '01-Jun-29', '04-Jun-29', '05-Jun-29', '06-Jun-29', '07-Jun-29', '08-Jun-29',
             '11-Jun-29', '12-Jun-29', '13-Jun-29', '14-Jun-29', '15-Jun-29', '18-Jun-29', '19-Jun-29', '20-Jun-29',
             '21-Jun-29', '22-Jun-29', '25-Jun-29', '26-Jun-29', '27-Jun-29', '28-Jun-29', '29-Jun-29', '02-Jul-29',
             '03-Jul-29', '04-Jul-29', '05-Jul-29', '06-Jul-29', '09-Jul-29', '10-Jul-29', '11-Jul-29', '12-Jul-29',
             '13-Jul-29', '16-Jul-29', '17-Jul-29', '18-Jul-29', '19-Jul-29', '20-Jul-29', '23-Jul-29', '24-Jul-29',
             '25-Jul-29', '26-Jul-29', '27-Jul-29', '30-Jul-29', '31-Jul-29', '01-Aug-29', '02-Aug-29', '03-Aug-29',
             '06-Aug-29', '07-Aug-29', '08-Aug-29', '09-Aug-29', '10-Aug-29', '13-Aug-29', '14-Aug-29', '15-Aug-29',
             '16-Aug-29', '17-Aug-29', '20-Aug-29', '21-Aug-29', '22-Aug-29', '23-Aug-29', '24-Aug-29', '27-Aug-29',
             '28-Aug-29', '29-Aug-29', '30-Aug-29', '31-Aug-29', '03-Sep-29', '04-Sep-29', '05-Sep-29', '06-Sep-29',
             '07-Sep-29', '10-Sep-29', '11-Sep-29', '12-Sep-29', '13-Sep-29', '14-Sep-29', '17-Sep-29', '18-Sep-29',
             '19-Sep-29', '20-Sep-29', '21-Sep-29', '24-Sep-29', '25-Sep-29', '26-Sep-29', '27-Sep-29', '28-Sep-29',
             '01-Oct-29', '02-Oct-29', '03-Oct-29', '04-Oct-29', '05-Oct-29', '08-Oct-29', '09-Oct-29', '10-Oct-29',
             '11-Oct-29', '12-Oct-29', '15-Oct-29', '16-Oct-29', '17-Oct-29', '18-Oct-29', '19-Oct-29', '22-Oct-29',
             '23-Oct-29', '24-Oct-29', '25-Oct-29', '26-Oct-29', '29-Oct-29', '30-Oct-29', '31-Oct-29', '01-Nov-29',
             '02-Nov-29', '05-Nov-29', '06-Nov-29', '07-Nov-29', '08-Nov-29', '09-Nov-29', '12-Nov-29', '13-Nov-29',
             '14-Nov-29', '15-Nov-29', '16-Nov-29', '19-Nov-29', '20-Nov-29', '21-Nov-29', '22-Nov-29', '23-Nov-29',
             '26-Nov-29', '27-Nov-29', '28-Nov-29', '29-Nov-29', '30-Nov-29', '03-Dec-29', '04-Dec-29', '05-Dec-29',
             '06-Dec-29', '07-Dec-29', '10-Dec-29', '11-Dec-29', '12-Dec-29', '13-Dec-29', '14-Dec-29', '17-Dec-29',
             '18-Dec-29', '19-Dec-29', '20-Dec-29', '21-Dec-29', '24-Dec-29', '25-Dec-29', '26-Dec-29', '27-Dec-29',
             '28-Dec-29', '31-Dec-29', '01-Jan-30', '02-Jan-30', '03-Jan-30', '04-Jan-30', '07-Jan-30', '08-Jan-30',
             '09-Jan-30', '10-Jan-30', '11-Jan-30', '14-Jan-30', '15-Jan-30', '16-Jan-30', '17-Jan-30', '18-Jan-30',
             '21-Jan-30', '22-Jan-30', '23-Jan-30', '24-Jan-30', '25-Jan-30', '28-Jan-30', '29-Jan-30', '30-Jan-30',
             '31-Jan-30', '01-Feb-30', '04-Feb-30', '05-Feb-30', '06-Feb-30', '07-Feb-30', '08-Feb-30', '11-Feb-30',
             '12-Feb-30', '13-Feb-30', '14-Feb-30', '15-Feb-30', '18-Feb-30', '19-Feb-30', '20-Feb-30', '21-Feb-30',
             '22-Feb-30', '25-Feb-30', '26-Feb-30', '27-Feb-30', '28-Feb-30', '01-Mar-30', '04-Mar-30', '05-Mar-30',
             '06-Mar-30', '07-Mar-30', '08-Mar-30', '11-Mar-30', '12-Mar-30', '13-Mar-30', '14-Mar-30', '15-Mar-30',
             '18-Mar-30', '19-Mar-30', '20-Mar-30', '21-Mar-30', '22-Mar-30', '25-Mar-30', '26-Mar-30', '27-Mar-30',
             '28-Mar-30', '29-Mar-30', '01-Apr-30', '02-Apr-30', '03-Apr-30', '04-Apr-30', '05-Apr-30', '08-Apr-30',
             '09-Apr-30', '10-Apr-30', '11-Apr-30', '12-Apr-30', '15-Apr-30', '16-Apr-30', '17-Apr-30', '18-Apr-30',
             '19-Apr-30', '22-Apr-30', '23-Apr-30', '24-Apr-30', '25-Apr-30', '26-Apr-30', '29-Apr-30', '30-Apr-30',
             '01-May-30', '02-May-30', '03-May-30', '06-May-30', '07-May-30', '08-May-30', '09-May-30', '10-May-30',
             '13-May-30', '14-May-30', '15-May-30', '16-May-30', '17-May-30', '20-May-30', '21-May-30', '22-May-30',
             '23-May-30', '24-May-30', '27-May-30', '28-May-30', '29-May-30', '30-May-30', '31-May-30', '03-Jun-30',
             '04-Jun-30', '05-Jun-30', '06-Jun-30', '07-Jun-30', '10-Jun-30', '11-Jun-30', '12-Jun-30', '13-Jun-30',
             '14-Jun-30', '17-Jun-30', '18-Jun-30', '19-Jun-30', '20-Jun-30', '21-Jun-30', '24-Jun-30', '25-Jun-30',
             '26-Jun-30', '27-Jun-30', '28-Jun-30', '01-Jul-30', '02-Jul-30', '03-Jul-30', '04-Jul-30', '05-Jul-30',
             '08-Jul-30', '09-Jul-30', '10-Jul-30', '11-Jul-30', '12-Jul-30', '15-Jul-30', '16-Jul-30', '17-Jul-30',
             '18-Jul-30', '19-Jul-30', '22-Jul-30', '23-Jul-30', '24-Jul-30', '25-Jul-30', '26-Jul-30', '29-Jul-30',
             '30-Jul-30', '31-Jul-30', '01-Aug-30', '02-Aug-30', '05-Aug-30', '06-Aug-30', '07-Aug-30', '08-Aug-30',
             '09-Aug-30', '12-Aug-30']

coords = {"date left": (1048, 512), "date right": (1348, 512), "left year": (982, 555), "right year": (1273, 555),
          "left left": (887, 552), "left jan": (898, 596), "right jan": (1184, 596), "jan 1 2020": (984, 614),
          "right left": (1172, 555), "nov 2023 right": (1294, 707), "left right": (1079, 555),
          "right right": (1368, 555),
          "jan 1 2021 right": (1331, 617), "jan 1 2021 left": (1048, 614), "jan 1 2022 right": (1366, 616),
          "jan 1 2022 left": (1081, 618), "jan 1 2023 right": (1174, 613), "jan 1 2023 left": (890, 617),
          "13 nov 2023 right": (1205, 680), "filter": (1460, 508), "download csv": (1467, 565)}


def csv_download_auto():
    pg.click(coords["date left"])
    sleep(2)
    pg.click(coords["left year"])
    sleep(2)
    pg.click(coords["left left"], clicks=3, interval=0.5)
    sleep(2)
    pg.click(coords["left jan"])
    sleep(2)
    pg.click(coords["jan 1 2020"])
    sleep(2)

    pg.click(coords["date right"])
    sleep(2)
    pg.click(coords["right year"])
    sleep(2)
    pg.click(coords["right left"], clicks=2, interval=0.5)
    sleep(2)
    pg.click(coords["right jan"])
    sleep(2)
    pg.click(coords["jan 1 2021 right"])
    sleep(2)

    pg.click(coords["filter"])
    sleep(2)
    pg.click(coords["download csv"])
    sleep(2)

    # 2nd csv
    pg.click(coords["date left"])
    sleep(2)
    pg.click(coords["left year"])
    sleep(2)
    pg.click(coords["left right"])
    sleep(2)
    pg.click(coords["left jan"])
    sleep(2)
    pg.click(coords["jan 1 2021 left"])
    sleep(2)

    pg.click(coords["date right"])
    sleep(2)
    pg.click(coords["right year"])
    sleep(2)
    pg.click(coords["right right"])
    sleep(2)
    pg.click(coords["right jan"])
    sleep(2)
    pg.click(coords["jan 1 2022 right"])
    sleep(2)
    pg.click(coords["filter"])
    sleep(2)
    pg.click(coords["download csv"])
    sleep(2)

    # jan 1 2022 - jan 1 2023
    pg.click(coords["date left"])
    sleep(2)
    pg.click(coords["left year"])
    sleep(2)
    pg.click(coords["left right"])
    sleep(2)
    pg.click(coords["left jan"])
    sleep(2)
    pg.click(coords["jan 1 2022 left"])
    sleep(2)

    pg.click(coords["date right"])
    sleep(2)
    pg.click(coords["right year"])
    sleep(2)
    pg.click(coords["right right"])
    sleep(2)
    pg.click(coords["right jan"])
    sleep(2)
    pg.click(coords["jan 1 2023 right"])
    sleep(2)
    pg.click(coords["filter"])
    sleep(2)
    pg.click(coords["download csv"])
    sleep(2)

    # jan 1 2023 - nov 13 2023
    pg.click(coords["date left"])
    sleep(2)
    pg.click(coords["left year"])
    sleep(2)
    pg.click(coords["left right"])
    sleep(2)
    pg.click(coords["left jan"])
    sleep(2)
    pg.click(coords["jan 1 2023 left"])
    sleep(2)

    pg.click(coords["date right"])
    sleep(2)
    pg.click(coords["right year"])
    sleep(2)
    # pg.click(coords["right right"])
    # sleep(2)
    pg.click(coords["nov 2023 right"])
    sleep(2)
    pg.click(coords["13 nov 2023 right"])
    sleep(2)
    pg.click(coords["filter"])
    sleep(2)
    pg.click(coords["download csv"])
    sleep(2)


def remove_heading_border(ws):
    # ws is not the worksheet name, but the worksheet object
    for r in range(1, 3):
        for c in range(1, 10):
            cell = sheet.cell(r, c)
            cell.border = None
            cell.font = Font("Arial", 11, bold=True, color='00ffffff')
            cell.fill = PatternFill(patternType='solid', fgColor="0000ff")
            cell.alignment = alignment


def bold_and_align():
    for r in range(3, 3001):
        for c in range(1, 12):
            sheet.cell(r, c).font = bold
            sheet.cell(r, c).alignment = alignment


for share in share_add_list:
    driver = webdriver.Chrome(options=options)

    driver.get(f"https://www.nseindia.com/get-quotes/equity?symbol={share}")
    try:
        sleep(2)
        myElem = WebDriverWait(driver, 10).until(ec.presence_of_element_located((By.ID, 'historic_data')))
        # sleep(5)
        driver.find_element(By.ID, 'historic_data').click()

        sleep(2)

        if share in double_name:
            for v in coords:
                coords[v] = (coords[v][0], coords[v][1]+27)

            csv_download_auto()

            for v in coords:
                coords[v] = (coords[v][0], coords[v][1]-27)

        else:
            csv_download_auto()

    except TimeoutException:
        print(f"{share} Loading took too much time!")
        continue

    # saving files as one
    # merging the files
    joined_files = os.path.join("C:\\Users\\admin\\Downloads", "Quote-Equity*.csv")

    # A list of all joined files is returned
    joined_list = glob.glob(joined_files)

    # Finally, the files are joined
    df = pd.concat(map(pd.read_csv, joined_list), ignore_index=True)

    # converting date column to datetime and sorting via that column
    df["Date "] = pd.to_datetime(df["Date "])
    df.sort_values('Date ', axis=0, inplace=True)

    # converting volume to number and then to lakhs
    df.replace(',', '', regex=True, inplace=True)
    df['VOLUME '] = df['VOLUME '].apply(pd.to_numeric, errors='coerce')
    df['VOLUME '] = df['VOLUME '] // 100000

    # renaming columns
    df.rename(
        columns={'Date ': 'Date', 'ltp ': 'Close', 'close ': 'LTP', 'HIGH ': 'High', 'LOW ': 'Low', 'VOLUME ': 'VOL'},
        inplace=True)

    # dropping duplicate date rows and diwali row
    df.drop_duplicates(subset="Date", keep='first', inplace=True)
    df.drop(df.index[-2], inplace=True)
    df.drop(df.index[220], inplace=True)  # 14 nov 2020

    # only keeping necessary columns
    df = df.iloc[:, [0, 3, 4, 6, 7, 11]]
    # print(df.to_string(max_cols=11))

    # saving to excel
    writer = pd.ExcelWriter(rf"C:\Users\admin\Downloads\current\{share}.xlsx", engine="xlsxwriter",
                            datetime_format="dd-mmm-yy")

    # Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name="D", index=False)
    wb1 = writer.book
    writer.close()

    insert_rows = [40, 52, 69, 71, 75, 77, 90, 106, 200, 231, 241, 260, 282, 314, 326, 330, 338, 343, 359, 408, 429,
                   445, 470, 485, 495, 543, 569, 580, 599, 600, 612, 682, 686, 698, 723, 738, 747, 804, 832, 849, 852,
                   855, 860, 871, 914, 947, 972, 981, 997]

    # formatting sheets with openpyxl (styling, inserting rows, headings, etc.)
    red = Font("Arial", 11, color='ff0000', bold=True)
    blue = Font("Arial", 11, color="0000ff", bold=True)
    bold = Font("Arial", 11, bold=True)
    alignment = Alignment(horizontal='center')

    wb = xl.load_workbook(rf'C:\Users\admin\Downloads\current\{share}.xlsx')
    sheet = wb['D']

    for row in insert_rows:
        sheet.insert_rows(row)

    sheet.insert_rows(1)  # heading row
    sheet.freeze_panes = sheet['A3']

    # merging heading columns
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=8)
    sheet.cell(1, 2).value = share

    # other headings
    sheet.cell(2, 7).value = "9:25 CL"
    sheet.cell(2, 8).value = "H/L DIFF"
    sheet.cell(2, 9).value = "CL DIFF"

    # bolding and aligning
    bold_and_align()

    # removing border formatting caused by pandas and adding style formatting to headings
    remove_heading_border(sheet)

    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(1, 12):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=13.57)

    sheet.column_dimensions = dim_holder

    # changing zoom scale to 115%
    sheet.sheet_view.zoomScale = 115

    # formatting and adding formulas row wise
    for i in range(len(date_list)):
        sheet.cell(i+3, 1).value = date_list[i]
        sheet.cell(i+3, 1).border = Border(right=Side(style='thin'))
        sheet.cell(i+3, 1).number_format = 'dd-mmm-yy'

        sheet.cell(i+3, 2).font = blue
        sheet.cell(i+3, 3).font = red

        # formulas for closing diff and h/l diff
        sheet.cell(i+3, 8).value = f'=B{i+3}-C{i+3}'
        sheet.cell(i+4, 9).value = f'=D{i+4}-D{i+3}'

    # TODO CHANGE THIS FOR ALGO SHARES!!!!!!!!!!!!!!!!!!
    wb.save(rf'E:\Daily Data work\ALGO\{share}.xlsx')

    # deleting .csv files from current after making a consolidated file
    dir_name = r'C:\Users\admin\Downloads'
    test = os.listdir(dir_name)

    for item in test:
        if item.endswith(".csv"):
            os.remove(os.path.join(dir_name, item))

    print(share)

end = time.time()
print(f"Total runtime of the program is {(end - begin) // 60} minutes and {(end - begin) % 60} seconds")