import os
import shutil

import openpyxl as xl
from openpyxl.styles import Font, Alignment
from xls2xlsx import XLS2XLSX

from date_variables import date, mnth, yr

cashHL_wb = xl.load_workbook(r'C:\Users\admin\PycharmProjects\daily data\cash high low.xlsx')
cashHL_sheet = cashHL_wb['Sheet1']

foHL_wb = xl.load_workbook(r'C:\Users\admin\PycharmProjects\daily data\fo high low.xlsx')
foHL_sheet = foHL_wb['Sheet1']

algoHL_wb = xl.load_workbook(r'C:\Users\admin\PycharmProjects\daily data\algo high low.xlsx')
algoHL_sheet = algoHL_wb['Sheet1']

# dict to store share names with their row in 'cash/algo/fo high low.xlsx' sheets respectively
cash_30_min_list = {"AARTIIND": 2, "ADANI": 3, "APOLLO": 4, "BAJFINSV": 5, "BAJFIN": 6, "BANBK": 7, "BARODA": 8, "BN": 4,
                    "DLF": 10, "EICHER": 11, "ESCORTS": 48, "FEDBANK": 12, "HCL": 13, "HINDALCO": 15, "IGL": 68, "INDUSIND": 17, "JIND": 19,
                    "LIC": 20, "M&M": 21, "M&MFIN": 22, "NIFTY": 10, "NTPC": 23, "SBIN": 25, "SUNTV": 26, "TM": 28,
                    "TP": 29, "TS": 30, "VEDL": 131}

# no decimal points in display
format_list = ["NIFTY", "EICHER", "BN"]

# shares that get their data from 'algo high low.xlsx', when adding shares that are in algo and not in cash, add to this
algo_shares = ["ESCORTS", "IGL", "VEDL"]

# copying hourlys (.xls) as backup
# path to source directory
src_dir = rf"E:\Daily Data work\hourlys 30 minute CASH\{yr}\{mnth}\{date}"

# path to destination directory
dest_dir = rf"C:\Users\admin\PycharmProjects\daily data\Daily Backup hourlys\30 min csh"

# getting all the files in the source directory
src_files = os.listdir(src_dir)
for file_name in src_files:
    full_file_name = os.path.join(src_dir, file_name)
    if os.path.isfile(full_file_name):
        shutil.copy(full_file_name, dest_dir)
print("Files copied as backup!")

# index for getting values from cash/fo high low sheets
# index_30_min = [2, 3, 4, 5, 6, 7, 4, 9, 10, 11, 12, 14, 16, 18, 19, 20, 21, 10, 22, 24, 25, 27, 28, 29]

# cash_30_min_list = ["ADANI"]

# idx = 0

# LTP and PREV
ltp_wb = xl.load_workbook(rf'C:\Users\admin\PycharmProjects\daily data\LTP PREV.xlsx')
ltp_sheet = ltp_wb["ltp"]
ltp_row = 2

while ltp_row <= len(ltp_sheet["A"]):
    ltp_sheet.cell(ltp_row, 3).value = ltp_sheet.cell(ltp_row, 2).value  # moving last day's LTP to 'PREV'

    share_name = ltp_sheet.cell(ltp_row, 1).value

    if share_name in ["BN", "NIFTY"]:    # separate for NIFTY and BN as their data LTP come from 'fo high low.xlsx'
        ltp_sheet.cell(ltp_row, 2).value = foHL_sheet.cell(cash_30_min_list[share_name], 5).value

    elif share_name in algo_shares:  # separate for shares who's LTP come from 'algo high low.xlsx'
        ltp_sheet.cell(ltp_row, 2).value = algoHL_sheet.cell(cash_30_min_list[share_name], 5).value
    else:
        ltp_sheet.cell(ltp_row, 2).value = cashHL_sheet.cell(cash_30_min_list[share_name], 5).value

    ltp_row += 1
    # idx += 1

idx = 0
ltp_row = 2

for share in cash_30_min_list:
    path = rf"E:\Daily Data work\hourlys 30 minute CASH\{yr}\{mnth}\{date}\{share}.xlsx"
    xls_path = rf"E:\Daily Data work\hourlys 30 minute CASH\{yr}\{mnth}\{date}\{share}.xls"
    x2x = XLS2XLSX(xls_path)

    cash_30_min_wb = x2x.to_xlsx()
    old_30_min_sheet = cash_30_min_wb[f"{share}-Sheet1"]

    new_30_min_sheet = cash_30_min_wb.create_sheet(f"{share}")

    # FIXED HEADINGS
    new_30_min_sheet.cell(6, 6).value = f'{share}'
    new_30_min_sheet.cell(6, 7).value = "HIGH"
    new_30_min_sheet.cell(6, 8).value = "LOW"
    new_30_min_sheet.cell(6, 9).value = "LTP"
    new_30_min_sheet.cell(6, 10).value = "PREV"

    new_30_min_sheet.cell(8, 6).value = "Time"
    new_30_min_sheet.cell(8, 7).value = "High Rate"
    new_30_min_sheet.cell(8, 8).value = "Low Rate"
    new_30_min_sheet.cell(8, 9).value = "Close Rate"

    old_sheet_row = 2

    while old_sheet_row <= len(old_30_min_sheet["A"]):
        # time
        new_30_min_sheet.cell(old_sheet_row + 7, 6).value = old_30_min_sheet.cell(old_sheet_row, 7).value
        new_30_min_sheet.cell(old_sheet_row + 7, 6).number_format = 'hh:mm AM/PM'

        # high
        new_30_min_sheet.cell(old_sheet_row + 7, 7).value = old_30_min_sheet.cell(old_sheet_row, 4).value

        # low
        new_30_min_sheet.cell(old_sheet_row + 7, 8).value = old_30_min_sheet.cell(old_sheet_row, 5).value

        # close
        new_30_min_sheet.cell(old_sheet_row + 7, 9).value = old_30_min_sheet.cell(old_sheet_row, 3).value

        old_sheet_row += 1

    del cash_30_min_wb[f"{share}-Sheet1"]

    # bolding the sheet
    for i in range(25):
        for j in range(15):
            new_30_min_sheet.cell(i+1, j+1).font = Font(bold=True)
            new_30_min_sheet.cell(i+1, j+1).alignment = Alignment(horizontal='center')

            # formatting to 0 decimal places if they are in format_list
            if share in format_list:
                if j > 5:
                    new_30_min_sheet.cell(i+1, j+1).number_format = '0'

    new_30_min_sheet.cell(7, 6).number_format = '0'     # for the 9:25 cl formatting

    # deleting 4:00 pm row
    new_30_min_sheet.delete_rows(22, 1)

    # filling LTP and PREV
    new_30_min_sheet.cell(7, 9).value = ltp_sheet.cell(ltp_row, 2).value  # LTP
    new_30_min_sheet.cell(7, 10).value = ltp_sheet.cell(ltp_row, 3).value  # PREV
    ltp_row += 1

    # filling rest of the data
    if share in ["BN", "NIFTY"]:    # separate for NIFTY and BN as their data will come from 'fo high low.xlsx'
        new_30_min_sheet.cell(7, 6).value = foHL_sheet.cell(cash_30_min_list[share], 7).value   # 9:25 cl
        new_30_min_sheet.cell(7, 7).value = foHL_sheet.cell(cash_30_min_list[share], 2).value   # HIGH
        new_30_min_sheet.cell(7, 8).value = foHL_sheet.cell(cash_30_min_list[share], 3).value   # LOW

    elif share in algo_shares:
        new_30_min_sheet.cell(7, 6).value = algoHL_sheet.cell(cash_30_min_list[share], 7).value  # 9:25 cl
        new_30_min_sheet.cell(7, 7).value = algoHL_sheet.cell(cash_30_min_list[share], 2).value  # HIGH
        new_30_min_sheet.cell(7, 8).value = algoHL_sheet.cell(cash_30_min_list[share], 3).value  # LOW

    else:
        new_30_min_sheet.cell(7, 6).value = cashHL_sheet.cell(cash_30_min_list[share], 7).value  # 9:25 cl
        new_30_min_sheet.cell(7, 7).value = cashHL_sheet.cell(cash_30_min_list[share], 2).value  # HIGH
        new_30_min_sheet.cell(7, 8).value = cashHL_sheet.cell(cash_30_min_list[share], 3).value  # LOW

    idx += 1

    cash_30_min_wb.save(path)
    os.remove(xls_path)

ltp_wb.save(rf'C:\Users\admin\PycharmProjects\daily data\LTP PREV.xlsx')
