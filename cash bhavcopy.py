import openpyxl as xl
import pandas as pd
from xls2xlsx import XLS2XLSX
from date_variables import date, mnth, yr
from zipfile import ZipFile

md_path_zipped = rf"E:\chrome downloads\cm{date[:2]}{mnth}20{date[6:]}bhav.csv.zip"     # .zip file path of downloaded cash bhavcpoy
md_path = rf"E:\chrome downloads"

# extracting .zip file
# with ZipFile(md_path_zipped, 'r') as zObject:
#     zObject.extractall(path=md_path)

# list of shares we want to check
tgt_shares = ['AARTIIND', 'ABB', 'ABCAPITAL', 'ABFRL', 'ADANIENT', 'ADANIPORTS', 'ALKEM', 'AMARAJABAT', 'AMBUJACEM',
              'APLLTD', 'APOLLOHOSP', 'APOLLOTYRE', 'ASHOKLEY', 'ASTRAL', 'ATUL', 'AUBANK', 'AUROPHARMA', 'BAJAJFINSV',
              'BAJFINANCE', 'BALKRISIND', 'BALRAMCHIN', 'BANDHANBNK', 'BANKBARODA', 'BATAINDIA', 'BEL', 'BHARATFORG',
              'BIOCON', 'BRITANNIA', 'BSOFT', 'CANBK', 'CANFINHOME', 'CHAMBLFERT', 'CHOLAFIN', 'CIPLA', 'COFORGE',
              'CONCOR', 'COROMANDEL', 'CROMPTON', 'CUMMINSIND', 'DABUR', 'DALBHARAT', 'DEEPAKNTR', 'DELTACORP',
              'DIVISLAB', 'DIXON', 'DLF', 'DRREDDY', 'ESCORTS', 'EXIDEIND', 'GLENMARK', 'GLS', 'GNFC', 'GODREJCP',
              'GODREJPROP', 'GRANULES', 'GRASIM', 'GUJGASLTD', 'HAL', 'HAVELLS', 'HCLTECH', 'HDFCAMC', 'HDFCLIFE',
              'HINDALCO', 'HINDCOPPER', 'ICICIGI', 'ICICIPRULI', 'IEX', 'IGL', 'INDHOTEL', 'INDIACEM', 'INDIAMART',
              'INDIGO', 'INDUSINDBK', 'INDUSTOWER', 'INTELLECT', 'IPCALAB', 'JINDALSTEL', 'JKCEMENT', 'JSWSTEEL',
              'JUBLFOOD', 'KOTAKBANK', 'LALPATHLAB', 'LAURUSLABS', 'LICHSGFIN', 'LTIM', 'LTTS', 'LUPIN',
              'M&MFIN', 'MANAPPURAM', 'MARICO', 'MCDOWELL-N', 'MCX', 'METROPOLIS', 'MFSL', 'MGL', 'MPHASIS',
              'MUTHOOTFIN', 'NAM-INDIA', 'NAUKRI', 'NAVINFLUOR', 'NMDC', 'NTPC', 'OBEROIRLTY', 'PEL', 'PERSISTENT',
              'PETRONET', 'PIDILITIND', 'POLYCAB', 'POWERGRID', 'RAIN', 'RAMCOCEM', 'RBLBANK', 'RECLTD',
              'SBICARD', 'SBILIFE', 'SIEMENS', 'SRF', 'STAR', 'SUNPHARMA', 'SYNGENE', 'TATACOMM', 'TECHM',
              'TORNTPHARM', 'TORNTPOWER', 'TRENT', 'TVSMOTOR', 'UBL', 'ULTRACEMCO', 'UPL', 'VEDL', 'VOLTAS', 'ZEEL',
              'ZYDUSLIFE']

# creating output workbook to save the data
op_wb = xl.Workbook()
op_sheet = op_wb.active
row = 2     # output workbook starting row

df = pd.read_csv(md_path_zipped[:len(md_path_zipped)-4])

# Fixed headings to be put in output sheet
op_sheet.cell(1, 1).value = "Share Name"
op_sheet.cell(1, 2).value = "High Value"
op_sheet.cell(1, 3).value = "Low Value"

op_sheet.cell(1, 5).value = "Buy Entry"
op_sheet.cell(1, 6).value = "Buy Target"
op_sheet.cell(1, 7).value = "Buy Stoploss"
op_sheet.cell(1, 8).value = "Buy Quantity"

op_sheet.cell(1, 10).value = "Sell Entry"
op_sheet.cell(1, 11).value = "Sell Target"
op_sheet.cell(1, 12).value = "Sell Stoploss"
op_sheet.cell(1, 13).value = "Sell Quantity"

for i in range(len(df)):
    share = str(df.iloc[i, 0])           # share name
    high = df.iloc[i, 3]            # high value
    low = df.iloc[i, 4]             # low value
    symbol = df.iloc[i, 1]          # symbol value

    if share in tgt_shares and symbol == 'EQ':
        op_sheet.cell(row, 1).value = share
        op_sheet.cell(row, 2).value = high
        op_sheet.cell(row, 3).value = low

        # adding trade entry prices and quantities
        b_entry = high + (high * 0.01)
        b_tgt = b_entry + (b_entry * 0.015)
        b_sl = b_entry - (b_entry * 0.0125)
        b_qty = 100000 // b_entry

        s_entry = low - (low * 0.01)
        s_tgt = s_entry - (s_entry * 0.015)
        s_sl = s_entry + (s_entry * 0.0125)
        s_qty = 100000 // s_entry

        # BUY
        op_sheet.cell(row, 5).value = b_entry           # Buy Entry
        op_sheet.cell(row, 6).value = b_tgt             # Buy Target
        op_sheet.cell(row, 7).value = b_sl              # Buy Stoploss
        op_sheet.cell(row, 8).value = b_qty             # Buy Quantity

        # SELL
        op_sheet.cell(row, 10).value = s_entry           # Sell Entry
        op_sheet.cell(row, 11).value = s_tgt             # Sell Target
        op_sheet.cell(row, 12).value = s_sl              # Sell Stoploss
        op_sheet.cell(row, 13).value = s_qty             # Sell Quantity

        row += 1

op_wb.save(rf"E:\Daily Data work\Trading Algorithm\{yr}\{mnth}\{date}\{date[:2]}{mnth}20{date[6:]}algo.xlsx")        # saving our file in Daily Data work


# next
# create a new sheet called 'actual trades', rename old sheet to 'possible trades' now go through day 1 file and compare if
# we entered a buy or a sell. then see if where we exited and with what profit/ loss. now also add what would happen if we had stayed till end and
# if we cut at max profit


