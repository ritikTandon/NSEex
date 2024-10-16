import shutil

import openpyxl as xl
import os
import requests
from openpyxl.styles import PatternFill
import datetime

from xls2xlsx import XLS2XLSX

from date_variables import date, mnth, yr
from time import sleep
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException

cash_share_list = ["AARTIIND", "ADANI", "APOLLO", "BAJFINSV", "BAJFIN", "BANBK", "BARODA", "COALIND", "DLF", "EICHER",
                   "FEDBANK", "HCL", "HDFC", "HIND", "ICICI", "INDUSIND", "INFY", "JIND", "LIC", "M&M", "M&MFIN", "NTPC",
                   "REL", "SBIN", "SUNTV", "TCHEM", "TM", "TP", "TS", "ULTRA"]

# cash_share_list = ["ADANI"]   # list to test few shares after making code changes

cashHL_wb = xl.load_workbook(r'C:\Users\admin\PycharmProjects\daily data\cash high low.xlsx')
cashHL_sheet = cashHL_wb['Sheet1']
cashHL_row = 2

# converting xls to xlsx for csh sheet
x2x = XLS2XLSX(r'E:\Daily Data work\csh.xls')
wb = x2x.to_xlsx()
wb.save(r'E:\Daily Data work\csh.xlsx')

csh_wb = xl.load_workbook(r'E:\Daily Data work\csh.xlsx')
csh_sheet = csh_wb['csh-Sheet1']
csh_row = 2

fl_9_25 = 0.392361111       # 9:25 time value in general format

# copying hourlys (.xls) as backup
# path to source directory
src_dir = rf"E:\Daily Data work\hourlys 1 minute CASH\{yr}\{mnth}\{date}"

# path to destination directory
dest_dir = rf"C:\Users\admin\PycharmProjects\daily data\Daily Backup hourlys\1 min csh"

# getting all the files in the source directory
src_files = os.listdir(src_dir)
for file_name in src_files:
    full_file_name = os.path.join(src_dir, file_name)
    if os.path.isfile(full_file_name):
        shutil.copy(full_file_name, dest_dir)

print("Files copied as backup!")

for share in cash_share_list:
    # converting .xls shares to .xlsx
    path = rf"E:\Daily Data work\hourlys 1 minute CASH\{yr}\{mnth}\{date}\{share}.xlsx"
    xls_path = rf"E:\Daily Data work\hourlys 1 minute CASH\{yr}\{mnth}\{date}\{share}.xls"
    x2x = XLS2XLSX(xls_path)

    wb = x2x.to_xlsx()
    sheet = wb[f"{share}-Sheet1"]

    start_row = 2
    time_cell = sheet.cell(start_row, 7)

    # incrementing starting row till we reach 9:25 am row
    while time_cell.value < fl_9_25:
        start_row += 1
        time_cell = sheet.cell(start_row, 7)

    print(f"starting row is {start_row}")

    start_row_2 = start_row

    # loop for changing the time cells format (have to close and reopen otherwise it doesn't change format)
    while time_cell.value is not None:
        time_cell = sheet.cell(start_row, 7)
        time_cell.number_format = 'hh:mm AM/PM'

        start_row += 1

    wb.save(path)

    # reloading the worksheet
    wb = xl.load_workbook(path)
    sheet = wb[f"{share}-Sheet1"]

    start_row = start_row_2

    time_cell = sheet.cell(start_row, 7)
    cur_time = time_cell.value
    end_time = datetime.time(15, 30, 0)

    # 9:25 close value
    cl_9_25 = sheet.cell(start_row, 3).value
    sheet.cell(start_row, 3).fill = PatternFill("solid", 'FFFF00')

    # reloading wb otherwise pattern fill doesn't work
    wb.save(path)
    wb = xl.load_workbook(path)
    sheet = wb[f"{share}-Sheet1"]

    HIGH = 0
    LOW = 9999999

    # HIGH and LOW value finding loop
    while cur_time and cur_time <= end_time:
        time_cell = sheet.cell(start_row, 7)
        high_cell = sheet.cell(start_row, 4)
        low_cell = sheet.cell(start_row, 5)

        cur_time = time_cell.value

        if high_cell.value is not None and high_cell.value > HIGH:
            HIGH = high_cell.value

        if low_cell.value is not None and low_cell.value < LOW and low_cell.value != 0:
            LOW = low_cell.value

        start_row += 1

    # high
    cashHL_sheet.cell(cashHL_row, 2).value = HIGH

    # low
    cashHL_sheet.cell(cashHL_row, 3).value = LOW

    # LTP
    cashHL_sheet.cell(cashHL_row, 5).value = csh_sheet.cell(csh_row, 2).value

    # vol
    volume = csh_sheet.cell(csh_row, 5).value
    volume //= 100000

    cashHL_sheet.cell(cashHL_row, 6).value = volume

    # 9:25 close
    cashHL_sheet.cell(cashHL_row, 7).value = cl_9_25

    cashHL_row += 1
    csh_row += 1

    HIGH = 0
    LOW = 9999999

    # 30 MIN FORMATTING IN 1 MIN SHEETS
    sheet.cell(1, 14).value = "HIGH"
    sheet.cell(1, 15).value = "LOW"
    sheet.cell(1, 16).value = "CLOSE"

    start_row = start_row_2     # actual start row

    time_cell = sheet.cell(start_row, 7)
    cur_time = time_cell.value

    count = 0

    while cur_time is not None and cur_time <= end_time:
        high_cell = sheet.cell(start_row, 4)
        low_cell = sheet.cell(start_row, 5)

        if high_cell.value is not None and high_cell.value > HIGH:
            HIGH = high_cell.value

        if low_cell.value is not None and low_cell.value < LOW and low_cell.value != 0:
            LOW = low_cell.value

        # resetting after 30 mins
        if count == 30:
            sheet.cell(start_row, 14).value = HIGH
            sheet.cell(start_row, 15).value = LOW

            # if 30 min close is empty or 0
            if sheet.cell(start_row, 3).value == 0 or sheet.cell(start_row, 3).value is None:
                temp_row = start_row

                # iterating backwards till we find a close value
                while sheet.cell(temp_row, 3).value == 0 or sheet.cell(temp_row, 3).value is None:
                    temp_row -= 1

                sheet.cell(start_row, 16).value = sheet.cell(temp_row, 3).value  # close

            else:
                sheet.cell(start_row, 16).value = sheet.cell(start_row, 3).value  # close

            count = 1
            HIGH = 0
            LOW = 9999999
            start_row += 1
            continue

        start_row += 1
        count += 1

        time_cell = sheet.cell(start_row, 7)
        cur_time = time_cell.value

    # last any left aggregate (< 30 mins)
    sheet.cell(start_row-1, 14).value = HIGH
    sheet.cell(start_row-1, 15).value = LOW
    sheet.cell(start_row-1, 16).value = sheet.cell(start_row-1, 3).value  # close

    sheet.freeze_panes = sheet["A2"]

    wb.save(path)
    os.remove(xls_path)

    print(f"{share} done")

# for close filling
options = Options()
options.add_argument("--disable-blink-features=AutomationControlled")

# Exclude the collection of enable-automation switches
options.add_experimental_option("excludeSwitches", ["enable-automation"])

# Turn-off userAutomationExtension
options.add_experimental_option("useAutomationExtension", False)

cash_close_list = ["AARTIIND", "ADANIENT", "APOLLOTYRE", "BAJAJFINSV", "BAJFINANCE", "BANDHANBNK", "BANKBARODA", "COALINDIA",
                   "DLF", "EICHERMOT", "FEDERALBNK", "HCLTECH", "HDFCBANK", "HINDALCO", "ICICIBANK", "INDUSINDBK", "INFY",
                   "JINDALSTEL", "LICHSGFIN", "M%26M", "M%26MFIN", "NTPC", "RELIANCE", "SBIN", "SUNTV", "TATACHEM", "TATAMOTORS",
                   "TATAPOWER", "TATASTEEL", "ULTRACEMCO"]

# cash_close_list1 = ["M%26M"]

manual = []         # list to keep track of the shares whose values selenium couldn't get
close = []

for share in cash_close_list:
    driver = webdriver.Chrome(options=options)

    driver.get(f"https://www.nseindia.com/get-quotes/equity?symbol={share}")

    try:
        sleep(2)
        myElem = WebDriverWait(driver, 20).until(ec.presence_of_element_located((By.ID, 'quoteLtp')))
        # sleep(5)
        close_val = driver.find_element(By.ID, "quoteLtp").text

        while close_val == '':
            driver.refresh()
            WebDriverWait(driver, 10).until(ec.presence_of_element_located((By.ID, 'quoteLtp')))
            close_val = driver.find_element(By.ID, "quoteLtp").text
            sleep(0.5)

        close_val = close_val.replace(",", "")

        # truncating last 0
        if close_val[len(close_val)-1:len(close_val)] == '0':
            close_val = close_val[:len(close_val)-1]

        close.append(close_val)

        print(f'{share}: {close_val}')
        if close_val == '':
            manual.append(share)

    except TimeoutException:        # added temp fix for timeoutexception, will need to check if it works properly or nah
        close.append('')
        manual.append(share)
        print(f"Loading took too much time for {share}!")

    driver.close()

print(close)
print(manual)

i = 0

while i < len(cash_close_list):
    close_cell = cashHL_sheet.cell(i+2, 4)

    if close[i] == '':
        close_cell.value = 0

    else:
        close_cell.value = float(close[i])
        close_cell.number_format = "0.00"

    i += 1

cashHL_wb.save(r'C:\Users\admin\PycharmProjects\daily data\cash high low.xlsx')
