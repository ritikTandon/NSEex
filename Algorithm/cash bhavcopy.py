import openpyxl as xl
import datetime
import pandas as pd
from win32com.client import Dispatch
from openpyxl.styles import Font, Alignment
from date_variables import date, mnth, yr
from zipfile import ZipFile
from decimal import Decimal, ROUND_UP


# up and down rounding functions (to nearest 0.05)
def round_up(var):
    x = Decimal(str(var))

    return float((x * 2).quantize(Decimal('.5'), rounding=ROUND_UP) / 2)


def round_down(var):
    x = Decimal(str(var))

    return float((x * 2).quantize(Decimal('.5')) / 2)


# reading previous dates from date.txt
with open("date.txt", "r") as file:
    data = file.readlines()
    old_date = rf"{data[0][:-1]}"
    old_mnth = rf"{data[1][:-1]}"
    old_yr = rf"{data[2]}"

md_path_zipped = rf"E:\chrome downloads\cm{date[:2]}{mnth}20{date[6:]}bhav.csv.zip"     # .zip file path of downloaded cash bhavcpoy
md_path = rf"E:\chrome downloads"

# extracting .zip file
with ZipFile(md_path_zipped, 'r') as zObject:
    zObject.extractall(path=md_path)

# list of shares we want to check
tgt_shares = ['AARTIIND', 'ABB', 'ABCAPITAL', 'ABFRL', 'ADANIENT', 'ADANIPORTS', 'ALKEM', 'AMBUJACEM',
               'APOLLOHOSP', 'APOLLOTYRE', 'ASHOKLEY', 'ASTRAL', 'ATUL', 'AUBANK', 'AUROPHARMA', 'BAJAJFINSV',
              'BAJFINANCE', 'BALKRISIND', 'BALRAMCHIN', 'BANDHANBNK', 'BANKBARODA', 'BATAINDIA', 'BEL', 'BHARATFORG',
              'BIOCON', 'BRITANNIA', 'BSOFT', 'CANBK', 'CANFINHOME', 'CHAMBLFERT', 'CHOLAFIN', 'CIPLA', 'COFORGE',
              'CONCOR', 'COROMANDEL', 'CROMPTON', 'CUMMINSIND', 'DABUR', 'DALBHARAT', 'DEEPAKFERT', 'DEEPAKNTR', 'DELTACORP',
              'DIVISLAB', 'DIXON', 'DLF', 'DRREDDY', 'ESCORTS', 'EXIDEIND', 'GLENMARK', 'GLS', 'GNFC', 'GODREJCP',
              'GODREJPROP', 'GRANULES', 'GRASIM', 'GUJGASLTD', 'HAL', 'HAVELLS', 'HCLTECH', 'HDFCAMC', 'HDFCLIFE',
              'HINDALCO', 'HINDCOPPER', 'ICICIGI', 'ICICIPRULI', 'IEX', 'IGL', 'INDHOTEL', 'INDIACEM', 'INDIAMART',
              'INDIGO', 'INDUSINDBK', 'INDUSTOWER', 'INTELLECT', 'IPCALAB', 'JINDALSTEL', 'JKCEMENT', 'JSWSTEEL',
              'JUBLFOOD', 'KOTAKBANK', 'LALPATHLAB', 'LAURUSLABS', 'LICHSGFIN', 'LTIM', 'LTTS', 'LUPIN',
              'M&MFIN', 'MANAPPURAM', 'MARICO', 'MCDOWELL-N', 'MCX', 'METROPOLIS', 'MFSL', 'MGL', 'MPHASIS',
              'MUTHOOTFIN', 'NAM-INDIA', 'NAUKRI', 'NAVINFLUOR', 'NMDC', 'NTPC', 'OBEROIRLTY', 'PEL', 'PERSISTENT',
              'PETRONET', 'PIDILITIND', 'POLYCAB', 'POWERGRID', 'RAIN', 'RAMCOCEM', 'RBLBANK', 'RECLTD',
              'SBICARD', 'SBILIFE', 'SIEMENS', 'SRF', 'STAR', 'SUNPHARMA', 'SYNGENE', 'TATACOMM', 'TATAMOTORS', 'TECHM',
              'TORNTPHARM', 'TORNTPOWER', 'TRENT', 'TVSMOTOR', 'UBL', 'ULTRACEMCO', 'UPL', 'VEDL', 'VOLTAS', 'ZEEL',
              'ZYDUSLIFE']

# creating output workbook to save the data
op_wb = xl.Workbook()

del op_wb['Sheet']          # deleting default sheet

all_tr_sheet = op_wb.create_sheet("All Trades")
actual_tr_sheet = op_wb.create_sheet("Actual Trades")

row = 2     # output workbook starting row

df = pd.read_csv(md_path_zipped[:-4])

# Fixed headings to be put in all_tr_sheet sheet
all_tr_sheet.cell(1, 1).value = "Share Name"
all_tr_sheet.cell(1, 2).value = "High Value"
all_tr_sheet.cell(1, 3).value = "Low Value"

all_tr_sheet.cell(1, 5).value = "Buy Entry"
all_tr_sheet.cell(1, 6).value = "Buy Target"
all_tr_sheet.cell(1, 7).value = "Buy Stoploss"
all_tr_sheet.cell(1, 8).value = "Buy Quantity"
all_tr_sheet.cell(1, 9).value = "Buy Amount"

all_tr_sheet.cell(1, 11).value = "Sell Entry"
all_tr_sheet.cell(1, 12).value = "Sell Target"
all_tr_sheet.cell(1, 13).value = "Sell Stoploss"
all_tr_sheet.cell(1, 14).value = "Sell Quantity"
all_tr_sheet.cell(1, 15).value = "Sell Amount"
all_tr_sheet.cell(1, 17).value = "3:30 Price"


# # Fixed headings to be put in actual_tr_sheet sheet
# actual_tr_sheet.cell(1, 1).value = "Share Name"
# actual_tr_sheet.cell(1, 2).value = "New High Value"
# actual_tr_sheet.cell(1, 3).value = "New Low Value"
#
# actual_tr_sheet.cell(1, 5).value = "Entry Price"
# actual_tr_sheet.cell(1, 6).value = "Trade Type"
# actual_tr_sheet.cell(1, 7).value = "Target Price"
# actual_tr_sheet.cell(1, 8).value = "Stoploss Price"
# actual_tr_sheet.cell(1, 9).value = "3:30 Price"
# actual_tr_sheet.cell(1, 10).value = "Exit P&L"
#
# actual_tr_sheet.cell(1, 12).value = "Peak Exit P&L"
#
# actual_tr_sheet.cell(1, 14).value = "3:30 Exit P&L"

for i in range(len(df)):
    share = str(df.iloc[i, 0])          # share name
    high = df.iloc[i, 3]                # high value
    low = df.iloc[i, 4]                 # low value
    symbol = df.iloc[i, 1]              # symbol value
    val_330 = df.iloc[i, 6]             # last value

    if share in tgt_shares and symbol == 'EQ':
        all_tr_sheet.cell(row, 1).value = share
        all_tr_sheet.cell(row, 2).value = high
        all_tr_sheet.cell(row, 3).value = low

        # adding trade entry prices and quantities
        b_entry = round_up((high + (high * 0.01)))
        b_tgt = round_down((b_entry + (b_entry * 0.015)))
        b_sl = round_up((b_entry - (b_entry * 0.0125)))
        # b_qty = 100000 // b_entry

        s_entry = round_down((low - (low * 0.01)))
        s_tgt = round_up((s_entry - (s_entry * 0.015)))
        s_sl = round_down((s_entry + (s_entry * 0.0125)))
        # s_qty = 100000 // s_entry

        # BUY
        all_tr_sheet.cell(row, 5).value = b_entry           # Buy Entry
        all_tr_sheet.cell(row, 6).value = b_tgt             # Buy Target
        all_tr_sheet.cell(row, 7).value = b_sl              # Buy Stoploss
        all_tr_sheet.cell(row, 8).value = f'=INT(I{row}/E{row})'             # Buy Quantity
        all_tr_sheet.cell(row, 9).value = 100000            # Buy Amount (100000 default)

        # SELL
        all_tr_sheet.cell(row, 11).value = s_entry           # Sell Entry
        all_tr_sheet.cell(row, 12).value = s_tgt             # Sell Target
        all_tr_sheet.cell(row, 13).value = s_sl              # Sell Stoploss
        all_tr_sheet.cell(row, 14).value = f'=INT(O{row}/K{row})'             # Sell Quantity
        all_tr_sheet.cell(row, 15).value = 100000            # Sell Amount (100000 default)

        # 3:30 Value
        all_tr_sheet.cell(row, 17).value = val_330           # LTP/ 3:30 close value

        row += 1


# formatting file
for i in range(200):
    for j in range(20):
        all_tr_sheet.cell(i + 1, j + 1).font = Font('Arial', 12, bold=True)
        actual_tr_sheet.cell(i + 1, j + 1).font = Font('Arial', 12, bold=True)

# fixing top heading row to be always visible
all_tr_sheet.freeze_panes = all_tr_sheet['A2']
actual_tr_sheet.freeze_panes = actual_tr_sheet['A2']

op_wb.save(rf"E:\Daily Data work\Trading Algorithm\{yr}\{mnth}\{date[:2]}{mnth}20{date[6:]}algo.xlsx")        # saving our file in Daily Data work


# auto-fitting columns
excel = Dispatch('Excel.Application')
wb = excel.Workbooks.Open(rf"E:\Daily Data work\Trading Algorithm\{yr}\{mnth}\{date[:2]}{mnth}20{date[6:]}algo.xlsx")

# Activating sheets
excel.Worksheets(1).Activate()
excel.ActiveSheet.Columns.AutoFit()
excel.Worksheets(2).Activate()
excel.ActiveSheet.Columns.AutoFit()

# saving
wb.Save()
wb.Close()

# opening previous day algo workbook, create actual happened trades, go through each 1 min file of algo of previous day and check if buy/sell happened
# if either happened, check if tgt or sl and if neither tgt nor sl, exit on the time papa said. Save these in similar format
#
# old all trades to get buy entry, sl tgt etc. and to save actual trades that happened, all from next day 1 min algo

# # loading yesterday's workbook
old_wb = xl.load_workbook(rf"E:\Daily Data work\Trading Algorithm\{old_yr}\{old_mnth}\{old_date[:2]}{old_mnth}20{old_date[6:]}algo.xlsx")
old_all_tr_sheet = old_wb["All Trades"]
old_actual_tr_sheet = old_wb["Actual Trades"]

old_all_row = 2                     # row for previous day's workbook's 'All Trades' sheet
# old_actual_row = 2              # row for previous day's workbook's 'Actual Trades' sheet


# headings
old_actual_tr_sheet.cell(1, 1).value = "Share Name"
old_actual_tr_sheet.cell(1, 2).value = "Trade Type"
old_actual_tr_sheet.cell(1, 3).value = "Exit Type"
old_actual_tr_sheet.cell(1, 4).value = "Profit/Loss Points"
old_actual_tr_sheet.cell(1, 5).value = "Capital"
old_actual_tr_sheet.cell(1, 6).value = "Profit/Loss Amount"

old_actual_tr_sheet.cell(1, 8).value = "3:15 Profit/Loss Points"
old_actual_tr_sheet.cell(1, 9).value = "3:15 Profit/Loss Amount"

old_actual_tr_sheet.cell(1, 11).value = "Trade Count"

old_actual_tr_sheet.cell(1, 14).value = "BUY"
old_actual_tr_sheet.cell(1, 15).value = "SELL"
old_actual_tr_sheet.cell(1, 17).value = "Trade Period"
old_actual_tr_sheet.cell(1, 18).value = "Holding Time"
old_actual_tr_sheet.cell(2, 13).value = "Target"
old_actual_tr_sheet.cell(3, 13).value = "Stoploss"
old_actual_tr_sheet.cell(4, 13).value = "3:15 exit"

old_actual_tr_sheet.cell(9, 13).value = "Actual Profit"
old_actual_tr_sheet.cell(10, 13).value = "3:15 exit Profit"


buy_tgt_count = 0
sell_tgt_count = 0
buy_sl_count = 0
sell_sl_count = 0
buy_315_count = 0
sell_315_count = 0

while old_all_row < 135:       # todo WILL NEED TO MAKE IT DYNAMIC OR CHANGE WHEN ADDING NEW SHARES
    inTrade = False
    completed = False
    trade_type = ''
    exit_type = ''  # exit on target, stoploss or 3:15 pm
    profit_pts = 0
    entry_time = None
    exit_time = None

    # fetching data
    share_name = old_all_tr_sheet.cell(old_all_row, 1).value

    # buy
    old_buy_entry = old_all_tr_sheet.cell(old_all_row, 5).value             # old buy entry
    buy_tgt = old_all_tr_sheet.cell(old_all_row, 6).value                   # buy target
    buy_sl = old_all_tr_sheet.cell(old_all_row, 7).value                    # buy stoploss

    # sell
    old_sell_entry = old_all_tr_sheet.cell(old_all_row, 11).value           # old sell entry
    sell_tgt = old_all_tr_sheet.cell(old_all_row, 12).value                 # sell target
    sell_sl = old_all_tr_sheet.cell(old_all_row, 13).value                  # sell stoploss

    # loading current 1 min sheet
    path = rf"E:\Daily Data work\hourlys 1 minute ALGO\{yr}\{mnth}\{date}\{share_name}.xlsx"
    wb_1_min = xl.load_workbook(path)
    sheet_1_min = wb_1_min[f'{share_name}-Sheet1']

    min_1_start_row = 2
    time = sheet_1_min.cell(min_1_start_row, 7).value

    # starting from 9:21 am
    while time < datetime.time(9, 21, 0):
        min_1_start_row += 1
        time = sheet_1_min.cell(min_1_start_row, 7).value

    while time <= datetime.time(15, 15, 0):         # time <= 3:15
        # high, low and time
        high = sheet_1_min.cell(min_1_start_row, 4).value
        low = sheet_1_min.cell(min_1_start_row, 5).value
        time = sheet_1_min.cell(min_1_start_row, 7).value

    #     # checking trade entry conditions

        if not inTrade and not completed and time <= datetime.time(14, 55, 0):      # no trade entry after 2:55 pm
            if high >= old_buy_entry >= low != 0:
                trade_type = 'B'
                inTrade = True
                entry_time = time
                min_1_start_row += 1
                continue
            elif high >= old_sell_entry >= low != 0:
                trade_type = 'S'
                inTrade = True
                entry_time = time
                min_1_start_row += 1
                continue

        if inTrade and not completed:
            if trade_type == 'B':
                if high >= buy_tgt >= low != 0:
                    exit_type = 'tgt'
                    inTrade = False
                    exit_time = time
                    completed = True
                elif high >= buy_sl >= low != 0:
                    exit_type = 'sl'
                    inTrade = False
                    exit_time = time
                    completed = True

            if trade_type == 'S':
                if high >= sell_tgt >= low != 0:
                    exit_type = 'tgt'
                    inTrade = False
                    exit_time = time
                    completed = True
                elif high >= sell_sl >= low != 0:
                    exit_type = 'sl'
                    completed = True
                    exit_time = time
                    inTrade = False

        min_1_start_row += 1

    # saving 3:15 pm close value
    close = sheet_1_min.cell(min_1_start_row-1, 3).value
    print(f"closing time of {share_name} is {time}")

    if trade_type == 'B':
        if inTrade:
            exit_type = '3:15'
            profit_pts = close - old_buy_entry
            buy_315_count += 1
        else:
            if exit_type == 'tgt':
                profit_pts = buy_tgt - old_buy_entry
                buy_tgt_count += 1
            elif exit_type == 'sl':
                profit_pts = buy_sl - old_buy_entry
                buy_sl_count += 1

    elif trade_type == 'S':
        if inTrade:
            exit_type = '3:15'
            profit_pts = old_sell_entry - close
            sell_315_count += 1
        else:
            if exit_type == 'tgt':
                profit_pts = old_sell_entry - sell_tgt
                sell_tgt_count += 1
            elif exit_type == 'sl':
                profit_pts = old_sell_entry - sell_sl
                sell_sl_count += 1

    else:
        trade_type = 'None'

    # filling in values of trades
    old_actual_tr_sheet.cell(old_all_row, 1).value = share_name

    if trade_type == "None":
        old_actual_tr_sheet.cell(old_all_row, 2).value = trade_type
        old_all_row += 1
        continue
    else:
        old_actual_tr_sheet.cell(old_all_row, 2).value = trade_type
        old_actual_tr_sheet.cell(old_all_row, 3).value = exit_type
        old_actual_tr_sheet.cell(old_all_row, 4).value = profit_pts
        old_actual_tr_sheet.cell(old_all_row, 5).value = 100000

        if trade_type == 'B':
            old_actual_tr_sheet.cell(old_all_row, 6).value = f"=INT('Actual Trades'!E{old_all_row}/'All Trades'!E{old_all_row})*'Actual Trades'!D{old_all_row}"
            old_actual_tr_sheet.cell(old_all_row, 8).value = close - old_buy_entry
            old_actual_tr_sheet.cell(old_all_row, 9).value = f"=INT('Actual Trades'!E{old_all_row}/'All Trades'!E{old_all_row})*'Actual Trades'!H{old_all_row}"

        elif trade_type == 'S':
            old_actual_tr_sheet.cell(old_all_row, 6).value = f"=INT('Actual Trades'!E{old_all_row}/'All Trades'!K{old_all_row})*'Actual Trades'!D{old_all_row}"
            old_actual_tr_sheet.cell(old_all_row, 8).value = old_sell_entry - close
            old_actual_tr_sheet.cell(old_all_row, 9).value = f"=INT('Actual Trades'!E{old_all_row}/'All Trades'!K{old_all_row})*'Actual Trades'!H{old_all_row}"

        # summary
        old_actual_tr_sheet.cell(2, 11).value = buy_tgt_count + buy_sl_count + buy_315_count + sell_tgt_count + sell_sl_count + sell_315_count
        old_actual_tr_sheet.cell(2, 14).value = buy_tgt_count
        old_actual_tr_sheet.cell(3, 14).value = buy_sl_count
        old_actual_tr_sheet.cell(4, 14).value = buy_315_count

        old_actual_tr_sheet.cell(2, 15).value = sell_tgt_count
        old_actual_tr_sheet.cell(3, 15).value = sell_sl_count
        old_actual_tr_sheet.cell(4, 15).value = sell_315_count

        # total profits
        old_actual_tr_sheet.cell(9, 15).value = "=sum(F2:F200)"
        old_actual_tr_sheet.cell(9, 15).number_format = '0'
        old_actual_tr_sheet.cell(10, 15).value = "=sum(I2:I200)"
        old_actual_tr_sheet.cell(10, 15).number_format = '0'

        # entry, exit times and holding time
        if exit_time is None:
            exit_time = datetime.time(15, 15)

        # if share_name == "APOLLOHOSP":
        #     print("er")

        # converting to datetime.datetime for easier subtraction
        entry_time = datetime.datetime.combine(datetime.datetime.today(), entry_time)
        exit_time = datetime.datetime.combine(datetime.datetime.today(), exit_time)

        old_actual_tr_sheet.cell(old_all_row, 17).value = f"{entry_time.time().strftime('%I:%M %p')}-{exit_time.time().strftime('%I:%M %p')}"

        time_difference = exit_time - entry_time

        # Get hours and minutes from the time difference
        hours = time_difference.seconds // 3600
        minutes = (time_difference.seconds % 3600) // 60

        old_actual_tr_sheet.cell(old_all_row, 18).value = f"{hours}h-{minutes}m"

    old_all_row += 1

# finally saving the old workbook
old_wb.save(rf"E:\Daily Data work\Trading Algorithm\{old_yr}\{old_mnth}\{old_date[:2]}{old_mnth}20{old_date[6:]}algo.xlsx")

# saving current date in file
with open('date.txt', 'w') as file:
    file.write(f"{date}\n{mnth}\n{yr}")


# auto fitting excel files
excel = Dispatch('Excel.Application')
wb = excel.Workbooks.Open(rf"E:\Daily Data work\Trading Algorithm\{old_yr}\{old_mnth}\{old_date[:2]}{old_mnth}20{old_date[6:]}algo.xlsx")

# Activating sheets
excel.Worksheets(1).Activate()
excel.ActiveSheet.Columns.AutoFit()
excel.Worksheets(2).Activate()
excel.ActiveSheet.Columns.AutoFit()

# saving
wb.Save()
wb.Close()



