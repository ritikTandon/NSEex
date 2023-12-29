import datetime
from datetime import timedelta
import calendar
import openpyxl as xl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension


algo_share_list = ['AARTIIND', 'ABB', 'ABCAPITAL', 'ABFRL', 'ADANIENT', 'ADANIPORTS', 'ALKEM', 'AMBUJACEM',
                     'APOLLOHOSP', 'APOLLOTYRE', 'ASHOKLEY', 'ASTRAL', 'ATUL', 'AUBANK', 'AUROPHARMA', 'BAJAJFINSV',
                     'BAJFINANCE', 'BALKRISIND', 'BALRAMCHIN', 'BANDHANBNK', 'BANKBARODA', 'BATAINDIA', 'BEL',
                     'BHARATFORG', 'BIOCON', 'BRITANNIA', 'BSOFT', 'CANBK', 'CANFINHOME', 'CHAMBLFERT', 'CHOLAFIN',
                     'CIPLA', 'COFORGE', 'CONCOR', 'COROMANDEL', 'CROMPTON', 'CUMMINSIND', 'DABUR', 'DALBHARAT',
                     'DEEPAKFERT', 'DEEPAKNTR', 'DELTACORP', 'DIVISLAB', 'DIXON', 'DLF', 'DRREDDY', 'ESCORTS',
                     'EXIDEIND', 'GLENMARK', 'GLS', 'GNFC', 'GODREJCP', 'GODREJPROP', 'GRANULES', 'GRASIM', 'GUJGASLTD',
                     'HAL', 'HAVELLS', 'HCLTECH', 'HDFCAMC', 'HDFCLIFE', 'HINDALCO', 'HINDCOPPER', 'ICICIGI',
                     'ICICIPRULI', 'IEX', 'IGL', 'INDHOTEL', 'INDIACEM', 'INDIAMART', 'INDIGO', 'INDUSINDBK',
                     'INDUSTOWER', 'INTELLECT', 'IPCALAB', 'JINDALSTEL', 'JKCEMENT', 'JSWSTEEL', 'JUBLFOOD',
                     'KOTAKBANK', 'LALPATHLAB', 'LAURUSLABS', 'LICHSGFIN', 'LTIM', 'LTTS', 'LUPIN', 'M%26MFIN',
                     'MANAPPURAM', 'MARICO', 'MCDOWELL-N', 'MCX', 'METROPOLIS', 'MFSL', 'MGL', 'MPHASIS', 'MUTHOOTFIN',
                     'NAM-INDIA', 'NAUKRI', 'NAVINFLUOR', 'NMDC', 'NTPC', 'OBEROIRLTY', 'PEL', 'PERSISTENT', 'PETRONET',
                     'PIDILITIND', 'POLYCAB', 'POWERGRID', 'RAIN', 'RAMCOCEM', 'RBLBANK', 'RECLTD', 'SBICARD',
                     'SBILIFE', 'SIEMENS', 'SRF', 'STAR', 'SUNPHARMA', 'SYNGENE', 'TATACOMM', 'TATAMOTORS', 'TECHM',
                     'TORNTPHARM', 'TORNTPOWER', 'TRENT', 'TVSMOTOR', 'UBL', 'ULTRACEMCO', 'UPL', 'VEDL', 'VOLTAS',
                     'ZEEL', 'ZYDUSLIFE']

cash_share_list = ['ADANIENT', 'APOLLOTYRE', 'BAJAJFINSERV', 'BAJAJFINANCE', 'BANDHANBANK', 'BANKBARODA', 'COAL INDIA',
                   'DLF CHL', 'EICHERMOTOR', 'FEDRAL BANK', 'HCLTECH', 'HDFC', 'HINDALCO', 'ICICIBANK', 'INDUSINDBANK',
                   'INFY', 'JINDALS chl', 'LICHSGFIN', 'M&M', 'M&MFINANCE', 'NTPC', '03 RELIANCE CHL', '04 SBIN CHL',
                   'SUNTV', 'TATACHEM', '07 TATAMOTOR CHL', 'TATAPOWER', '05 TATASTEEL chl', 'ULTRACHEM']

add_share_list = []

# algo_share_list = ['ADANIENT']

red = Font("Arial", 12, color='ff0000', bold=True)
blue = Font("Arial", 12, color="0000ff", bold=True)
bold = Font("Arial", 12, bold=True)
alignment = Alignment(horizontal='center')


def weekly_create():
    for share in add_share_list:
        path = rf'E:\Daily Data work\ALGORITHM\{share}.xlsx'

        wb = xl.load_workbook(path)

        # del wb['W']

        wb.create_sheet('W')

        d_sheet = wb['D']
        w_sheet = wb['W']

        # weekly
        w_start_date = datetime.datetime(2020, 1, 6)
        d_row = 6
        w_row = 3

        while d_row < 2700:
            cur_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, '%d-%b-%y')
            high = 0
            low = 999999
            c = 0

            start_date = cur_date

            if cur_date == datetime.datetime(2020, 1, 27):
                end_date = cur_date + timedelta(days=5)

            else:
                end_date = cur_date + timedelta(days=4)

            while cur_date < end_date:
                cur_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, '%d-%b-%y')
                try:
                    h = float(d_sheet.cell(d_row, 2).value)
                    l = float(d_sheet.cell(d_row, 3).value)
                    c = float(d_sheet.cell(d_row, 4).value)
                except TypeError:
                    d_row += 1
                    continue

                if h > high:
                    high = h

                if l < low and l != 0:
                    low = l

                d_row += 1

            buff = 0
            close = c
            if not close or close == 0:
                while not c or c == 0:
                    c = d_sheet.cell(d_row, 4).value

                    d_row -= 1
                    buff += 1
                close = c

            d_row += buff

            w_sheet.cell(w_row, 1).value = f"{start_date.strftime('%d.%m.%y')} TO {end_date.strftime('%d.%m.%y')}"
            if d_row < 1018:
                w_sheet.cell(w_row, 2).value = high
                w_sheet.cell(w_row, 3).value = low
                w_sheet.cell(w_row, 4).value = close

            w_row += 1

        # formatting and headings
        w_sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=7)
        w_sheet.cell(1, 1).value = share
        w_sheet.cell(1, 1).fill = PatternFill(patternType='solid', fgColor="0000ff")
        w_sheet.cell(1, 1).font = Font("Arial", 11, bold=True, color='00ffffff')
        w_sheet.cell(1, 1).alignment = alignment

        for c in range(1, 7):
            w_sheet.cell(2, c).fill = PatternFill(patternType='solid', fgColor="0000ff")

        w_sheet.freeze_panes = w_sheet["A4"]

        w_sheet.cell(3, 1).value = 'SETTLEMENT PERIOD'
        w_sheet.cell(3, 2).value = 'HIGH'
        w_sheet.cell(3, 3).value = 'LOW'
        w_sheet.cell(3, 4).value = 'CL'
        w_sheet.cell(3, 5).value = 'TREND'
        w_sheet.cell(3, 6).value = 'H/L D'
        w_sheet.cell(3, 7).value = 'W/D'

        w_row = 2
        while w_row < 1000:
            if w_row >= 4:
                w_sheet.cell(w_row, 6).value = f'=B{w_row}-C{w_row}'
                w_sheet.cell(w_row, 7).value = f'=D{w_row}-D{w_row-1}'

            col = 1

            while col < 8:
                if col == 2 and w_row >= 3:
                    w_sheet.cell(w_row, col).font = blue
                elif col == 3 and w_row >= 3:
                    w_sheet.cell(w_row, col).font = red
                else:
                    w_sheet.cell(w_row, col).font = bold
                w_sheet.cell(w_row, col).alignment = alignment
                w_sheet.cell(w_row, col).border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'), top=Side(style='thin'))

                col += 1

            w_row += 1

        dim_holder = DimensionHolder(worksheet=w_sheet)

        for col in range(2, 12):
            dim_holder[get_column_letter(col)] = ColumnDimension(w_sheet, min=col, max=col, width=13.57)

        dim_holder[1] = ColumnDimension(w_sheet, min=1, max=1, width=23)
        w_sheet.column_dimensions = dim_holder

        w_sheet.sheet_view.zoomScale = 115
        d_sheet.sheet_view.zoomScale = 115

        wb.save(path)
        print(f'{share} done')


def calc_m_end_date(s_date):
    days = calendar.monthrange(s_date.year, s_date.month)[1]

    return datetime.datetime(s_date.year, s_date.month, days)


def monthly_create():
    for share in add_share_list:
        path = rf'E:\Daily Data work\ALGORITHM\{share}.xlsx'

        wb = xl.load_workbook(path)

        if 'M' in wb.sheetnames:
            del wb['M']

        wb.create_sheet('M')

        d_sheet = wb['D']
        m_sheet = wb['M']
        m_sheet.freeze_panes = m_sheet["A4"]

        # Monthly
        m_start_date = datetime.datetime(2020, 1, 1)
        start_date = m_start_date
        d_row = 3
        m_row = 4

        while d_row < 2700:
            cur_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, '%d-%b-%y')
            high = 0
            low = 999999
            c = 0

            end_date = calc_m_end_date(start_date)

            while cur_date <= end_date:
                try:
                    h = float(d_sheet.cell(d_row, 2).value)
                    l = float(d_sheet.cell(d_row, 3).value)
                    c = float(d_sheet.cell(d_row, 4).value)
                except TypeError:
                    d_row += 1
                    cur_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, '%d-%b-%y')
                    continue

                if h > high:
                    high = h

                if l < low and l != 0:
                    low = l

                d_row += 1
                cur_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, '%d-%b-%y')

            buff = 0
            close = c
            if not close or close == 0:
                while not c or c == 0:
                    c = d_sheet.cell(d_row, 4).value

                    d_row -= 1
                    buff += 1
                close = c

            d_row += buff

            m_sheet.cell(m_row, 1).value = f"{start_date.strftime('%d-%m-%y')} TO {end_date.strftime('%d-%m-%y')}"
            if d_row < 1018:            # todo change this row num when doing this for shares in future, it ignores h,l,c writing after this row
                m_sheet.cell(m_row, 2).value = f'=E{m_row-1}'
                m_sheet.cell(m_row, 3).value = high
                m_sheet.cell(m_row, 4).value = low
                m_sheet.cell(m_row, 5).value = close

            start_date = datetime.datetime((cur_date + timedelta(days=1)).year, (cur_date + timedelta(days=1)).month, 1)
            m_row += 1

        # formatting and headings
        m_sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=8)
        m_sheet.cell(1, 1).value = share
        m_sheet.cell(1, 1).fill = PatternFill(patternType='solid', fgColor="0000ff")
        m_sheet.cell(1, 1).font = Font("Arial", 11, bold=True, color='00ffffff')
        m_sheet.cell(1, 1).alignment = alignment

        for c in range(1, 8):
            m_sheet.cell(2, c).fill = PatternFill(patternType='solid', fgColor="0000ff")

        m_sheet.cell(3, 1).value = 'SETTLEMENT PERIOD'
        m_sheet.cell(3, 2).value = 'CL START'
        m_sheet.cell(3, 3).value = 'HIGH'
        m_sheet.cell(3, 4).value = 'LOW'
        m_sheet.cell(3, 5).value = 'CL END'
        m_sheet.cell(3, 6).value = 'TREND'
        m_sheet.cell(3, 7).value = 'H/L D'
        m_sheet.cell(3, 8).value = 'W/D'

        m_row = 2
        while m_row < 251:
            if m_row >= 4:
                m_sheet.cell(m_row, 7).value = f'=C{m_row}-D{m_row}'
                m_sheet.cell(m_row, 8).value = f'=E{m_row}-E{m_row-1}'
                m_sheet.cell(m_row, 2).value = f'=E{m_row-1}'

            col = 1

            while col < 9:
                if col == 3 and m_row >= 3:
                    m_sheet.cell(m_row, col).font = blue
                elif col == 4 and m_row >= 3:
                    m_sheet.cell(m_row, col).font = red
                else:
                    m_sheet.cell(m_row, col).font = bold
                m_sheet.cell(m_row, col).alignment = alignment
                m_sheet.cell(m_row, col).border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'), top=Side(style='thin'))

                col += 1

            m_row += 1

        # putting 0 in first cl start cell and cl diff cell
        m_sheet.cell(4, 2). value = 0
        m_sheet.cell(4, 7). value = 0

        dim_holder = DimensionHolder(worksheet=m_sheet)

        for col in range(2, 12):
            dim_holder[get_column_letter(col)] = ColumnDimension(m_sheet, min=col, max=col, width=13.57)

        dim_holder[1] = ColumnDimension(m_sheet, min=1, max=1, width=23)
        m_sheet.column_dimensions = dim_holder

        m_sheet.sheet_view.zoomScale = 115
        d_sheet.sheet_view.zoomScale = 115

        wb.save(path)
        # wb.save('m_test.xlsx')
        print(f'{share} done')


def calc_cl_end_date(s_date: datetime.datetime):
    s_date += timedelta(days=21)
    mnth = s_date.month

    while (s_date + timedelta(days=1)).month == mnth:
        s_date += timedelta(days=1)

    while s_date.weekday() != 3:
        s_date -= timedelta(days=1)

    return s_date


def closing_create():
    for share in add_share_list:
        path = rf'E:\Daily Data work\ALGORITHM\{share}.xlsx'

        wb = xl.load_workbook(path)

        if 'Cl' in wb.sheetnames:
            del wb['Cl']

        wb.create_sheet('Cl')

        d_sheet = wb['D']
        cl_sheet = wb['Cl']
        cl_sheet.freeze_panes = cl_sheet["A4"]

        # Closing
        cl_start_date = datetime.datetime(2020, 1, 31)
        start_date = cl_start_date
        d_row = 25      # 31-JAN-2020
        cl_row = 4

        while d_row < 2700:
            cur_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, '%d-%b-%y')
            high = 0
            low = 999999
            c = 0

            end_date = calc_cl_end_date(start_date)

            while cur_date <= end_date:
                try:
                    h = float(d_sheet.cell(d_row, 2).value)
                    l = float(d_sheet.cell(d_row, 3).value)
                    c = float(d_sheet.cell(d_row, 4).value)
                except TypeError:
                    d_row += 1
                    cur_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, '%d-%b-%y')
                    continue

                if h > high:
                    high = h

                if l < low and l != 0:
                    low = l

                d_row += 1
                cur_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, '%d-%b-%y')

            buff = 0
            close = c
            if not close or close == 0:
                while not c or c == 0:
                    c = d_sheet.cell(d_row, 4).value

                    d_row -= 1
                    buff += 1
                close = c

            d_row += buff

            cl_sheet.cell(cl_row, 1).value = f"{start_date.strftime('%d-%m-%y')} TO {end_date.strftime('%d-%m-%y')}"
            if d_row < 1018:            # todo change this row num when doing this for shares in future, it ignores h,l,c writing after this row
                cl_sheet.cell(cl_row, 2).value = f'=E{cl_row-1}'
                cl_sheet.cell(cl_row, 3).value = high
                cl_sheet.cell(cl_row, 4).value = low
                cl_sheet.cell(cl_row, 5).value = close

            start_date = cur_date + timedelta(days=1)
            cl_row += 1

        # formatting and headings
        cl_sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=8)
        cl_sheet.cell(1, 1).value = share
        cl_sheet.cell(1, 1).fill = PatternFill(patternType='solid', fgColor="0000ff")
        cl_sheet.cell(1, 1).font = Font("Arial", 11, bold=True, color='00ffffff')
        cl_sheet.cell(1, 1).alignment = alignment

        for c in range(1, 8):
            cl_sheet.cell(2, c).fill = PatternFill(patternType='solid', fgColor="0000ff")

        cl_sheet.cell(3, 1).value = 'SETTLEMENT PERIOD'
        cl_sheet.cell(3, 2).value = 'CL START'
        cl_sheet.cell(3, 3).value = 'HIGH'
        cl_sheet.cell(3, 4).value = 'LOW'
        cl_sheet.cell(3, 5).value = 'CL END'
        cl_sheet.cell(3, 6).value = 'TREND'
        cl_sheet.cell(3, 7).value = 'H/L D'
        cl_sheet.cell(3, 8).value = 'W/D'

        cl_row = 2
        while cl_row < 251:
            if cl_row >= 4:
                cl_sheet.cell(cl_row, 7).value = f'=C{cl_row}-D{cl_row}'
                cl_sheet.cell(cl_row, 8).value = f'=E{cl_row}-E{cl_row-1}'
                cl_sheet.cell(cl_row, 2).value = f'=E{cl_row-1}'

            col = 1

            while col < 9:
                if col == 3 and cl_row >= 3:
                    cl_sheet.cell(cl_row, col).font = blue
                elif col == 4 and cl_row >= 3:
                    cl_sheet.cell(cl_row, col).font = red
                else:
                    cl_sheet.cell(cl_row, col).font = bold
                cl_sheet.cell(cl_row, col).alignment = alignment
                cl_sheet.cell(cl_row, col).border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'), top=Side(style='thin'))

                col += 1

            cl_row += 1

        # putting 0 in first cl start cell and cl diff cell
        cl_sheet.cell(4, 2). value = 0
        cl_sheet.cell(4, 7). value = 0

        dim_holder = DimensionHolder(worksheet=cl_sheet)

        for col in range(2, 12):
            dim_holder[get_column_letter(col)] = ColumnDimension(cl_sheet, min=col, max=col, width=13.57)

        dim_holder[1] = ColumnDimension(cl_sheet, min=1, max=1, width=23)
        cl_sheet.column_dimensions = dim_holder

        cl_sheet.sheet_view.zoomScale = 115
        d_sheet.sheet_view.zoomScale = 115

        wb.save(path)
        # wb.save('m_test.xlsx')
        print(f'{share} done')


# by default returns last empty row unless specifically the empty param is made False
def get_last_row(sheet, empty=True):
    row = sheet.max_row

    while True:
        if sheet.cell(row, 3).value is not None:
            if empty:
                return row+1
            else:
                return row

        row -= 1


def weekly_update(typ):
    if typ == 'C':
        share_list = cash_share_list
    else:
        share_list = algo_share_list

    for share in share_list:
        if typ == 'C':
            path = rf'E:\Daily Data work\CASH\{share}.xlsx'
        else:
            path = rf'E:\Daily Data work\ALGORITHM\{share}.xlsx'

        wb = xl.load_workbook(path)

        d_sheet = wb['D']
        w_sheet = wb['W']

        w_row = get_last_row(w_sheet)
        d_row = get_last_row(d_sheet, empty=False)
        print(w_row)

        high = 0
        low = 999999
        close = 0
        cl_found = False

        for i in range(5):
            try:
                h = float(d_sheet.cell(d_row, 2).value)
                l = float(d_sheet.cell(d_row, 3).value)

                if not cl_found:
                    close = float(d_sheet.cell(d_row, 4).value)
                    cl_found = True
            except TypeError:
                d_row -= 1
                continue

            if h > high:
                high = h

            if l < low and l != 0:
                low = l

            d_row -= 1

        w_sheet.cell(w_row, 2).value = high
        w_sheet.cell(w_row, 2).font = blue
        w_sheet.cell(w_row, 2).alignment = alignment
        w_sheet.cell(w_row, 2).number_format = '0'

        w_sheet.cell(w_row, 3).value = low
        w_sheet.cell(w_row, 3).font = red
        w_sheet.cell(w_row, 3).alignment = alignment
        w_sheet.cell(w_row, 3).number_format = '0'

        w_sheet.cell(w_row, 4).value = close
        w_sheet.cell(w_row, 4).font = bold
        w_sheet.cell(w_row, 4).alignment = alignment
        w_sheet.cell(w_row, 4).number_format = '0'

        wb.save(path)
        # wb.save('m_test.xlsx')
        print(f'{share} done')


def monthly_update(typ):
    if typ == 'C':
        share_list = cash_share_list
    else:
        share_list = algo_share_list

    for share in share_list:
        if typ == 'C':
            path = rf'E:\Daily Data work\CASH\{share}.xlsx'
        else:
            path = rf'E:\Daily Data work\ALGORITHM\{share}.xlsx'

        wb = xl.load_workbook(path)

        d_sheet = wb['D']
        m_sheet = wb['M']

        m_row = get_last_row(m_sheet)
        d_row = get_last_row(d_sheet, empty=False)

        high = 0
        low = 999999
        close = 0
        cl_found = False
        date_format = "%d-%m-%y"

        # Extract start and end dates
        date_range = str(m_sheet.cell(m_row, 1).value)
        start_date_str, end_date_str = date_range.split(" TO ")

        # Convert strings to datetime objects
        start_date = datetime.datetime.strptime(start_date_str, date_format)
        end_date = datetime.datetime.strptime(end_date_str, date_format)

        # # -5 days to ensure its working even if doing 1 or 2 day after when the month has changed
        # date = datetime.datetime.now() - timedelta(days=5)
        # month_length = (calendar.monthrange(date.year, date.month)[1])
        while start_date <= end_date:
            try:
                h = float(d_sheet.cell(d_row, 2).value)
                l = float(d_sheet.cell(d_row, 3).value)

                if not cl_found:
                    close = float(d_sheet.cell(d_row, 4).value)
                    cl_found = True
            except TypeError:
                d_row -= 1
                if type(d_sheet.cell(d_row, 1).value) == str:
                    end_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, "%d-%b-%y")
                else:  # if already in datetime.datetime format
                    end_date = d_sheet.cell(d_row, 1).value
                continue

            if h > high:
                high = h

            if l < low and l != 0:
                low = l

            d_row -= 1
            if type(d_sheet.cell(d_row, 1).value) == str:
                end_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, "%d-%b-%y")
            else:       # if already in datetime.datetime format
                end_date = d_sheet.cell(d_row, 1).value
        # for i in range(month_length):
        #     try:
        #         h = float(d_sheet.cell(d_row, 2).value)
        #         l = float(d_sheet.cell(d_row, 3).value)
        #
        #         if not cl_found:
        #             close = float(d_sheet.cell(d_row, 4).value)
        #             cl_found = True
        #     except TypeError:
        #         d_row -= 1
        #         continue
        #
        #     if h > high:
        #         high = h
        #
        #     if l < low and l != 0:
        #         low = l
        #
        #     d_row -= 1

        m_sheet.cell(m_row, 3).value = high
        m_sheet.cell(m_row, 3).font = blue
        m_sheet.cell(m_row, 3).alignment = alignment

        m_sheet.cell(m_row, 4).value = low
        m_sheet.cell(m_row, 4).font = red
        m_sheet.cell(m_row, 4).alignment = alignment

        m_sheet.cell(m_row, 5).value = close
        m_sheet.cell(m_row, 5).font = bold
        m_sheet.cell(m_row, 5).alignment = alignment

        wb.save(path)
        # wb.save('m_test.xlsx')
        print(f'{share} done')


def closing_update(typ):
    if typ == 'C':
        share_list = cash_share_list
    else:
        share_list = algo_share_list

    for share in share_list:
        if typ == 'C':
            path = rf'E:\Daily Data work\CASH\{share}.xlsx'
        else:
            path = rf'E:\Daily Data work\ALGORITHM\{share}.xlsx'

        wb = xl.load_workbook(path)

        d_sheet = wb['D']
        cl_sheet = wb['Cl']

        cl_row = get_last_row(cl_sheet)
        d_row = get_last_row(d_sheet, empty=False)

        high = 0
        low = 999999
        close = 0
        cl_found = False
        date_format = "%d-%m-%y"

        # Extract start and end dates
        date_range = str(cl_sheet.cell(cl_row, 1).value)
        start_date_str, end_date_str = date_range.split(" TO ")

        # Convert strings to datetime objects
        start_date = datetime.datetime.strptime(start_date_str, date_format)
        end_date = datetime.datetime.strptime(end_date_str, date_format)

        if typ != 'C':
            start_date = start_date - timedelta(days=1)     # todo untested hotfix for closing dates being wrong
            cl_sheet.cell(cl_row, 1).value = f'{start_date.date()} TO {end_date.date()}'
            print(f'{start_date.date().strftime(date_format)} TO {end_date.date().strftime(date_format)}')

        while start_date <= end_date:
            try:
                h = float(d_sheet.cell(d_row, 2).value)
                l = float(d_sheet.cell(d_row, 3).value)

                if not cl_found:
                    close = float(d_sheet.cell(d_row, 4).value)
                    cl_found = True
            except TypeError:
                d_row -= 1
                if type(d_sheet.cell(d_row, 1).value) == str:
                    end_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, "%d-%b-%y")
                else:  # if already in datetime.datetime format
                    end_date = d_sheet.cell(d_row, 1).value
                continue

            if h > high:
                high = h

            if l < low and l != 0:
                low = l

            d_row -= 1
            if type(d_sheet.cell(d_row, 1).value) == str:
                end_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, "%d-%b-%y")
            else:       # if already in datetime.datetime format
                end_date = d_sheet.cell(d_row, 1).value

        cl_sheet.cell(cl_row, 3).value = high
        cl_sheet.cell(cl_row, 3).font = blue
        cl_sheet.cell(cl_row, 3).alignment = alignment

        cl_sheet.cell(cl_row, 4).value = low
        cl_sheet.cell(cl_row, 4).font = red
        cl_sheet.cell(cl_row, 4).alignment = alignment

        cl_sheet.cell(cl_row, 5).value = close
        cl_sheet.cell(cl_row, 5).font = bold
        cl_sheet.cell(cl_row, 5).alignment = alignment

        wb.save(path)
        # wb.save('m_test.xlsx')
        print(f'{share} done')


# weekly_create()
# monthly_create()
# closing_create()

# weekly_update("A")
# weekly_update("C")

# monthly_update("A")
# monthly_update("C")

closing_update("A")
# closing_update("C")

