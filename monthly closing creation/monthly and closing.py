import datetime
from datetime import timedelta
from win32com.client import Dispatch
import openpyxl as xl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension


algo_share_list = ['AARTIIND', 'ABB', 'ABCAPITAL', 'ABFRL', 'ADANIENT', 'ADANIPORTS', 'ALKEM', 'AMBUJACEM',
                     'APOLLOHOSP', 'APOLLOTYRE', 'ASHOKLEY', 'ASTRAL', 'ATUL', 'AUBANK', 'AUROPHARMA', 'BAJAJFINSV',
                     'BAJFINANCE', 'BALKRISIND', 'BALRAMCHIN', 'BANDHANBNK', 'BANKBARODA', 'BATAINDIA', 'BEL',
                     'BHARATFORG', 'BIOCON', 'BRITANNIA', 'BSOFT', 'CANBK', 'CANFINHOME', 'CHAMBLFERT', 'CHOLAFIN',
                     'CIPLA', 'COFORGE', 'CONCOR', 'COROMANDEL', 'CROMPTON', 'CUMMINSIND', 'DABUR', 'DALBHARAT',
                     'DEEPAKFERT', 'DEEPAKNTR', 'DELTACORP', 'DIVISLAB', 'DIXON', 'DLF', 'DRREDDY', 'ESCORTS',
                     'EXIDEIND', 'GLENMARK', 'GLS', 'GNFC', 'GODREJCP', 'GODREJPROP', 'GRANULES', 'GRASIM', 'GUJGASLTD',
                     'HAL', 'HAVELLS', 'HCLTECH', 'HDFCAMC', 'HDFCLIFE', 'HINDALCO', 'HINDCOPPER', 'ICICIGI',
                     'ICICIPRULI', 'IEX', 'IGL', 'INDHOTEL', 'INDIACEM', 'INDIAMART', 'INDIGO', 'INDUSINDBK',
                     'INDUSTOWER', 'INTELLECT', 'IPCALAB', 'JINDALSTEL', 'JKCEMENT', 'JSWSTEEL', 'JUBLFOOD',
                     'KOTAKBANK', 'LALPATHLAB', 'LAURUSLABS', 'LICHSGFIN', 'LTIM', 'LTTS', 'LUPIN', 'M%26MFIN',
                     'MANAPPURAM', 'MARICO', 'MCDOWELL-N', 'MCX', 'METROPOLIS', 'MFSL', 'MGL', 'MPHASIS', 'MUTHOOTFIN',
                     'NAM-INDIA', 'NAUKRI', 'NAVINFLUOR', 'NMDC', 'NTPC', 'OBEROIRLTY', 'PEL', 'PERSISTENT', 'PETRONET',
                     'PIDILITIND', 'POLYCAB', 'POWERGRID', 'RAIN', 'RAMCOCEM', 'RBLBANK', 'RECLTD', 'SBICARD',
                     'SBILIFE', 'SIEMENS', 'SRF', 'STAR', 'SUNPHARMA', 'SYNGENE', 'TATACOMM', 'TATAMOTORS', 'TECHM',
                     'TORNTPHARM', 'TORNTPOWER', 'TRENT', 'TVSMOTOR', 'UBL', 'ULTRACEMCO', 'UPL', 'VEDL', 'VOLTAS',
                     'ZEEL', 'ZYDUSLIFE']

algo_share_list = ['ADANIENT']

red = Font("Arial", 11, color='ff0000', bold=True)
blue = Font("Arial", 11, color="0000ff", bold=True)
bold = Font("Arial", 11, bold=True)
alignment = Alignment(horizontal='center')


def weekly_create():
    for share in algo_share_list:
        path = rf'E:\Daily Data work\ALGORITHM\{share}.xlsx'

        wb = xl.load_workbook(path)

        # del wb['W']

        wb.create_sheet('W')

        d_sheet = wb['D']
        w_sheet = wb['W']

        # weekly
        w_start_date = datetime.datetime(2020, 1, 6)
        d_row = 6
        w_row = 3

        while d_row < 2700:
            cur_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, '%d-%b-%y')
            high = 0
            low = 999999
            close = 0

            start_date = cur_date

            if cur_date == datetime.datetime(2020, 1, 27):
                end_date = cur_date + timedelta(days=5)

            else:
                end_date = cur_date + timedelta(days=4)

            while cur_date < end_date:
                cur_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, '%d-%b-%y')
                try:
                    h = float(d_sheet.cell(d_row, 2).value)
                    l = float(d_sheet.cell(d_row, 3).value)
                    c = float(d_sheet.cell(d_row, 4).value)
                except TypeError:
                    d_row += 1
                    continue

                if h > high:
                    high = h

                if l < low and l != 0:
                    low = l

                d_row += 1

            buff = 0
            close = c
            if not close or close == 0:
                while not c or c == 0:
                    c = d_sheet.cell(d_row, 4).value

                    d_row -= 1
                    buff += 1
                close = c

            d_row += buff

            w_sheet.cell(w_row, 1).value = f"{start_date.strftime('%d.%m.%y')} TO {end_date.strftime('%d.%m.%y')}"
            if d_row < 1018:
                w_sheet.cell(w_row, 2).value = high
                w_sheet.cell(w_row, 3).value = low
                w_sheet.cell(w_row, 4).value = close

            w_row += 1

        # formatting and headings
        w_sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=7)
        w_sheet.cell(1, 1).value = share
        w_sheet.cell(1, 1).fill = PatternFill(patternType='solid', fgColor="0000ff")
        w_sheet.cell(1, 1).font = Font("Arial", 11, bold=True, color='00ffffff')
        w_sheet.cell(1, 1).alignment = alignment

        for c in range(1, 7):
            w_sheet.cell(2, c).fill = PatternFill(patternType='solid', fgColor="0000ff")

        w_sheet.freeze_panes = w_sheet["A3"]

        w_sheet.cell(3, 1).value = 'SETTLEMENT PERIOD'
        w_sheet.cell(3, 2).value = 'HIGH'
        w_sheet.cell(3, 3).value = 'LOW'
        w_sheet.cell(3, 4).value = 'CL'
        w_sheet.cell(3, 5).value = 'TREND'
        w_sheet.cell(3, 6).value = 'H/L D'
        w_sheet.cell(3, 7).value = 'W/D'

        w_row = 2
        while w_row < 1000:
            if w_row >= 4:
                w_sheet.cell(w_row, 6).value = f'=B{w_row}-C{w_row}'
                w_sheet.cell(w_row, 7).value = f'=D{w_row}-D{w_row-1}'

            col = 1

            while col < 8:
                if col == 2 and w_row >= 3:
                    w_sheet.cell(w_row, col).font = blue
                elif col == 3 and w_row >= 3:
                    w_sheet.cell(w_row, col).font = red
                else:
                    w_sheet.cell(w_row, col).font = bold
                w_sheet.cell(w_row, col).alignment = alignment
                w_sheet.cell(w_row, col).border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'), top=Side(style='thin'))

                col += 1

            w_row += 1

        dim_holder = DimensionHolder(worksheet=w_sheet)

        for col in range(2, 12):
            dim_holder[get_column_letter(col)] = ColumnDimension(w_sheet, min=col, max=col, width=13.57)

        dim_holder[1] = ColumnDimension(w_sheet, min=1, max=1, width=23)
        w_sheet.column_dimensions = dim_holder

        w_sheet.sheet_view.zoomScale = 115
        d_sheet.sheet_view.zoomScale = 115

        wb.save(path)
        print(f'{share} done')


def monthly_create():
    for share in algo_share_list:
        path = rf'E:\Daily Data work\ALGORITHM\{share}.xlsx'

        wb = xl.load_workbook(path)

        # del wb['W']

        wb.create_sheet('W')

        d_sheet = wb['D']
        w_sheet = wb['W']

        # weekly
        w_start_date = datetime.datetime(2020, 1, 6)
        d_row = 6
        w_row = 3

        while d_row < 2700:
            cur_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, '%d-%b-%y')
            high = 0
            low = 999999
            close = 0

            start_date = cur_date

            if cur_date == datetime.datetime(2020, 1, 27):
                end_date = cur_date + timedelta(days=5)

            else:
                end_date = cur_date + timedelta(days=4)

            while cur_date < end_date:
                cur_date = datetime.datetime.strptime(d_sheet.cell(d_row, 1).value, '%d-%b-%y')
                try:
                    h = float(d_sheet.cell(d_row, 2).value)
                    l = float(d_sheet.cell(d_row, 3).value)
                    c = float(d_sheet.cell(d_row, 4).value)
                except TypeError:
                    d_row += 1
                    continue

                if h > high:
                    high = h

                if l < low and l != 0:
                    low = l

                d_row += 1

            buff = 0
            close = c
            if not close or close == 0:
                while not c or c == 0:
                    c = d_sheet.cell(d_row, 4).value

                    d_row -= 1
                    buff += 1
                close = c

            d_row += buff

            w_sheet.cell(w_row, 1).value = f"{start_date.strftime('%d.%m.%y')} TO {end_date.strftime('%d.%m.%y')}"
            if d_row < 1018:
                w_sheet.cell(w_row, 2).value = high
                w_sheet.cell(w_row, 3).value = low
                w_sheet.cell(w_row, 4).value = close

            w_row += 1

        # formatting and headings
        w_sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=7)
        w_sheet.cell(1, 1).value = share
        w_sheet.cell(1, 1).fill = PatternFill(patternType='solid', fgColor="0000ff")
        w_sheet.cell(1, 1).font = Font("Arial", 11, bold=True, color='00ffffff')
        w_sheet.cell(1, 1).alignment = alignment

        for c in range(1, 7):
            w_sheet.cell(2, c).fill = PatternFill(patternType='solid', fgColor="0000ff")

        w_sheet.freeze_panes = w_sheet["A3"]

        w_sheet.cell(3, 1).value = 'SETTLEMENT PERIOD'
        w_sheet.cell(3, 2).value = 'HIGH'
        w_sheet.cell(3, 3).value = 'LOW'
        w_sheet.cell(3, 4).value = 'CL'
        w_sheet.cell(3, 5).value = 'TREND'
        w_sheet.cell(3, 6).value = 'H/L D'
        w_sheet.cell(3, 7).value = 'W/D'

        w_row = 2
        while w_row < 1000:
            if w_row >= 4:
                w_sheet.cell(w_row, 6).value = f'=B{w_row}-C{w_row}'
                w_sheet.cell(w_row, 7).value = f'=D{w_row}-D{w_row-1}'

            col = 1

            while col < 8:
                if col == 2 and w_row >= 3:
                    w_sheet.cell(w_row, col).font = blue
                elif col == 3 and w_row >= 3:
                    w_sheet.cell(w_row, col).font = red
                else:
                    w_sheet.cell(w_row, col).font = bold
                w_sheet.cell(w_row, col).alignment = alignment
                w_sheet.cell(w_row, col).border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'), top=Side(style='thin'))

                col += 1

            w_row += 1

        dim_holder = DimensionHolder(worksheet=w_sheet)

        for col in range(2, 12):
            dim_holder[get_column_letter(col)] = ColumnDimension(w_sheet, min=col, max=col, width=13.57)

        dim_holder[1] = ColumnDimension(w_sheet, min=1, max=1, width=23)
        w_sheet.column_dimensions = dim_holder

        w_sheet.sheet_view.zoomScale = 115
        d_sheet.sheet_view.zoomScale = 115

        wb.save(path)
        print(f'{share} done')


# weekly_create()

# for share in algo_share_list:
#     path = rf'E:\Daily Data work\ALGORITHM\{share}.xlsx'
#
#     wb = xl.load_workbook(path)
#     w_sheet = wb['W']
#     w_sheet.freeze_panes = w_sheet["A4"]
#     wb.save(path)
