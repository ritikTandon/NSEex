# # from time import sleep
# # from selenium import webdriver
# # from selenium.webdriver.support.ui import WebDriverWait
# # from selenium.webdriver.support import expected_conditions as ec
# # from selenium.webdriver.common.by import By
# # from selenium.webdriver.chrome.options import Options
# # from selenium.common.exceptions import TimeoutException
# #
# # # headers = {'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:52.0) Gecko/20100101 Firefox/52.0'}
# #
# # cash_share_list = ["AARTIIND", "ADANI", "APOLLO", "BAJFINSV", "BAJFIN", "BANBK", "BARODA", "COALIND", "DLF", "EICHER",
# #                    "FEDBANK", "HCL", "HDFC", "HIND", "ICICI", "INDUSIND", "INFY", "JIND", "LIC", "M&M", "M&MFIN", "NTPC",
# #                    "REL", "SBIN", "SUNTV", "TCHEM", "TM", "TP", "TS", "ULTRA"]
# #
# # # for close filling
# # options = Options()
# # # options.add_argument('--headless=new')
# # options.add_argument("--disable-blink-features=AutomationControlled")
# #
# # # Exclude the collection of enable-automation switches
# # options.add_experimental_option("excludeSwitches", ["enable-automation"])
# #
# # # Turn-off userAutomationExtension
# # options.add_experimental_option("useAutomationExtension", False)
# #
# # cash_close_list = ['AARTIIND', 'ABB', 'ABCAPITAL', 'ABFRL', 'ADANIENT', 'ADANIPORTS', 'ALKEM', 'AMBUJACEM',
# #                      'APOLLOHOSP', 'APOLLOTYRE', 'ASHOKLEY', 'ASTRAL', 'ATUL', 'AUBANK', 'AUROPHARMA', 'BAJAJFINSV',
# #                      'BAJFINANCE', 'BALKRISIND', 'BALRAMCHIN', 'BANDHANBNK', 'BANKBARODA', 'BATAINDIA', 'BEL',
# #                      'BHARATFORG', 'BIOCON', 'BRITANNIA', 'BSOFT', 'CANBK', 'CANFINHOME', 'CHAMBLFERT', 'CHOLAFIN',
# #                      'CIPLA', 'COFORGE', 'CONCOR', 'COROMANDEL', 'CROMPTON', 'CUMMINSIND', 'DABUR', 'DALBHARAT',
# #                      'DEEPAKFERT', 'DEEPAKNTR', 'DELTACORP', 'DIVISLAB', 'DIXON', 'DLF', 'DRREDDY', 'ESCORTS',
# #                      'EXIDEIND', 'GLENMARK', 'GLS', 'GNFC', 'GODREJCP', 'GODREJPROP', 'GRANULES', 'GRASIM', 'GUJGASLTD',
# #                      'HAL', 'HAVELLS', 'HCLTECH', 'HDFCAMC', 'HDFCLIFE', 'HINDALCO', 'HINDCOPPER', 'ICICIGI',
# #                      'ICICIPRULI', 'IEX', 'IGL', 'INDHOTEL', 'INDIACEM', 'INDIAMART', 'INDIGO', 'INDUSINDBK',
# #                      'INDUSTOWER', 'INTELLECT', 'IPCALAB', 'JINDALSTEL', 'JKCEMENT', 'JSWSTEEL', 'JUBLFOOD',
# #                      'KOTAKBANK', 'LALPATHLAB', 'LAURUSLABS', 'LICHSGFIN', 'LTIM', 'LTTS', 'LUPIN', 'M%26MFIN',
# #                      'MANAPPURAM', 'MARICO', 'MCDOWELL-N', 'MCX', 'METROPOLIS', 'MFSL', 'MGL', 'MPHASIS', 'MUTHOOTFIN',
# #                      'NAM-INDIA', 'NAUKRI', 'NAVINFLUOR', 'NMDC', 'NTPC', 'OBEROIRLTY', 'PEL', 'PERSISTENT', 'PETRONET',
# #                      'PIDILITIND', 'POLYCAB', 'POWERGRID', 'RAIN', 'RAMCOCEM', 'RBLBANK', 'RECLTD', 'SBICARD',
# #                      'SBILIFE', 'SIEMENS', 'SRF', 'STAR', 'SUNPHARMA', 'SYNGENE', 'TATACOMM', 'TATAMOTORS', 'TECHM',
# #                      'TORNTPHARM', 'TORNTPOWER', 'TRENT', 'TVSMOTOR', 'UBL', 'ULTRACEMCO', 'UPL', 'VEDL', 'VOLTAS',
# #                      'ZEEL', 'ZYDUSLIFE']
# #
# # # cash_close_list1 = ["M%26M"]
# #
# # manual = []         # list to keep track of the shares whose values selenium couldn't get
# # close = []
# # # close = ['2446.95', '398.5', '1478.0', '7005.0', '227.65', '188.4', '234.15', '470.1', '3337.8', '132.8', '1170.0',
# # # '1608.5', '448.85', '959.0', '1387.8', '1392.95', '662.9', '421.45', '1544.95', '1544.95', '2573.2', '561.0', '548.9',
# # # '1000.05', '606.8', '230.95', '117.85', '8024.95']
# #
# # for share in cash_close_list:
# #     driver = webdriver.Chrome(options=options)
# #
# #     driver.get(f"https://www.nseindia.com/get-quotes/equity?symbol={share}")
# #
# #     try:
# #         sleep(2)
# #         WebDriverWait(driver, 10).until(ec.presence_of_element_located((By.ID, 'quoteLtp')))
# #
# #         close_val = driver.find_element(By.ID, "quoteLtp").text
# #
# #         while close_val == '':
# #             driver.refresh()
# #             WebDriverWait(driver, 10).until(ec.presence_of_element_located((By.ID, 'quoteLtp')))
# #             close_val = driver.find_element(By.ID, "quoteLtp").text
# #             sleep(0.5)
# #
# #         close_val = close_val.replace(",", "")
# #
# #         # truncating last 0
# #         if close_val[len(close_val)-1:len(close_val)] == '0':
# #             close_val = close_val[:len(close_val)-1]
# #
# #         close.append(close_val)
# #
# #         print(f'{share}: {close_val}')
# #         if close_val == '':
# #             manual.append(share)
# #
# #     except TimeoutException:        # added temp fix for timeoutexception, will need to check if it works properly or nah
# #         close.append('')
# #         manual.append(share)
# #         print(f"Loading took too much time for {share}!")
# #
# #     driver.close()
# #
# # print(close)
# # print(manual)
# #
# # for i in close:
# #     print(i)
#
#
# # import pyautogui as pg
# # import numpy as np
# # from time import sleep
# # #
# # # pg.click((797, 1058))
# # # sleep(1)
# # # # # Take a screenshot of the specified region
# # # region = (929, 212, 6, 8)
# # # screenshot = pg.screenshot(region=region)
# # # #
# # # # # Convert the screenshot to a NumPy array
# # # pixels = np.array(screenshot)
# # #
# # # # Calculate the average color
# # # average_color = pixels.mean(axis=(0, 1)).astype(int)
# # #
# # # # Print the average color as an RGB tuple
# # # print(f"Average color: {tuple(average_color)}")
# #
# #
# # pg.click((797, 1058))
# # sleep(1)
#
# # def check_change():
# #     changed = False  # bool representing if screen has changed from black to showing vwap statistics
# #
# #     # waiting for screen to change before saving
# #     while not changed:
# #         # # Take a screenshot of the specified region
# #         region = (929, 212, 6, 8)
# #         screenshot = pg.screenshot(region=region)
# #         pixels = np.array(screenshot)
# #         # Calculate the average color
# #         average_color = pixels.mean(axis=(0, 1)).astype(int)
# #
# #         # Print the average color as an RGB tuple
# #         print(f"Average color: {tuple(average_color)}")
# #
# #         if 200 > average_color[1] > 50:  # if color is green (screen has updated)
# #             print("Displayed the vwap data")
# #             print(f"Average color: {tuple(average_color)}")
# #             changed = True
# #
# #
# #
# # check_change()
#
#
# # import os
# # from time import sleep
# #
# # f = ['AARTIIND', 'ABB', 'AMBUJACEM', 'APOLLOHOSP', 'APOLLOTYRE', 'ASHOKLEY', 'AUROPHARMA', 'BANKBARODA', 'BEL',
# #      'BSOFT', 'CANBK', 'CANFINHOME', 'CHAMBLFERT', 'CHOLAFIN', 'DIXON', 'DLF', 'ESCORTS', 'EXIDEIND', 'GNFC',
# #      'GODREJPROP', 'HAL', 'HAVELLS', 'HCLTECH', 'HINDALCO', 'HINDCOPPER', 'IGL', 'INDIACEM', 'INDUSINDBK', 'INDUSTOWER',
# #      'JINDALSTEL', 'JUBLFOOD', 'LALPATHLAB', 'LICHSGFIN', 'LUPIN', 'MANAPPURAM', 'MCX', 'METROPOLIS', 'MGL',
# #      'MUTHOOTFIN', 'NAM-INDIA', 'NMDC', 'NTPC', 'PETRONET', 'RAMCOCEM', 'RBLBANK', 'RECLTD', 'STAR', 'TRENT', 'VEDL']
# #
# #
# # for fl in f:
# #     path = rf'E:\Daily Data work\ALGORITHM\{fl}.xlsx'
# #     os.startfile(path)
#
#
# import pandas as pd
# import re
# from date_variables import yr, date, mnth
# from zipfile import ZipFile
#
# share_list = ["BANKNIFTY", "NIFTY"]
#
# md_path_zipped = rf"E:\chrome downloads\BhavCopy_NSE_FO_0_0_0_{yr}{date[3:5]}{date[:2]}_F_0000.csv.zip"     # .zip file path of downloaded cash bhavcopy
# md_path = rf"E:\chrome downloads"
#
# # take data from col 22 and col 14 has month as cur month in caps and col 13 is null
# #
# # # extracting .zip file
# # with ZipFile(md_path_zipped, 'r') as zObject:
# #     zObject.extractall(path=md_path)
# #
# md_file_path = rf"E:\Daily Data work\MD files\{yr}\{mnth}\fo{date[:2]}{mnth}20{date[6:]}bhav.xlsx"
#
# df = pd.read_csv(md_path_zipped[:-4])       # removing the .zip extension after unzipping
# df = df.drop(df.columns[-9:], axis=1)
# df = df.drop(columns=['BizDt', 'Src', 'FinInstrmTp', 'FinInstrmId', 'ISIN', 'SctySrs', 'OpnIntrst', 'ChngInOpnIntrst'])
# df2 = df.copy()     # this one will actually be saved as md file
# df = df[df['OptnTp'].isnull()]
#
# import datetime
# import calendar
# from date_variables import date,mnth, yr
#
#
# def checkExpiry(day, month_abbr, year):
#     # Mapping of month abbreviations to month numbers
#     month_map = {
#         "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
#         "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12
#     }
#
#     # Convert the month abbreviation to a month number
#     month = month_map.get(month_abbr.upper())
#
#     if month is None:
#         raise ValueError("Invalid month abbreviation")
#
#     # Get the number of days in the month
#     _, last_day_of_month = calendar.monthrange(year, month)
#
#     # Find the last Thursday of the month
#     last_date_of_month = datetime.date(year, month, last_day_of_month)
#     while last_date_of_month.weekday() != 3:  # 3 corresponds to Thursday
#         last_date_of_month -= datetime.timedelta(days=1)
#
#     # Check if the given day is greater than or equal to the last Thursday
#     given_date = datetime.date(year, month, day)
#     if given_date >= last_date_of_month:
#         # Move to the next month
#         if month == 12:
#             next_month = "JAN"
#         else:
#             next_month = list(month_map.keys())[list(month_map.values()).index(month) + 1]
#         return next_month
#     else:
#         return month_abbr
#
#
# d = int(date[:2])
# m = mnth
# y = int(yr)
#
#
# for share in share_list:
#     ltp = df.loc[(df['TckrSymb'] == share) & df['FinInstrmNm'].str.contains(checkExpiry(d, m, y)), 'SttlmPric'].iloc[0]
#     print(f"{share}: {ltp}")
#
#
#
# # BANKNIFTY: 52560.5
# # NIFTY: 23638.9
# # ADANIPORTS: 1482.5
# # AUROPHARMA: 1308.35
# # CANBK: 115.45
# # DLF: 839.6
# # HINDALCO: 698.6
# # ICICIBANK: 1234.9
# # JINDALSTEL: 1031.3
# # RELIANCE: 3208.1
# # SBIN: 860.05
# # TATACONSUM: 1154.4
# # TATAMOTORS: 1006.75
# # TATASTEEL: 172.7
# # TCS: 3996.8
# # TITAN: 3171.35
import re
# //////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
# 1160, 233, 57, 56
# 1157-1161, 1157-1176, 1156-1175

# hl, hl%, cl, cl%, trend

import time
import openpyxl as xl
import pyautogui as pg
from time import sleep

from openpyxl.formatting.rule import CellIsRule
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from selenium import webdriver
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

algo_share_list = ['AARTIIND', '02 ABB', 'ABCAPITAL', 'ABFRL', 'ADANIENT', 'ADANIPORTS', 'ALKEM', 'AMBUJACEM',
                         'APOLLOHOSP', 'APOLLOTYRE', '03 ASHOKLEY', 'ASTRAL', 'ATUL', 'AUBANK', 'AUROPHARMA', 'BAJAJFINSV',
                         'BAJFINANCE', 'BALKRISIND', 'BALRAMCHIN', 'BANDHANBNK', 'BANKBARODA', 'BATAINDIA', 'BEL',
                         'BHARATFORG', '04 BHEL', 'BIOCON', 'BRITANNIA', 'BSOFT', 'CANBK', 'CANFINHOME', 'CHAMBLFERT', 'CHOLAFIN',
                         'CIPLA', 'COFORGE', 'CONCOR', 'COROMANDEL', 'CROMPTON', 'CUMMINSIND', 'DABUR', 'DALBHARAT',
                         'DEEPAKFERT', 'DEEPAKNTR', 'DELTACORP', 'DIVISLAB', '05 DIXON', 'DLF', 'DRREDDY', 'ESCORTS',
                         'EXIDEIND', 'GLENMARK', 'GLS', 'GNFC', 'GODREJCP', 'GODREJPROP', 'GRANULES', 'GRASIM', 'GUJGASLTD',
                         'HAL', 'HAVELLS', 'HCLTECH', 'HDFCAMC', 'HDFCLIFE', 'HINDALCO', 'HINDCOPPER', 'ICICIGI',
                         'ICICIPRULI', 'IEX', 'IGL', 'INDHOTEL', 'INDIACEM', 'INDIAMART', 'INDIGO', 'INDUSINDBK',
                         'INDUSTOWER', 'INTELLECT', 'IPCALAB', 'JINDALSTEL', 'JKCEMENT', 'JSWSTEEL', 'JUBLFOOD',
                         'KOTAKBANK', 'LALPATHLAB', 'LAURUSLABS', 'LICHSGFIN', 'LTIM', 'LTTS', 'LUPIN', 'M%26MFIN',
                         'MANAPPURAM', 'MARICO', 'MCDOWELL-N', 'MCX', 'METROPOLIS', 'MFSL', 'MGL', 'MPHASIS', 'MUTHOOTFIN',
                         'NAM-INDIA', 'NAUKRI', 'NAVINFLUOR', 'NMDC', 'NTPC', 'OBEROIRLTY', '09 ONGC', 'PEL', 'PERSISTENT', 'PETRONET',
                         'PIDILITIND', 'POLYCAB', 'POWERGRID', 'RAIN', 'RAMCOCEM', 'RBLBANK', '10 RECLTD', 'SBICARD',
                         'SBILIFE', 'SIEMENS', 'SRF', 'STAR', 'SUNPHARMA', 'SYNGENE', 'TATACOMM', 'TATAMOTORS', 'TECHM',
                         'TORNTPHARM', 'TORNTPOWER', 'TRENT', 'TVSMOTOR', 'UBL', 'ULTRACEMCO', 'UPL', 'VEDL', 'VOLTAS',
                         'ZEEL', 'ZYDUSLIFE']

daily_cell_range = "H3:K3000"
weekly_cell_range = "E4:I1000"
other_cell_range = "F4:J1000"

font = Font(name='Arial', size=11, bold=True)  # text font
heading_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")  # heading
fill = PatternFill(patternType='solid', fgColor="0000ff")  # blue
red = Font("Arial", 11, color='ff0000', bold=True)
blue = Font("Arial", 11, color="0000ff", bold=True)
alignment = Alignment(horizontal='center')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

red_text_rule = CellIsRule(operator='lessThan', formula=['0'], font=red)
blue_text_rule = CellIsRule(operator='greaterThanOrEqual', formula=['0'], font=blue)

sh_row = 2
d_row = 1101
w_row = 221
m_row = 54
cl_row = 53
w_range = [1097, 1101]
m_range = [1091, 1110]
cl_range = [1091, 1110]
sh_count = 0

xp_high = '/html/body/div[11]/div/div/section/div/div/div/div/div/div[2]/div/div/div/div[4]/div/div[1]/section/div/div[3]/div/table/tbody/tr/td[4]'
xp_low = '/html/body/div[11]/div/div/section/div/div/div/div/div/div[2]/div/div/div/div[4]/div/div[1]/section/div/div[3]/div/table/tbody/tr/td[5]'

options = Options()
# options.add_argument('--headless=new')
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--start-maximized")

# Exclude the collection of enable-automation switches
options.add_experimental_option("excludeSwitches", ["enable-automation"])

# Turn-off userAutomationExtension
options.add_experimental_option("useAutomationExtension", False)

# algo_close_list = [
#                      'CIPLA', 'COFORGE', 'CONCOR', 'COROMANDEL', 'CROMPTON', 'CUMMINSIND', 'DABUR', 'DALBHARAT',
#                      'DEEPAKFERT', 'DEEPAKNTR', 'DELTACORP', 'DIVISLAB', 'DIXON', 'DLF', 'DRREDDY', 'ESCORTS',
#                      'EXIDEIND', 'GLENMARK', 'GLS', 'GNFC', 'GODREJCP', 'GODREJPROP', 'GRANULES', 'GRASIM', 'GUJGASLTD',
#                      'HAL', 'HAVELLS', 'HCLTECH', 'HDFCAMC', 'HDFCLIFE', 'HINDALCO', 'HINDCOPPER', 'ICICIGI',
#                      'ICICIPRULI', 'IEX', 'IGL', 'INDHOTEL', 'INDIACEM', 'INDIAMART', 'INDIGO', 'INDUSINDBK',
#                      'INDUSTOWER', 'INTELLECT', 'IPCALAB', 'JINDALSTEL', 'JKCEMENT', 'JSWSTEEL', 'JUBLFOOD',
#                      'KOTAKBANK', 'LALPATHLAB', 'LAURUSLABS', 'LICHSGFIN', 'LTIM', 'LTTS', 'LUPIN', 'M%26MFIN',
#                      'MANAPPURAM', 'MARICO', 'UNITDSPR', 'MCX', 'METROPOLIS', 'MFSL', 'MGL', 'MPHASIS', 'MUTHOOTFIN',
#                      'NAM-INDIA', 'NAUKRI', 'NAVINFLUOR', 'NMDC', 'NTPC', 'OBEROIRLTY', 'ONGC', 'PEL', 'PERSISTENT', 'PETRONET',
#                      'PIDILITIND', 'POLYCAB', 'POWERGRID', 'RAIN', 'RAMCOCEM', 'RBLBANK', 'RECLTD', 'SBICARD',
#                      'SBILIFE', 'SIEMENS', 'SRF', 'STAR', 'SUNPHARMA', 'SYNGENE', 'TATACOMM', 'TATAMOTORS', 'TECHM',
#                      'TORNTPHARM', 'TORNTPOWER', 'TRENT', 'TVSMOTOR', 'UBL', 'ULTRACEMCO', 'UPL', 'VEDL', 'VOLTAS',
#                      'ZEEL', 'ZYDUSLIFE']
#
#
# doub = ['ADANIPORTS', 'CHOLAFIN']
# add = 0
# sh_row = 34
# for share in algo_close_list:
#     # if share takes 2 line as name
#     if share in doub:
#         add = 20
#     left = (1105, 530+add)
#     ll = (888, 575+add)  # 6
#     ldate = (1052, 701+add)
#     hist_data = (593, 439+add)
#     right = (1387, 532+add)
#     rl = (1172, 573+add)  # 7
#     rdate = (1334, 702+add)
#     flter = (1467, 527+add)
#     hlwb = xl.load_workbook(rf'C:\Users\admin\PycharmProjects\daily data\test\hl_test.xlsx')
#     hlsh = hlwb['Sheet2']
#     driver = webdriver.Chrome(options=options)
#
#     driver.get(f"https://www.nseindia.com/get-quotes/equity?symbol={share}")
#
#
#
#     try:
#         myElem = WebDriverWait(driver, 20).until(ec.presence_of_element_located((By.ID, 'historic_data')))
#         sleep(2)
#         pg.click(hist_data)
#         sleep(2)
#         pg.click(left)
#         sleep(2)
#         pg.click(ll, clicks=6, interval=0.6)
#         sleep(2)
#         pg.click(ldate)
#         sleep(2)
#         pg.click(right)
#         sleep(2)
#         pg.click(rl, clicks=7, interval=0.6)
#         sleep(2)
#         pg.click(rdate)
#         sleep(2)
#         pg.click(flter)
#         sleep(2)
#         he = WebDriverWait(driver, 20).until(ec.presence_of_element_located((By.XPATH, xp_high)))
#         high = float(driver.find_element(By.XPATH, xp_high).text.replace(",", ""))
#         low = float(driver.find_element(By.XPATH, xp_low).text.replace(",", ""))
#
#         hlsh.cell(sh_row, 2).value = high
#         hlsh.cell(sh_row, 3).value = low
#         sh_row += 1
#
#         hlwb.save(rf'C:\Users\admin\PycharmProjects\daily data\test\hl_test.xlsx')
#         print(f"{high} {low} {share}")
#
#     except TimeoutException:
#         print(f"Loading took too much time for {share}!")
#
#     driver.close()


# # copying high and low from algo hl sheet (after putting correct high and low values for that date)
# for share in algo_share_list:
#     hlwb = xl.load_workbook(rf'C:\Users\admin\PycharmProjects\daily data\test\hl_test.xlsx')
#     hlsh = hlwb['Sheet2']
#
#     wb = xl.load_workbook(rf'E:\Daily Data work\ALGORITHM\ALGORITHM OLD\{share}.xlsx')
#     sh = wb['D']
#
#     sh.cell(d_row, 2).value = hlsh.cell(sh_row, 2).value        # high copy
#     sh.cell(d_row, 3).value = hlsh.cell(sh_row, 3).value        # low copy
#
#     wb.save(rf'C:\Users\admin\PycharmProjects\daily data\test\algo\{share}.xlsx')
#     sh_row += 1
#     sh_count += 1
#
#     print(f"{share} done")
#
# print(f"{sh_count} shares done")


for share in algo_share_list:
    wb = xl.load_workbook(rf'E:\Daily Data work\ALGORITHM\ALGORITHM OLD\{share}.xlsx')
    d = wb['D']
    w = wb['W']
    m = wb['M']
    cl = wb['Cl']

    high = 0
    low = 9999999

    # weekly
    for row in range(w_range[0], w_range[1]+1):
        high_cell = d.cell(row, 2)
        low_cell = d.cell(row, 3)

        if high_cell.value is not None and high_cell.value > high:
            high = high_cell.value

        if low_cell.value is not None and low_cell.value < low and low_cell.value != 0:
            low = low_cell.value

    w.cell(w_row, 2).value = high
    w.cell(w_row, 3).value = low

    high = 0
    low = 9999999

    # weekly
    for row in range(m_range[0], m_range[1] + 1):
        high_cell = d.cell(row, 2)
        low_cell = d.cell(row, 3)

        if high_cell.value is not None and high_cell.value > high:
            high = high_cell.value

        if low_cell.value is not None and low_cell.value < low and low_cell.value != 0:
            low = low_cell.value

    m.cell(m_row, 3).value = high
    m.cell(m_row, 4).value = low

    high = 0
    low = 9999999

    # closing
    for row in range(cl_range[0], cl_range[1] + 1):
        high_cell = d.cell(row, 2)
        low_cell = d.cell(row, 3)

        if high_cell.value is not None and high_cell.value > high:
            high = high_cell.value

        if low_cell.value is not None and low_cell.value < low and low_cell.value != 0:
            low = low_cell.value

    cl.cell(cl_row, 3).value = high
    cl.cell(cl_row, 4).value = low

    wb.save(rf'C:\Users\admin\PycharmProjects\daily data\test\algo\{share}.xlsx')
    print(f'{share} done')

# for share in algo_share_list:
#     wb = xl.load_workbook(rf'E:\Daily Data work\ALGORITHM\ALGORITHM OLD\{share}.xlsx')
#     sh = wb['D']
#
#     h = sh.cell(mar_15, 2).value
#     l = sh.cell(mar_15, 3).value
#     c = sh.cell(mar_15, 4).value
#     ltp = sh.cell(mar_15, 5).value
#
#     if h > c > l and h > ltp > l:
#         print(f"{share} good!")
#     else:
#         print(f"{share} bad!")


# for share in algo_share_list:
#     wb1 = xl.load_workbook(rf'E:\Daily Data work\ALGORITHM\{share}.xlsx')
#     d = wb1['D']
#     w = wb1['W']
#     m = wb1['M']
#     cl = wb1['Cl']
#
#     # headings daily
#     d.cell(2, 8).value = "H/L DIFF"
#     d.cell(2, 9).value = "H/L %"
#     d.cell(2, 10).value = "LTP DIFF"
#     d.cell(2, 11).value = "LTP %"
#
#     # headings weekly
#     w.cell(3, 5).value = "H/L DIFF"
#     w.cell(3, 6).value = "H/L %"
#     w.cell(3, 7).value = "LTP DIFF"
#     w.cell(3, 8).value = "LTP %"
#     w.cell(3, 9).value = "TREND"
#
#     # headings monthly
#     m.cell(3, 6).value = "H/L DIFF"
#     m.cell(3, 7).value = "H/L %"
#     m.cell(3, 8).value = "LTP DIFF"
#     m.cell(3, 9).value = "LTP %"
#     m.cell(3, 10).value = "TREND"
#
#     # headings closing
#     cl.cell(3, 6).value = "H/L DIFF"
#     cl.cell(3, 7).value = "H/L %"
#     cl.cell(3, 8).value = "LTP DIFF"
#     cl.cell(3, 9).value = "LTP %"
#     cl.cell(3, 10).value = "TREND"
#
#     # daily pattern filling
#     for row in range(2, 3000):
#         for col in range(8, 12):
#             cell = d.cell(row, col)
#             cell.font = font
#             cell.alignment = alignment
#             cell.number_format = '0.##'
#
#             if row >= 4:
#                 d.cell(row, 8).value = f'=IF(B{row}="", "", B{row}-C{row})'    # hl diff
#                 d.cell(row, 9).value = f'=IF(B{row}="", "",H{row}/E{row}*100)'    # %
#                 d.cell(row, 10).value = f'=IF(B{row}="", "",IF(E{row-1}="", IF(E{row-2}="", E{row}-E{row-3}, E{row}-E{row-2}), E{row}-E{row-1}))'    # ltp diff
#                 d.cell(row, 11).value = f'=IF(B{row}="", "", J{row}*100/(IF(E{row-1}="", IF(E{row-2}="", E{row-3}, E{row-2}), E{row-1})))'    # %
#
#             if row < 3:
#                 cell.fill = fill
#                 cell.font = heading_font
#                 if col > 9:
#                     d.cell(1, col).fill = fill
#
#     d.conditional_formatting.add(daily_cell_range, red_text_rule)
#     d.conditional_formatting.add(daily_cell_range, blue_text_rule)
#
#     # # weekly pattern filling
#     # for row in range(3, 1001):
#     #     for col in range(5, 10):
#     #         cell = w.cell(row, col)
#     #         cell.font = font
#     #         cell.alignment = alignment
#     #         cell.number_format = '0.##'
#     #         cell.border = border
#     #
#     #         if row >= 4:
#     #             w.cell(row, 5).value = f'=B{row}-C{row}'  # hl diff
#     #             w.cell(row, 6).value = f'=E{row}/D{row}*100'  # %
#     #             w.cell(row, 7).value = f'=D{row}-D{row-1}'  # ltp diff
#     #             w.cell(row, 8).value = f'=G{row}*100/D{row-1}'  # %
#     #
#     # w.conditional_formatting.add(weekly_cell_range, red_text_rule)
#     # w.conditional_formatting.add(weekly_cell_range, blue_text_rule)
#     #
#     # # monthly and cl pattern filling
#     # for row in range(3, 1001):
#     #     for col in range(6, 11):
#     #         cell = m.cell(row, col)
#     #         cell.font = font
#     #         cell.alignment = alignment
#     #         cell.number_format = '0.##'
#     #         cell.border = border
#     #
#     #         cell2 = cl.cell(row, col)
#     #         cell2.font = font
#     #         cell2.alignment = alignment
#     #         cell2.number_format = '0.##'
#     #         cell2.border = border
#     #
#     #         if row >= 4:
#     #             m.cell(row, 5).value = f'=C{row}-D{row}'  # hl diff
#     #             m.cell(row, 6).value = f'=E{row}/D{row}*100'  # %
#     #             m.cell(row, 7).value = f'=D{row}-D{row - 1}'  # ltp diff
#     #             m.cell(row, 8).value = f'=G{row}*100/D{row - 1}'  # %
#     #
#     #             cl.cell(row, 5).value = f'=C{row}-D{row}'  # hl diff
#     #             cl.cell(row, 6).value = f'=E{row}/D{row}*100'  # %
#     #             cl.cell(row, 7).value = f'=D{row}-D{row - 1}'  # ltp diff
#     #             cl.cell(row, 8).value = f'=G{row}*100/D{row - 1}'  # %
#     #
#     # m.conditional_formatting.add(other_cell_range, red_text_rule)
#     # m.conditional_formatting.add(other_cell_range, blue_text_rule)
#     # cl.conditional_formatting.add(other_cell_range, red_text_rule)
#     # cl.conditional_formatting.add(other_cell_range, blue_text_rule)
#
#     print(f'{share} done')
#     sh_count += 1
#     wb1.save(rf'C:\Users\admin\PycharmProjects\daily data\test\algo\{share}.xlsx')
#
# print(sh_count)
#
#
# # def apply_borders(ws, cell_range):
# #     for row in ws[cell_range]:
# #         for cell in row:
# #             cell.border = border
# #
# #
# # for share in algo_share_list:
# #     wb1 = xl.load_workbook(rf'E:\Daily Data work\ALGORITHM\{share}.xlsx')
# #     d = wb1['D']
# #     w = wb1['W']
# #     m = wb1['M']
# #     cl = wb1['Cl']
# #
# #     apply_borders(w, weekly_cell_range)
# #     apply_borders(m, other_cell_range)
# #     apply_borders(cl, other_cell_range)
# #
# #     print(f'{share} done')
# #     wb1.save(rf'C:\Users\admin\PycharmProjects\daily data\test\algo\{share}.xlsx')








