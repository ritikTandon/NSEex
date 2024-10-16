import os
import shutil

import openpyxl as xl
import send2trash
from openpyxl.styles import Font, Alignment

append = 307  # increment this daily. 307 is for 16-OCT-2024

# styles
red = Font("Arial", 11, color='ff0000', bold=True)
blue = Font("Arial", 11, color="0000ff", bold=True)
bold = Font("Arial", 11, bold=True)
alignment = Alignment(horizontal='center')


# deleting files of csh, fo1 and algo
def delete_file(path):
    try:
        send2trash.send2trash(path)
    except FileNotFoundError:
        print(f"No file in {path}")


# CASH
def cash():
    i = 0  # main iterator variable

    cash_shares = {'AARTIIND': 947, 'ADANIENT': 1579, 'APOLLOTYRE': 2946, 'BAJAJFINSERV': 1579, 'BAJAJFINANCE': 1579,
                   'BANDHANBANK': 1579, 'BANKBARODA': 1579, 'COAL INDIA': 3232, '06 DLF CHL': 4058, 'EICHERMOTOR': 2715,
                   'FEDRAL BANK': 1579, 'HCLTECH': 1579, 'HDFC': 3936, 'HINDALCO': 947, 'ICICIBANK': 1579, 'INDUSINDBANK': 1579,
                   'INFY': 2765, 'JINDALS chl': 5195, 'LICHSGFIN': 1579, 'M&M': 1579, '07 M&MFINANCE': 1579,
                   '08 NTPC': 947, 'RELIANCE CHL': 4793, 'SBIN CHL': 4860, 'SUNTV': 1579, 'TATACHEM': 1579,
                   '11 TATAMOTOR CHL': 4434, '12 TATAPOWER': 1579, '13 TATASTEEL chl': 4570, 'ULTRACHEM': 2696}

    cash_no_format_list = ['APOLLOTYRE', 'BANDHANBANK', 'BANKBARODA', 'COAL INDIA', 'DLF CHL', 'TATAMOTOR CHL',
                           'TATASTEEL chl', 'TATAPOWER', 'M&MFINANCE', 'FEDRAL BANK', 'HINDALCO', 'NTPC']

    # loading 'cash high low.xlsx'
    cashHL_wb = xl.load_workbook(r'C:\Users\admin\PycharmProjects\daily data\cash high low.xlsx')
    cashHL_sheet = cashHL_wb['Sheet1']
    cashHL_row = 2

    for share in cash_shares:
        path = rf'E:\Daily Data work\CASH\{share}.xlsx'

        wb = xl.load_workbook(path)
        sheet = wb['D']

        input_row = cash_shares[share] + append  # incrementing row values from base row(start row)

        # data filling
        sheet.cell(input_row, 2).value = cashHL_sheet.cell(cashHL_row, 2).value  # high
        sheet.cell(input_row, 3).value = cashHL_sheet.cell(cashHL_row, 3).value  # low
        sheet.cell(input_row, 4).value = cashHL_sheet.cell(cashHL_row, 4).value  # close
        sheet.cell(input_row, 5).value = cashHL_sheet.cell(cashHL_row, 5).value  # LTP
        sheet.cell(input_row, 6).value = cashHL_sheet.cell(cashHL_row, 6).value  # vol
        sheet.cell(input_row, 7).value = cashHL_sheet.cell(cashHL_row, 7).value  # 9:25 close

        # number formatting
        if share not in cash_no_format_list:
            sheet.cell(input_row, 2).number_format = '0'  # high
            sheet.cell(input_row, 3).number_format = '0'  # low
            sheet.cell(input_row, 4).number_format = '0'  # close
            sheet.cell(input_row, 5).number_format = '0'  # LTP
            sheet.cell(input_row, 7).number_format = '0'  # 9:25 close

        # style formatting
        sheet.cell(input_row, 2).font = blue  # high
        sheet.cell(input_row, 2).alignment = alignment

        sheet.cell(input_row, 3).font = red  # low
        sheet.cell(input_row, 3).alignment = alignment

        sheet.cell(input_row, 4).font = bold  # close
        sheet.cell(input_row, 4).alignment = alignment

        sheet.cell(input_row, 5).font = bold  # LTP
        sheet.cell(input_row, 5).alignment = alignment

        sheet.cell(input_row, 6).font = bold  # vol
        sheet.cell(input_row, 6).alignment = alignment

        sheet.cell(input_row, 7).font = bold  # 9:25 close
        sheet.cell(input_row, 7).alignment = alignment

        i += 1
        cashHL_row += 1

        print(f"{share} done!")

        wb.save(path)

    print("----------------------------------CASH DONE----------------------------------")


# FO
def fo():
    i = 0  # main iterator variable

    # nifty and bn new data won't be updated in the future sheets but in FUT sheets in cash wb from now on (12/10/24)
    # changed nifty row to reflect the correct row in nifty cash "FUT" sheet
    fo_shares = {'ADANI PORT': 2279, 'AUROPHARMA': 2789, '02 BANKNIFTY F': 3309, 'CANBK': 2016, 'DLF': 2831,
                 'HINDALCO': 3989, 'ICICIBANK': 1093, 'JINDS': 2274, '01 NIFTY F': 5390, '03 RELIANCE': 2792, 'SBIN': 2793,
                 'TATACONSUM': 2276, '05 TATAMOTOR': 2791, '04 TATASTEEL': 2793, 'TCS': 4826, 'TITAN': 1762}

    fo_no_format_list = ['04 TATASTEEL']

    # loading 'fo high low.xlsx'
    foHL_wb = xl.load_workbook(r'C:\Users\admin\PycharmProjects\daily data\fo high low.xlsx')
    foHL_sheet = foHL_wb['Sheet1']
    foHL_row = 2

    for share in fo_shares:
        if share == '02 BANKNIFTY F':
            path = rf'E:\Daily Data work\CASH\BANKNIFTY C.xlsx'         # MAKE A LIST WHERE CHANGING A SHEET NAME AFFECTS THE CODE
            wb = xl.load_workbook(path)
            sheet = wb['FUT']
        elif share == '01 NIFTY F':
            path = rf'E:\Daily Data work\CASH\01 NIFTY CHL C.xlsx'
            wb = xl.load_workbook(path)
            sheet = wb['FUT']
        else:
            path = rf'E:\Daily Data work\FO\{share}.xlsx'

            wb = xl.load_workbook(path)
            sheet = wb['D']

        input_row = fo_shares[share] + append

        # data filling
        sheet.cell(input_row, 2).value = foHL_sheet.cell(foHL_row, 2).value  # high
        sheet.cell(input_row, 3).value = foHL_sheet.cell(foHL_row, 3).value  # low
        sheet.cell(input_row, 4).value = foHL_sheet.cell(foHL_row, 4).value  # close
        sheet.cell(input_row, 5).value = foHL_sheet.cell(foHL_row, 5).value  # LTP
        sheet.cell(input_row, 6).value = foHL_sheet.cell(foHL_row, 6).value  # vol
        sheet.cell(input_row, 7).value = foHL_sheet.cell(foHL_row, 7).value  # 9:25 close

        # number formatting
        if share not in fo_no_format_list:
            sheet.cell(input_row, 2).number_format = '0'  # high
            sheet.cell(input_row, 3).number_format = '0'  # low
            sheet.cell(input_row, 4).number_format = '0'  # close
            sheet.cell(input_row, 5).number_format = '0'  # LTP
            sheet.cell(input_row, 7).number_format = '0'  # 9:25 close

        # style formatting
        sheet.cell(input_row, 2).font = blue  # high
        sheet.cell(input_row, 2).alignment = alignment

        sheet.cell(input_row, 3).font = red  # low
        sheet.cell(input_row, 3).alignment = alignment

        sheet.cell(input_row, 4).font = bold  # close
        sheet.cell(input_row, 4).alignment = alignment

        sheet.cell(input_row, 5).font = bold  # LTP
        sheet.cell(input_row, 5).alignment = alignment

        sheet.cell(input_row, 6).font = bold  # vol
        sheet.cell(input_row, 6).alignment = alignment

        sheet.cell(input_row, 7).font = bold  # 9:25 close
        sheet.cell(input_row, 7).alignment = alignment

        i += 1
        foHL_row += 1

        print(f"{share} done!")

        wb.save(path)

    print("----------------------------------FO DONE----------------------------------")


# algo
def algo():
    i = 0  # main iterator variable
    algo_row = 947

    algo_copy_to_cash_list = ['02 ABB', '03 ASHOKLEY', '04 BHEL', '05 DIXON', '09 ONGC', '10 RECLTD']

    algo_share_list = ['AARTIIND', '02 ABB', 'ABCAPITAL', 'ABFRL', 'ADANIENT', 'ADANIPORTS', 'ALKEM', 'AMBUJACEM',
                         'APOLLOHOSP', 'APOLLOTYRE', '03 ASHOKLEY', 'ASTRAL', 'ATUL', 'AUBANK', 'AUROPHARMA', 'BAJAJFINSV',
                         'BAJFINANCE', 'BALKRISIND', 'BALRAMCHIN', 'BANDHANBNK', 'BANKBARODA', 'BATAINDIA', 'BEL',
                         'BHARATFORG', '04 BHEL', 'BIOCON', 'BRITANNIA', 'BSOFT', 'CANBK', 'CANFINHOME', 'CHAMBLFERT', 'CHOLAFIN',
                         'CIPLA', 'COFORGE', 'CONCOR', 'COROMANDEL', 'CROMPTON', 'CUMMINSIND', 'DABUR', 'DALBHARAT',
                         'DEEPAKFERT', 'DEEPAKNTR', 'DELTACORP', 'DIVISLAB', '05 DIXON', 'DLF', 'DRREDDY', 'ESCORTS',
                         'EXIDEIND', 'GLENMARK', 'GLS', 'GNFC', 'GODREJCP', 'GODREJPROP', 'GRANULES', 'GRASIM', 'GUJGASLTD',
                         'HAL', 'HAVELLS', 'HCLTECH', 'HDFCAMC', 'HDFCLIFE', 'HINDALCO', 'HINDCOPPER', 'ICICIGI',
                         'ICICIPRULI', 'IEX', 'IGL', 'INDHOTEL', 'INDIACEM', 'INDIAMART', 'INDIGO', 'INDUSINDBK',
                         'INDUSTOWER', 'INTELLECT', 'IPCALAB', 'JINDALSTEL', 'JKCEMENT', 'JSWSTEEL', 'JUBLFOOD',
                         'KOTAKBANK', 'LALPATHLAB', 'LAURUSLABS', 'LICHSGFIN', 'LTIM', 'LTTS', 'LUPIN', 'M%26MFIN',
                         'MANAPPURAM', 'MARICO', 'MCDOWELL-N', 'MCX', 'METROPOLIS', 'MFSL', 'MGL', 'MPHASIS', 'MUTHOOTFIN',
                         'NAM-INDIA', 'NAUKRI', 'NAVINFLUOR', 'NMDC', 'NTPC', 'OBEROIRLTY', '09 ONGC', 'PEL', 'PERSISTENT', 'PETRONET',
                         'PIDILITIND', 'POLYCAB', 'POWERGRID', 'RAIN', 'RAMCOCEM', 'RBLBANK', '10 RECLTD', 'SBICARD',
                         'SBILIFE', 'SIEMENS', 'SRF', 'STAR', 'SUNPHARMA', 'SYNGENE', 'TATACOMM', 'TATAMOTORS', 'TECHM',
                         'TORNTPHARM', 'TORNTPOWER', 'TRENT', 'TVSMOTOR', 'UBL', 'ULTRACEMCO', 'UPL', 'VEDL', 'VOLTAS',
                         'ZEEL', 'ZYDUSLIFE']

    algo_copy_list = ['AARTIIND', '02 ABB', 'ABFRL', 'ADANIENT', 'ADANIPORTS', 'AMBUJACEM','APOLLOHOSP', 'APOLLOTYRE',
                      '03 ASHOKLEY', 'AUROPHARMA', 'BAJAJFINSV','BAJFINANCE', 'BALKRISIND', 'BALRAMCHIN', 'BANDHANBNK',
                      'BANKBARODA', 'BEL','BHARATFORG', '04 BHEL', 'BIOCON', 'BRITANNIA', 'BSOFT', 'CANBK', 'CANFINHOME',
                      'CHAMBLFERT', 'CHOLAFIN', 'CIPLA', 'COFORGE', 'CROMPTON', 'CUMMINSIND', 'DIVISLAB', '05 DIXON',
                      'DLF', 'DRREDDY', 'ESCORTS','EXIDEIND', 'GLENMARK', 'GNFC', 'GODREJPROP', 'HAL', 'HAVELLS',
                      'HCLTECH', 'HDFCAMC', 'HDFCLIFE', 'HINDALCO', 'HINDCOPPER', 'ICICIGI', 'ICICIPRULI', 'IEX', 'IGL',
                      'INDIACEM', 'INDIGO', 'INDUSINDBK', 'INDUSTOWER', 'INTELLECT', 'IPCALAB', 'KOTAKBANK', 'LALPATHLAB',
                      'LAURUSLABS', 'LICHSGFIN', 'LTIM', 'LTTS', 'LUPIN', 'M%26MFIN', 'MANAPPURAM', 'MCDOWELL-N', 'MCX',
                      'METROPOLIS', 'MFSL', 'MGL', 'MPHASIS', 'MUTHOOTFIN', 'NAM-INDIA', 'NAUKRI', 'NMDC', 'NTPC',
                      '09 ONGC', 'PEL', 'PERSISTENT', 'PETRONET', 'POLYCAB', 'POWERGRID', 'RBLBANK', '10 RECLTD',
                      'SBICARD', 'SIEMENS', 'SUNPHARMA', 'TATAMOTORS', 'TECHM', 'TRENT', 'TVSMOTOR', 'UBL', 'ULTRACEMCO',
                      'VEDL', 'VOLTAS']

    algo_no_format_list = []

    # loading 'algo high low.xlsx'
    algoHL_wb = xl.load_workbook(r'C:\Users\admin\PycharmProjects\daily data\algo high low.xlsx')
    algoHL_sheet = algoHL_wb['Sheet1']
    algoHL_row = 2

    for share in algo_share_list:
        path = rf'E:\Daily Data work\ALGORITHM\ALGORITHM OLD\{share}.xlsx'

        wb = xl.load_workbook(path)
        sheet = wb['D']

        input_row = algo_row + append

        # data filling
        sheet.cell(input_row, 2).value = algoHL_sheet.cell(algoHL_row, 2).value  # high
        sheet.cell(input_row, 3).value = algoHL_sheet.cell(algoHL_row, 3).value  # low
        sheet.cell(input_row, 4).value = algoHL_sheet.cell(algoHL_row, 4).value  # close
        sheet.cell(input_row, 5).value = algoHL_sheet.cell(algoHL_row, 5).value  # LTP
        sheet.cell(input_row, 6).value = algoHL_sheet.cell(algoHL_row, 6).value  # vol
        sheet.cell(input_row, 7).value = algoHL_sheet.cell(algoHL_row, 7).value  # 9:25 close

        # number formatting
        if share not in algo_no_format_list:
            sheet.cell(input_row, 2).number_format = '0'  # high
            sheet.cell(input_row, 3).number_format = '0'  # low
            sheet.cell(input_row, 4).number_format = '0'  # close
            sheet.cell(input_row, 5).number_format = '0'  # LTP
            sheet.cell(input_row, 7).number_format = '0'  # 9:25 close

        # style formatting
        sheet.cell(input_row, 2).font = blue  # high
        sheet.cell(input_row, 2).alignment = alignment

        sheet.cell(input_row, 3).font = red  # low
        sheet.cell(input_row, 3).alignment = alignment

        sheet.cell(input_row, 4).font = bold  # close
        sheet.cell(input_row, 4).alignment = alignment

        sheet.cell(input_row, 5).font = bold  # LTP
        sheet.cell(input_row, 5).alignment = alignment

        sheet.cell(input_row, 6).font = bold  # vol
        sheet.cell(input_row, 6).alignment = alignment

        sheet.cell(input_row, 7).font = bold  # 9:25 close
        sheet.cell(input_row, 7).alignment = alignment

        i += 1
        algoHL_row += 1

        print(f"{share} done!")

        if share in algo_copy_to_cash_list:
            wb.save(rf'E:\Daily Data work\CASH\{share}.xlsx')

        # copying relevant shares to ALGO (new) folder
        if share in algo_copy_list:
            wb.save(rf'E:\Daily Data work\ALGORITHM\{share}.xlsx')

        wb.save(path)

    print("----------------------------------ALGO DONE----------------------------------")


# easier to make these into functions as I can easily run separately during testing
cash()
fo()
algo()

# finally deleting files
delete_file(r'E:\Daily Data work\csh.xls')
delete_file(r'E:\Daily Data work\fo1.xls')
delete_file(r'E:\Daily Data work\algo.xls')

# path = r'D:\Transfer'
#
# try:
#     if os.path.exists(path):
#         shutil.copy()
#     else:
#         print("Path doesn't exist!")
# except FileNotFoundError:
#     print("Pendrive not inserted!")